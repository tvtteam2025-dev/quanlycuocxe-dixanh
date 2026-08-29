import csv
import base64
import copy
import json
import math
import os
import re
import secrets
import threading
import time
import unicodedata
from datetime import datetime, timedelta, timezone
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any
from urllib.parse import quote
from urllib.request import Request as UrlRequest, urlopen

from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, Response
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field

try:
    import gspread
except ImportError:  # pragma: no cover
    gspread = None

load_dotenv()

BASE_DIR = Path(__file__).resolve().parent
PUBLIC_DIR = BASE_DIR / "public"

SHEET_ID = os.getenv("GOOGLE_SHEET_ID", "")
LEGACY_SHEET_ID = os.getenv("LEGACY_GOOGLE_SHEET_ID", "")
LEGACY_READ_ENABLED = os.getenv("LEGACY_READ_ENABLED", "true").strip().lower() in {"1", "true", "yes", "on"}
SERVICE_ACCOUNT_FILE = os.getenv("GOOGLE_SERVICE_ACCOUNT_FILE", "service-account.json")
DEMO_PUBLIC_USERNAME = os.getenv("DEMO_PUBLIC_USERNAME", "demo")
DEMO_PUBLIC_PASSWORD = os.getenv("DEMO_PUBLIC_PASSWORD", "")
DEFAULT_ADMIN_USERNAME = os.getenv("DEFAULT_ADMIN_USERNAME", "admin")
DEFAULT_ADMIN_PASSWORD = os.getenv("DEFAULT_ADMIN_PASSWORD", "")
_GSPREAD_CLIENT: Any | None = None
_SPREADSHEET: Any | None = None
_LEGACY_SPREADSHEET: Any | None = None
_WORKSHEET_CACHE: dict[str, Any] = {}
_HEADER_CHECKED_SHEETS: set[str] = set()
SHEET_VALUES_CACHE_TTL_SECONDS = max(30, int(os.getenv("SHEET_VALUES_CACHE_TTL_SECONDS", "300")))
LEGACY_SHEET_VALUES_CACHE_TTL_SECONDS = max(
    300,
    int(os.getenv("LEGACY_SHEET_VALUES_CACHE_TTL_SECONDS", "86400")),
)
_SHEET_VALUES_CACHE: dict[str, dict[str, Any]] = {}
_SHEET_VALUES_CACHE_LOCK = threading.RLock()
_SHEET_REFRESH_LOCKS: dict[str, threading.Lock] = {}
_WORKSHEET_WRITE_LOCKS: dict[str, threading.Lock] = {}
_WORKSHEET_WRITE_LOCKS_LOCK = threading.RLock()
_LEGACY_VALUES_CACHE: dict[str, dict[str, Any]] = {}
_LEGACY_VALUES_CACHE_LOCK = threading.RLock()
_LEGACY_REFRESH_LOCKS: dict[str, threading.Lock] = {}
SESSIONS: dict[str, dict[str, Any]] = {}
SESSION_IDLE_TIMEOUT_SECONDS = max(300, int(os.getenv("SESSION_IDLE_TIMEOUT_SECONDS", str(8 * 60 * 60))))
USER_CACHE_TTL_SECONDS = 30
_USER_CACHE: dict[str, Any] = {"expires": datetime.min.replace(tzinfo=timezone.utc), "rows": []}

ROSTER_SHEET_NAME = os.getenv("ROSTER_SHEET_NAME", "DANH_SACH_LEN_CA")
FRANCHISE_VEHICLES_SHEET_NAME = os.getenv("FRANCHISE_VEHICLES_SHEET_NAME", "XE_THUONG_QUYEN_HOP_TAC")
CUSTOMERS_SHEET_NAME = os.getenv("CUSTOMERS_SHEET_NAME", "KHACH_HANG")
TOURS_SHEET_NAME = os.getenv("TOURS_SHEET_NAME", "HOP_DONG_TOUR")
ORDERS_SHEET_NAME = os.getenv("ORDERS_SHEET_NAME", "DON_HANG")
SHARED_RIDE_SHEET_NAME = os.getenv("SHARED_RIDE_SHEET_NAME", "KHACH_XE_GHEP")
VOUCHERS_SHEET_NAME = os.getenv("VOUCHERS_SHEET_NAME", "VOUCHER")
PROMOTIONS_SHEET_NAME = os.getenv("PROMOTIONS_SHEET_NAME", "CHUONG_TRINH_KHUYEN_MAI")
ORDER_BENEFITS_SHEET_NAME = os.getenv("ORDER_BENEFITS_SHEET_NAME", "DON_HANG_UU_DAI")
USERS_SHEET_NAME = os.getenv("USERS_SHEET_NAME", "USERS")
REOPEN_REQUESTS_SHEET_NAME = os.getenv("REOPEN_REQUESTS_SHEET_NAME", "YEU_CAU_MO_LAI_DON")
SYSTEM_LOGS_SHEET_NAME = os.getenv("SYSTEM_LOGS_SHEET_NAME", "NHAT_KY_HE_THONG")
ORDER_FEEDBACK_SHEET_NAME = os.getenv("ORDER_FEEDBACK_SHEET_NAME", "PHAN_HOI_KHACH_HANG")
INVOICE_GROUPS_SHEET_NAME = os.getenv("INVOICE_GROUPS_SHEET_NAME", "NHOM_HOA_DON")
SYSTEM_CATALOGS_SHEET_NAME = os.getenv("SYSTEM_CATALOGS_SHEET_NAME", "DANH_MUC_HE_THONG")
CSKH_SHIFT_REPORTS_SHEET_NAME = os.getenv("CSKH_SHIFT_REPORTS_SHEET_NAME", "BAO_CAO_CA_CSKH")
CALENDAR_VEHICLE_ORDER_SHEET_NAME = os.getenv("CALENDAR_VEHICLE_ORDER_SHEET_NAME", "THU_TU_LICH_DIEU_XE")
CONTRACT_PRICING_SHEET_NAME = os.getenv("CONTRACT_PRICING_SHEET_NAME", "BANG_GIA_HOP_DONG")

ROLE_LABELS = {
    "admin": "Admin",
    "ke_toan": "Kế toán",
    "cskh": "CSKH",
    "marketing": "Marketing",
    "ban_giam_doc": "Ban Giám đốc",
    "kinh_doanh": "Kinh Doanh",
}

ROLE_PERMISSIONS = {
    "admin": {
        "views": [
            "dashboard",
            "customers",
            "contracts",
            "contractPricing",
            "overnightCalculator",
            "vouchers",
            "promotions",
            "orders",
            "reports",
            "calendar",
            "vehicles",
            "franchiseVehicles",
            "invoiceOrders",
            "debtOrders",
            "orderFeedback",
            "commissionOrders",
            "reopenApprovals",
            "permissions",
            "systemCatalogs",
            "cskhShiftReports",
        ],
        "actions": [
            "manage_users",
            "approve_reopen",
            "reports_all",
            "reports_revenue",
            "reports_b2c",
            "export_excel",
            "export_b2b",
            "export_b2c",
            "manage_benefits",
            "manage_invoices",
            "create_invoice_groups",
            "export_invoices",
            "export_debts",
            "manage_debts",
            "manage_order_feedback",
            "manage_commissions",
            "export_commissions",
            "manage_remittance_status",
            "manage_system_catalogs",
        ],
    },
    "ke_toan": {
        "views": ["orders", "invoiceOrders", "debtOrders", "commissionOrders", "reports", "overnightCalculator"],
        "actions": ["manage_invoices", "create_invoice_groups", "export_invoices", "export_excel", "export_debts", "manage_debts", "manage_commissions", "export_commissions", "manage_remittance_status"],
    },
    "cskh": {
        "views": [
            "dashboard",
            "customers",
            "contracts",
            "contractPricing",
            "overnightCalculator",
            "orders",
            "reports",
            "calendar",
            "vehicles",
            "franchiseVehicles",
            "invoiceOrders",
            "orderFeedback",
            "cskhShiftReports",
        ],
        "actions": [
            "edit_data",
            "dispatch",
            "complete_order",
            "request_reopen",
            "view_invoices",
            "create_invoice_groups",
            "export_orders",
            "export_customers",
            "manage_order_feedback",
        ],
    },
    "marketing": {
        "views": ["customers", "vouchers", "promotions", "reports", "cskhShiftReports", "overnightCalculator"],
        "actions": ["manage_benefits", "export_customers", "export_vouchers", "export_orders", "export_cskh_shift_reports"],
    },
    "ban_giam_doc": {
        "views": [
            "dashboard",
            "customers",
            "contracts",
            "contractPricing",
            "overnightCalculator",
            "vouchers",
            "promotions",
            "orders",
            "reports",
            "calendar",
            "vehicles",
            "franchiseVehicles",
            "invoiceOrders",
        ],
        "actions": ["reports_all", "export_excel"],
    },
    "kinh_doanh": {
        "views": ["dashboard", "customers", "orders", "reports", "overnightCalculator"],
        "actions": ["reports_all", "export_excel", "export_b2b", "export_b2c"],
    },
}

CUSTOMER_HEADERS = [
    "id",
    "tenKhach",
    "soDienThoai",
    "soCCCD",
    "diaChi",
    "loaiKhachHang",
    "namSinh",
    "gioiTinh",
    "nguonKhach",
    "nhanVienNhap",
    "createdAt",
    "deletedAt",
    "deletedBy",
]

TOUR_HEADERS = [
    "id",
    "diemDi",
    "diemDen",
    "tuyen",
    "ghiChu",
    "createdAt",
    "deletedAt",
    "deletedBy",
]

ORDER_HEADERS = [
    "id",
    "khachHangId",
    "tenKhach",
    "soDienThoai",
    "hopDongTourId",
    "tuyen",
    "diemDon",
    "diemTra",
    "giaTien",
    "bienKiemSoat",
    "soHieuXe",
    "hoTenLaiXe",
    "maNVLaiXe",
    "ngayGioDi",
    "ngayGioDuKienKetThuc",
    "loaiHopDong",
    "soVe",
    "loaiKhach",
    "yeuCauHoaDon",
    "tenCongTy",
    "maSoThue",
    "diaChiHoaDon",
    "emailHoaDon",
    "ghiChu",
    "trangThai",
    "ngayGioHoanThanh",
    "createdAt",
    "giamGia",
    "daCoc",
    "thucThu",
    "loaiXeDieuDong",
    "tyLeNopLai",
    "soTienNopLai",
    "voucherCodes",
    "khuyenMai",
    "tongUuDai",
    "khuVucDatXe",
    "deletedAt",
    "deletedBy",
    "trangThaiHoaDon",
    "ngayXuatHoaDon",
    "nguoiXuatHoaDon",
    "ghiChuGiamGia",
    "congNo",
    "congNoChoAi",
    "thueVAT",
    "tongThanhToan",
    "trangThaiCongNo",
    "ngayThuHoiCongNo",
    "nguoiThuHoiCongNo",
    "phuThu",
    "lyDoPhuThu",
    "trangThaiHoaHong",
    "ngayThuHoaHong",
    "nguoiThuHoaHong",
    "nhomHoaDonId",
    "soCho",
    "trangThaiGuiTaiXe",
    "trangThaiNopTien",
    "ngayXacNhanNopTien",
    "nguoiXacNhanNopTien",
    "nguoiTaoDon",
]

FRANCHISE_VEHICLE_HEADERS = [
    "id",
    "bienKiemSoat",
    "dongXe",
    "hieuXe",
    "soCho",
    "tenChuXe",
    "soDienThoaiChuXe",
    "hoTenLaiXe",
    "soDienThoaiLaiXe",
    "diaChiLaiXe",
    "trangThai",
    "ghiChu",
    "createdAt",
    "deletedAt",
    "deletedBy",
]

SHARED_RIDE_HEADERS = [
    "id",
    "donHangId",
    "hopDongTourId",
    "tuyen",
    "bienKiemSoat",
    "ngayGioDi",
    "ngayGioDuKienKetThuc",
    "hoTen",
    "soDienThoai",
    "soCCCD",
    "diaChi",
    "gioiTinh",
    "namSinh",
    "nguonKhach",
    "nhanVienNhap",
    "diemDon",
    "diemTra",
    "soTien",
    "giamGia",
    "daCoc",
    "thucThu",
    "voucherIds",
    "voucherCodes",
    "promotionIds",
    "khuyenMai",
    "tongUuDai",
    "yeuCauHoaDon",
    "tenCongTy",
    "maSoThue",
    "diaChiHoaDon",
    "emailHoaDon",
    "createdAt",
    "trangThaiHoaDon",
    "ngayXuatHoaDon",
    "nguoiXuatHoaDon",
    "ghiChuGiamGia",
    "thueVAT",
    "tongThanhToan",
    "congNo",
    "congNoChoAi",
    "trangThaiCongNo",
    "ngayThuHoiCongNo",
    "nguoiThuHoiCongNo",
    "phuThu",
    "lyDoPhuThu",
    "loaiKhach",
]

VOUCHER_HEADERS = [
    "id",
    "maVoucher",
    "tenVoucher",
    "loaiGiaTri",
    "giaTri",
    "ngayBatDau",
    "ngayHetHan",
    "trangThai",
    "ghiChu",
    "createdAt",
    "deletedAt",
    "deletedBy",
]

PROMOTION_HEADERS = [
    "id",
    "tenChuongTrinh",
    "loaiGiaTri",
    "giaTri",
    "ngayBatDau",
    "ngayHetHan",
    "trangThai",
    "ghiChu",
    "createdAt",
    "deletedAt",
    "deletedBy",
]

ORDER_BENEFIT_HEADERS = [
    "id",
    "donHangId",
    "khachHangId",
    "tenKhach",
    "loaiUuDai",
    "uuDaiId",
    "maUuDai",
    "tenUuDai",
    "loaiGiaTri",
    "giaTri",
    "soTienGiam",
    "createdAt",
]

USER_HEADERS = [
    "id",
    "username",
    "password",
    "displayName",
    "role",
    "status",
    "createdAt",
    "createdBy",
    "extraPermissions",
]

SYSTEM_CATALOG_HEADERS = [
    "id", "loaiDanhMuc", "giaTri", "thuTu", "trangThai", "createdAt", "createdBy", "deletedAt", "deletedBy",
]

CSKH_SHIFT_REPORT_HEADERS = [
    "Ngày",
    "Nhân Viên Trực",
    "Ca Làm Việc",
    "Thời Gian",
    "Số lượng tin nhắn meta",
    "Số lượng khách phản hồi",
    "Số lượng cuộc gọi",
    "Số lượng chat zalo",
    "Số lượng khách từ website",
    "Số lượng khách từ Email",
    "Số lượng tin nhắn khách vãng lai",
    "Số lượng khách phản hồi từ tiktok",
    "Số lượng đơn chốt từ tiktok",
    "Tổng số lượng đơn chốt",
    "Trạng Thái",
    "Ngày Xóa",
    "Nhân Viên Xóa",
]

CALENDAR_VEHICLE_ORDER_HEADERS = ["bienKiemSoat", "thuTu", "updatedAt", "updatedBy"]
CONTRACT_PRICING_HEADERS = ["id", "configJson", "updatedAt", "updatedBy"]

DEFAULT_CONTRACT_PRICING = {
    "oneWay": [
        {"minKm": 1, "maxKm": 20, "rates": {"4": 18000, "7": 20000, "16": 20000}},
        {"minKm": 21, "maxKm": 50, "rates": {"4": 12000, "7": 14000, "16": 20000}},
        {"minKm": 51, "maxKm": 80, "rates": {"4": 11000, "7": 12000, "16": 20000}},
        {"minKm": 81, "maxKm": 100, "rates": {"4": 9000, "7": 10000, "16": 20000}},
        {"minKm": 101, "maxKm": 120, "rates": {"4": 8000, "7": 9000, "16": 18000}},
        {"minKm": 121, "maxKm": None, "rates": {"4": 7500, "7": 8500, "16": 18000}},
    ],
    "roundTrip": [
        {"minKm": 1, "maxKm": 19, "percentages": {"4": 100, "7": 100, "16": 100}},
        {"minKm": 20, "maxKm": 30, "percentages": {"4": 50, "7": 50, "16": 50}},
        {"minKm": 31, "maxKm": 70, "percentages": {"4": 40, "7": 40, "16": 40}},
        {"minKm": 71, "maxKm": 150, "percentages": {"4": 30, "7": 30, "16": 30}},
        {"minKm": 151, "maxKm": 200, "percentages": {"4": 20, "7": 20, "16": 20}},
        {"minKm": 201, "maxKm": None, "percentages": {"4": 10, "7": 10, "16": 10}},
    ],
    "waiting": [
        {"minAmount": 0, "maxAmount": 500000, "minutes": 30},
        {"minAmount": 500001, "maxAmount": 1000000, "minutes": 60},
        {"minAmount": 1000001, "maxAmount": 2000000, "minutes": 120},
        {"minAmount": 2000001, "maxAmount": None, "minutes": 180},
    ],
    "rounding": {"threshold": 1000000, "belowStep": 10000, "fromStep": 100000},
}

SYSTEM_CATALOG_TYPES = {
    "nguonKhach": "Nguồn khách",
    "dongXe": "Dòng xe",
    "hieuXe": "Hiệu xe",
}

DEFAULT_SYSTEM_CATALOGS = {
    "nguonKhach": ["Facebook Ads", "Tổng đài", "Tiktok", "Lái xe giới thiệu", "Khách cũ", "Khác"],
    "dongXe": ["Xe điện", "Xe xăng", "Xe dầu", "Xe hybrid"],
    "hieuXe": [
        "Minio Green", "Herio Green", "Nerio Green", "Limo Green",
        "VF 3", "VF 5", "VF 6", "VF 7", "VF 8", "VF MPV 7",
        "Vios", "Innova", "Innova Cross", "Veloz Cross", "Avanza Premio", "Fortuner",
        "Attrage", "Xpander", "Xpander Cross", "Xforce",
        "Grand i10", "Accent", "Stargazer", "Custin", "Solati",
        "K3", "Carens", "Carnival",
        "City", "BR-V", "XL7 Hybrid", "Ertiga Hybrid",
        "Mazda2", "Mazda3", "CX-5", "Transit", "Everest",
    ],
}

REOPEN_REQUEST_HEADERS = [
    "id",
    "orderId",
    "orderCode",
    "requestedBy",
    "requestedRole",
    "reason",
    "status",
    "adminNote",
    "createdAt",
    "reviewedAt",
    "reviewedBy",
]

SYSTEM_LOG_HEADERS = [
    "id",
    "createdAt",
    "username",
    "role",
    "action",
    "targetType",
    "targetId",
    "note",
    "before",
    "after",
]

ORDER_FEEDBACK_HEADERS = [
    "id",
    "donHangId",
    "maDon",
    "khachHangId",
    "tenKhach",
    "kenhChamSoc",
    "diemDanhGia",
    "noiDungPhanHoi",
    "hinhThucXuLy",
    "ketQuaXuLy",
    "chuThich",
    "createdAt",
    "createdBy",
    "updatedAt",
    "updatedBy",
]

INVOICE_GROUP_HEADERS = [
    "id", "khachHangId", "tenKhach", "soDienThoai", "orderIds", "tenCongTy", "maSoThue",
    "diaChiHoaDon", "emailHoaDon", "tongTruocVAT", "tongVAT", "tongThanhToan",
    "trangThai", "createdAt", "createdBy", "ngayXuatHoaDon", "nguoiXuatHoaDon",
]

app = FastAPI(title="Di Xanh Order Manager")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


def is_local_request(host: str) -> bool:
    hostname = host.split(":", 1)[0].strip().lower()
    return hostname in {"127.0.0.1", "localhost", "::1"}


def unauthorized_response() -> Response:
    return JSONResponse({"detail": "Cần đăng nhập để sử dụng hệ thống."}, status_code=401)


def demo_admin_user() -> dict[str, Any]:
    return {
        "id": "default-admin",
        "username": DEFAULT_ADMIN_USERNAME,
        "displayName": "Admin",
        "role": "admin",
        "status": "active",
    }


def public_user(row: dict[str, Any]) -> dict[str, Any]:
    return {key: value for key, value in row.items() if key != "password"}


def default_admin_row() -> dict[str, Any]:
    return {
        "id": "USER-DEFAULT-ADMIN",
        "username": DEFAULT_ADMIN_USERNAME,
        "password": DEFAULT_ADMIN_PASSWORD,
        "displayName": "Admin",
        "role": "admin",
        "status": "active",
        "createdAt": now_iso(),
        "createdBy": "system",
        "extraPermissions": "manage_system_catalogs",
    }


def permission_list(value: Any) -> list[str]:
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    return [item.strip() for item in str(value or "").split(",") if item.strip()]


def compare_secret(left: Any, right: Any) -> bool:
    """Compare credentials safely while supporting Vietnamese/Unicode text."""
    return secrets.compare_digest(
        str(left or "").encode("utf-8"),
        str(right or "").encode("utf-8"),
    )


def user_permissions(role: str, extra_permissions: Any = "") -> dict[str, list[str]]:
    base = ROLE_PERMISSIONS.get(role, ROLE_PERMISSIONS["ban_giam_doc"])
    views = list(base.get("views", []))
    actions = list(base.get("actions", []))
    for action in permission_list(extra_permissions):
        if action not in actions:
            actions.append(action)
    if "manage_system_catalogs" in actions and "systemCatalogs" not in views:
        views.append("systemCatalogs")
    return {"views": views, "actions": actions}


def json_forbidden_response(message: str = "Bạn không có quyền thao tác chức năng này.") -> Response:
    return Response(
        content=json.dumps({"detail": message}, ensure_ascii=False),
        status_code=403,
        media_type="application/json",
    )


def is_google_quota_error(exc: Exception) -> bool:
    message = str(exc)
    return "Quota exceeded" in message or "[429]" in message


def is_google_transient_error(exc: Exception) -> bool:
    message = str(exc).lower()
    transient_markers = (
        "timed out",
        "timeout",
        "connection reset",
        "connection aborted",
        "temporarily unavailable",
        "service unavailable",
        "bad gateway",
        "gateway timeout",
        "[429]",
        "[500]",
        "[502]",
        "[503]",
        "[504]",
        "quota exceeded",
    )
    return any(marker in message for marker in transient_markers)


def google_quota_exception() -> HTTPException:
    return HTTPException(
        status_code=429,
        detail="Google Sheet dang gioi han so lan doc trong 1 phut. Vui long cho khoang 60 giay roi thu lai.",
    )


def user_rows_cached(force: bool = False) -> list[dict[str, Any]]:
    now = datetime.now(timezone.utc)
    if not force and _USER_CACHE["expires"] > now:
        return _USER_CACHE["rows"]
    try:
        worksheet = users_worksheet()
        rows = worksheet_records(worksheet, USER_HEADERS)
        if not rows:
            admin_row = default_admin_row()
            worksheet.append_row([admin_row.get(header, "") for header in USER_HEADERS], value_input_option="RAW")
            rows = [admin_row]
    except Exception as exc:
        if is_google_quota_error(exc):
            if _USER_CACHE["rows"]:
                return _USER_CACHE["rows"]
            raise google_quota_exception() from exc
        rows = []
    _USER_CACHE["expires"] = now + timedelta(seconds=USER_CACHE_TTL_SECONDS)
    _USER_CACHE["rows"] = rows
    return rows


def authenticate_user(username: str, password: str) -> dict[str, Any] | None:
    username = str(username or "").strip()
    password = str(password or "")
    active_default_admin_exists = False
    for row in user_rows_cached():
        if normalize_text(row.get("status")) not in {"", "active", "dang hoat dong", "hoat dong"}:
            continue
        user_ok = compare_secret(row.get("username", ""), username)
        if compare_secret(row.get("username", ""), DEFAULT_ADMIN_USERNAME):
            active_default_admin_exists = True
        pass_ok = compare_secret(row.get("password", ""), password)
        if user_ok and pass_ok:
            return row
    if active_default_admin_exists and compare_secret(username, DEFAULT_ADMIN_USERNAME):
        return None
    user_ok = compare_secret(username, DEFAULT_ADMIN_USERNAME)
    pass_ok = compare_secret(password, DEFAULT_ADMIN_PASSWORD)
    return demo_admin_user() if user_ok and pass_ok else None


def session_user_from_request(request: Request) -> dict[str, Any] | None:
    token = request.cookies.get("dx_session", "")
    auth_header = request.headers.get("authorization", "")
    if auth_header.lower().startswith("bearer "):
        token = auth_header.split(" ", 1)[1].strip()
    if not token:
        return None
    session = SESSIONS.get(token)
    if not session:
        return None
    now = datetime.now(timezone.utc)
    last_seen = session.get("lastSeen")
    if not isinstance(last_seen, datetime) or (now - last_seen).total_seconds() > SESSION_IDLE_TIMEOUT_SECONDS:
        SESSIONS.pop(token, None)
        return None
    session["lastSeen"] = now
    return session.get("user")


def has_action(user: dict[str, Any], action: str) -> bool:
    role = str(user.get("role") or "")
    return role == "admin" or action in user_permissions(role, user.get("extraPermissions")).get("actions", [])


def can_export_report_path(user: dict[str, Any], path: str) -> bool:
    if path.endswith("/driver-remittance.xlsx"):
        return any(has_action(user, action) for action in ["reports_all", "reports_revenue", "export_excel"])
    if path.endswith("/summary.xlsx"):
        return any(has_action(user, action) for action in ["reports_all", "reports_revenue", "export_excel"])
    if path.endswith("/customers.xlsx"):
        return any(has_action(user, action) for action in ["reports_all", "export_customers"])
    if path.endswith("/invoices.xlsx"):
        return any(has_action(user, action) for action in ["reports_all", "export_invoices"])
    if path.endswith("/debts.xlsx"):
        return any(has_action(user, action) for action in ["reports_all", "export_debts"])
    if path.endswith("/commissions.xlsx"):
        return any(has_action(user, action) for action in ["reports_all", "export_commissions"])
    if path.endswith("/vouchers.xlsx"):
        return any(has_action(user, action) for action in ["reports_all", "export_excel", "export_vouchers"])
    if path.endswith("/orders.xlsx"):
        return any(has_action(user, action) for action in ["reports_all", "export_excel", "export_orders"])
    if path.endswith("/work-performance.xlsx"):
        return any(has_action(user, action) for action in ["reports_all", "export_excel", "export_orders"])
    return any(
        has_action(user, action)
        for action in ["reports_all", "reports_revenue", "reports_b2c", "export_excel", "export_b2b", "export_b2c"]
    )


def request_path_allowed_for_role(request: Request, user: dict[str, Any]) -> bool:
    path = request.url.path
    method = request.method.upper()
    role = str(user.get("role") or "")
    if not path.startswith("/api"):
        return True
    if path in {"/api/me", "/api/logout"}:
        return True
    if path == "/api/me/password" and method == "POST":
        return True
    if path.startswith("/api/users") or path.startswith("/api/logs"):
        return role == "admin"
    if path.startswith("/api/system-catalogs"):
        return method == "GET" or has_action(user, "manage_system_catalogs")
    if path.startswith("/api/calendar-vehicle-order"):
        return "calendar" in user_permissions(role, user.get("extraPermissions")).get("views", [])
    if path.startswith("/api/cskh-shift-reports"):
        if method == "GET":
            return "cskhShiftReports" in user_permissions(role, user.get("extraPermissions")).get("views", [])
        return role in {"admin", "cskh"}
    if path.startswith("/api/reopen-requests") and method != "GET":
        return has_action(user, "approve_reopen")
    if path.endswith("/reopen-requests") and method == "POST":
        return role == "cskh" and has_action(user, "request_reopen")
    if path.startswith("/api/reports"):
        return can_export_report_path(user, path)
    if path.startswith("/api/invoice-orders"):
        return any(has_action(user, action) for action in ["view_invoices", "manage_invoices", "export_invoices", "reports_all"])
    if path.startswith("/api/invoice-groups"):
        return any(has_action(user, action) for action in ["view_invoices", "manage_invoices", "export_invoices", "reports_all"])
    if path.startswith("/api/debt-orders"):
        if method == "GET":
            return any(has_action(user, action) for action in ["manage_debts", "export_debts", "reports_all"])
        return has_action(user, "manage_debts")
    if path.startswith("/api/commission-orders"):
        if method == "GET":
            return any(has_action(user, action) for action in ["manage_commissions", "export_commissions", "reports_all"])
        return has_action(user, "manage_commissions")
    if path.startswith("/api/orders") and path.endswith("/invoice-status") and method == "POST":
        return has_action(user, "manage_invoices")
    if path.startswith("/api/orders") and path.endswith("/remittance-status") and method == "POST":
        return has_action(user, "manage_remittance_status")
    if path.startswith("/api/shared-passengers") and path.endswith("/invoice-status") and method == "POST":
        return has_action(user, "manage_invoices")
    if path.startswith("/api/orders") and method in {"POST", "PUT", "PATCH", "DELETE"}:
        return role in {"admin", "cskh"}
    if path.startswith("/api/order-feedback") and method in {"POST", "PUT", "PATCH", "DELETE"}:
        return role in {"admin", "cskh"} and has_action(user, "manage_order_feedback")
    if method in {"POST", "PUT", "PATCH", "DELETE"}:
        if role == "admin":
            return True
        if path.startswith(("/api/vouchers", "/api/promotions")):
            return has_action(user, "manage_benefits")
        if role == "cskh" and path.startswith(
            (
                "/api/customers",
                "/api/tours",
                "/api/orders",
                "/api/franchise-vehicles",
            )
        ):
            return True
        return False
    return True


def current_user(request: Request) -> dict[str, Any]:
    user = getattr(request.state, "current_user", None)
    if not user:
        raise HTTPException(status_code=401, detail="Cần đăng nhập để sử dụng hệ thống.")
    return user


def current_user_display_name(request: Request) -> str:
    user = current_user(request)
    return str(user.get("displayName") or user.get("username") or "").strip()


def require_admin(request: Request) -> dict[str, Any]:
    user = current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Chỉ admin được thao tác chức năng này.")
    return user


@app.middleware("http")
async def public_demo_auth(request: Request, call_next):
    path = request.url.path
    if not path.startswith("/api/"):
        return await call_next(request)
    if path in {"/api/login", "/api/app-version"}:
        return await call_next(request)

    user = session_user_from_request(request)
    if not user:
        return unauthorized_response()
    request.state.current_user = user
    if not request_path_allowed_for_role(request, user):
        return json_forbidden_response()

    return await call_next(request)

app.mount("/static", StaticFiles(directory=PUBLIC_DIR), name="static")


class CustomerInput(BaseModel):
    tenKhach: str = Field(min_length=1)
    soDienThoai: str = Field(min_length=1)
    soCCCD: str = ""
    diaChi: str = ""
    loaiKhachHang: str = ""
    namSinh: str = ""
    gioiTinh: str = Field(min_length=1)
    nguonKhach: str = Field(min_length=1)
    nhanVienNhap: str = ""


class TourInput(BaseModel):
    diemDi: str = ""
    diemDen: str = ""
    tuyen: str = ""
    ghiChu: str = ""


class ContractPricingInput(BaseModel):
    oneWay: list[dict[str, Any]]
    roundTrip: list[dict[str, Any]]
    waiting: list[dict[str, Any]]
    rounding: dict[str, Any]


class FranchiseVehicleInput(BaseModel):
    bienKiemSoat: str = Field(min_length=1)
    dongXe: str = Field(min_length=1)
    hieuXe: str = Field(min_length=1)
    soCho: str = ""
    tenChuXe: str = Field(min_length=1)
    soDienThoaiChuXe: str = ""
    hoTenLaiXe: str = Field(min_length=1)
    soDienThoaiLaiXe: str = ""
    diaChiLaiXe: str = ""
    trangThai: str = "Đang hợp tác"
    ghiChu: str = ""


class OrderInput(BaseModel):
    khachHangId: str = ""
    tenKhach: str = ""
    soDienThoai: str = ""
    soCCCD: str = ""
    diaChi: str = ""
    loaiKhachHang: str = ""
    namSinh: str = ""
    gioiTinh: str = ""
    nguonKhach: str = ""
    nhanVienNhap: str = ""
    hopDongTourId: str = ""
    diemDon: str = ""
    diemTra: str = ""
    khuVucDatXe: str = ""
    giaTien: float = Field(default=0, ge=0)
    giamGia: float = Field(default=0, ge=0)
    ghiChuGiamGia: str = ""
    phuThu: float = Field(default=0, ge=0)
    lyDoPhuThu: str = ""
    daCoc: float = Field(default=0, ge=0)
    tyLeNopLai: float = Field(default=0, ge=0)
    bienKiemSoat: str = ""
    ngayGioDi: str = Field(min_length=1)
    ngayGioDuKienKetThuc: str = ""
    soCho: str = Field(min_length=1)
    loaiHopDong: str = Field(pattern="^(xe_nguyen_chuyen|xe_ghep)$")
    soVe: int = Field(default=0, ge=0)
    loaiKhach: str = ""
    yeuCauHoaDon: bool = False
    congNo: bool = False
    congNoChoAi: str = ""
    tenCongTy: str = ""
    maSoThue: str = ""
    diaChiHoaDon: str = ""
    emailHoaDon: str = ""
    ghiChu: str = ""
    khachXeGhep: list["SharedPassengerInput"] = Field(default_factory=list)
    voucherIds: list[str] = Field(default_factory=list)
    promotionIds: list[str] = Field(default_factory=list)


class OrderFeedbackInput(BaseModel):
    kenhChamSoc: str = Field(min_length=1)
    diemDanhGia: int = Field(ge=1, le=10)
    noiDungPhanHoi: str = Field(min_length=1)
    hinhThucXuLy: str = ""
    ketQuaXuLy: str = ""
    chuThich: str = ""


class AssignVehicleInput(BaseModel):
    bienKiemSoat: str = ""
    ngayGioDi: str = Field(min_length=1)
    ngayGioDuKienKetThuc: str = Field(min_length=1)
    tyLeNopLai: float = Field(default=0, ge=0)


class SharedPassengerInput(BaseModel):
    hoTen: str = Field(min_length=1)
    soDienThoai: str = Field(min_length=1)
    soCCCD: str = ""
    diaChi: str = ""
    gioiTinh: str = Field(min_length=1)
    namSinh: str = ""
    nguonKhach: str = Field(min_length=1)
    nhanVienNhap: str = ""
    loaiKhach: str = Field(pattern="^(B2C|B2B)$")
    diemDon: str = Field(min_length=1)
    diemTra: str = Field(min_length=1)
    soTien: float = Field(ge=0)
    giamGia: float = Field(default=0, ge=0)
    ghiChuGiamGia: str = ""
    phuThu: float = Field(default=0, ge=0)
    lyDoPhuThu: str = ""
    daCoc: float = Field(default=0, ge=0)
    voucherIds: list[str] = Field(default_factory=list)
    promotionIds: list[str] = Field(default_factory=list)
    yeuCauHoaDon: bool = False
    congNo: bool = False
    congNoChoAi: str = ""
    tenCongTy: str = ""
    maSoThue: str = ""
    diaChiHoaDon: str = ""
    emailHoaDon: str = ""


class CompleteOrderInput(BaseModel):
    ngayGioHoanThanh: str = Field(min_length=1)


class DriverNotificationStatusInput(BaseModel):
    trangThaiGuiTaiXe: str = Field(pattern="^(Chưa gửi tài xế|Đã gửi tài xế)$")


class RemittanceStatusInput(BaseModel):
    trangThaiNopTien: str = Field(pattern="^(Chưa nộp tiền|Đã nộp tiền)$")


class InvoiceStatusInput(BaseModel):
    trangThaiHoaDon: str = Field(pattern="^(Chưa xuất|Đã xuất)$")


class InvoiceGroupInput(BaseModel):
    orderIds: list[str] = Field(min_length=2)
    tenCongTy: str = Field(min_length=1)
    maSoThue: str = Field(min_length=1)
    diaChiHoaDon: str = Field(min_length=1)
    emailHoaDon: str = ""


class DebtStatusInput(BaseModel):
    trangThaiCongNo: str = Field(pattern="^(Chưa thu hồi|Đã thu hồi)$")


class CommissionStatusInput(BaseModel):
    trangThaiHoaHong: str = Field(pattern="^(Chưa thu|Đã thu)$")


class UserInput(BaseModel):
    username: str = Field(min_length=1)
    password: str = Field(min_length=4)
    displayName: str = Field(min_length=1)
    role: str = Field(min_length=1)
    status: str = "active"
    extraPermissions: list[str] = Field(default_factory=list)


class UserUpdateInput(BaseModel):
    username: str = Field(min_length=1)
    displayName: str = Field(min_length=1)
    role: str = Field(min_length=1)
    status: str = Field(pattern="^(active|inactive)$")
    extraPermissions: list[str] = Field(default_factory=list)


class SystemCatalogInput(BaseModel):
    loaiDanhMuc: str = Field(min_length=1)
    giaTri: str = Field(min_length=1)


class CalendarVehicleOrderInput(BaseModel):
    bienKiemSoat: list[str] = Field(default_factory=list)


class CskhShiftReportInput(BaseModel):
    ngay: str = Field(min_length=1)
    caLamViec: int = Field(ge=1, le=2)
    soLuongTinNhanMeta: int = Field(ge=0)
    soLuongKhachPhanHoi: int = Field(ge=0)
    soLuongCuocGoi: int = Field(ge=0)
    soLuongChatZalo: int = Field(ge=0)
    soLuongKhachTuWebsite: int = Field(ge=0)
    soLuongKhachTuEmail: int = Field(ge=0)
    soLuongTinNhanKhachVangLai: int = Field(ge=0)
    soLuongKhachPhanHoiTuTiktok: int = Field(ge=0)
    soLuongDonChotTuTiktok: int = Field(default=0, ge=0)
    tongSoLuongDonChot: int = Field(ge=0)


class CskhShiftReportDeleteInput(BaseModel):
    ngay: str = Field(min_length=1)
    caLamViec: int = Field(ge=1, le=2)
    nhanVienTruc: str = ""


class ChangePasswordInput(BaseModel):
    currentPassword: str = Field(min_length=1)
    newPassword: str = Field(min_length=4)


class ResetPasswordInput(BaseModel):
    newPassword: str = Field(min_length=4)


class LoginInput(BaseModel):
    username: str = Field(min_length=1)
    password: str = Field(min_length=1)


class ReopenRequestInput(BaseModel):
    reason: str = Field(min_length=1)


class ReopenReviewInput(BaseModel):
    adminNote: str = ""


class VoucherInput(BaseModel):
    maVoucher: str = Field(min_length=1)
    tenVoucher: str = Field(min_length=1)
    loaiGiaTri: str = Field(pattern="^(fixed|percent)$")
    giaTri: float = Field(ge=0)
    ngayBatDau: str = ""
    ngayHetHan: str = ""
    trangThai: str = "Đang áp dụng"
    ghiChu: str = ""


class VoucherBatchInput(BaseModel):
    idChienDich: str = Field(min_length=1, max_length=30, pattern=r"^[A-Za-z0-9]+$")
    tenLoPhatHanh: str = Field(min_length=1)
    loaiGiaTri: str = Field(pattern="^(fixed|percent)$")
    menhGia: float = Field(ge=0)
    soLuong: int = Field(ge=1, le=1000)
    ngayBatDau: str = ""
    ngayHetHan: str = ""
    ghiChu: str = ""


class PromotionInput(BaseModel):
    tenChuongTrinh: str = Field(min_length=1)
    loaiGiaTri: str = Field(pattern="^(fixed|percent)$")
    giaTri: float = Field(ge=0)
    ngayBatDau: str = ""
    ngayHetHan: str = ""
    trangThai: str = "Đang áp dụng"
    ghiChu: str = ""


OrderInput.model_rebuild()


def read_public_sheet(sheet_name: str) -> list[dict[str, str]]:
    selector = f"sheet={quote(sheet_name)}"
    url = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/gviz/tq?tqx=out:csv&{selector}"
    request = UrlRequest(url, headers={"User-Agent": "DixanhOrderManager/1.0"})
    with urlopen(request, timeout=20) as response:
        csv_text = response.read().decode("utf-8-sig")
    return list(csv.DictReader(StringIO(csv_text)))


def credentials_path() -> Path:
    return BASE_DIR / SERVICE_ACCOUNT_FILE


def get_gspread_client() -> Any:
    global _GSPREAD_CLIENT
    if _GSPREAD_CLIENT is not None:
        return _GSPREAD_CLIENT
    if gspread is None:
        raise HTTPException(status_code=500, detail="Chưa cài thư viện gspread.")
    path = credentials_path()
    if not path.exists():
        raise HTTPException(status_code=503, detail="Chưa cấu hình service-account.json.")
    _GSPREAD_CLIENT = gspread.service_account(
        filename=str(path),
        http_client=gspread.BackOffHTTPClient,
    )
    return _GSPREAD_CLIENT


def get_spreadsheet() -> Any:
    global _SPREADSHEET
    if _SPREADSHEET is None:
        _SPREADSHEET = get_gspread_client().open_by_key(SHEET_ID)
    return _SPREADSHEET


def get_legacy_spreadsheet() -> Any | None:
    global _LEGACY_SPREADSHEET
    if not LEGACY_READ_ENABLED or not LEGACY_SHEET_ID or LEGACY_SHEET_ID == SHEET_ID:
        return None
    if _LEGACY_SPREADSHEET is None:
        try:
            _LEGACY_SPREADSHEET = get_gspread_client().open_by_key(LEGACY_SHEET_ID)
        except Exception:
            return None
    return _LEGACY_SPREADSHEET


def legacy_worksheet_records(sheet_name: str, headers: list[str]) -> list[dict[str, Any]]:
    spreadsheet = get_legacy_spreadsheet()
    if spreadsheet is None:
        return []
    now = time.monotonic()
    with _LEGACY_VALUES_CACHE_LOCK:
        cached = copy.deepcopy(_LEGACY_VALUES_CACHE.get(sheet_name))
    if cached and cached.get("expires_at", 0) > now:
        values = copy.deepcopy(cached.get("values") or [])
    else:
        with _LEGACY_VALUES_CACHE_LOCK:
            refresh_lock = _LEGACY_REFRESH_LOCKS.setdefault(sheet_name, threading.Lock())
        # Nhiều API có thể cần cùng một tab lưu trữ. Chỉ cho phép một request
        # đọc Google Sheet; các request còn lại chờ và dùng chung kết quả cache.
        with refresh_lock:
            now = time.monotonic()
            with _LEGACY_VALUES_CACHE_LOCK:
                refreshed = copy.deepcopy(_LEGACY_VALUES_CACHE.get(sheet_name))
            if refreshed and refreshed.get("expires_at", 0) > now:
                values = copy.deepcopy(refreshed.get("values") or [])
            else:
                try:
                    worksheet = spreadsheet.worksheet(sheet_name)
                    values = worksheet.get_all_values()
                except Exception:
                    # Dữ liệu lưu trữ không thay đổi thường xuyên; khi Google
                    # phản hồi chậm vẫn ưu tiên bản cache cũ thay vì làm treo app.
                    if refreshed:
                        values = copy.deepcopy(refreshed.get("values") or [])
                    else:
                        return []
                else:
                    with _LEGACY_VALUES_CACHE_LOCK:
                        _LEGACY_VALUES_CACHE[sheet_name] = {
                            "expires_at": now + LEGACY_SHEET_VALUES_CACHE_TTL_SECONDS,
                            "values": copy.deepcopy(values),
                        }
    if len(values) < 2:
        return []
    actual_headers = values[0]
    indexes = {header: actual_headers.index(header) for header in headers if header in actual_headers}
    records = []
    for row in values[1:]:
        record = {header: row[index] if index < len(row) else "" for header, index in indexes.items()}
        for header in headers:
            record.setdefault(header, "")
        if not is_deleted_row(record) and any(str(value).strip() for value in record.values()):
            record["_archived"] = True
            records.append(record)
    return records


def merge_records_by_id(current: list[dict[str, Any]], archived: list[dict[str, Any]]) -> list[dict[str, Any]]:
    seen = {str(row.get("id") or "").strip() for row in current if str(row.get("id") or "").strip()}
    return current + [row for row in archived if not str(row.get("id") or "").strip() or str(row.get("id") or "").strip() not in seen]


def get_worksheet(sheet_name: str, headers: list[str]) -> Any:
    spreadsheet = get_spreadsheet()
    worksheet = _WORKSHEET_CACHE.get(sheet_name)
    if worksheet is None:
        try:
            worksheet = spreadsheet.worksheet(sheet_name)
        except gspread.WorksheetNotFound:
            worksheet = spreadsheet.add_worksheet(title=sheet_name, rows=1000, cols=len(headers))
            worksheet.append_row(headers, value_input_option="RAW")
            _WORKSHEET_CACHE[sheet_name] = worksheet
            _HEADER_CHECKED_SHEETS.add(sheet_name)
            return worksheet
        _WORKSHEET_CACHE[sheet_name] = worksheet

    if sheet_name not in _HEADER_CHECKED_SHEETS and worksheet.row_values(1) != headers:
        managed_sheet = sheet_name in {
            CUSTOMERS_SHEET_NAME,
            TOURS_SHEET_NAME,
            FRANCHISE_VEHICLES_SHEET_NAME,
            ORDERS_SHEET_NAME,
            SHARED_RIDE_SHEET_NAME,
            VOUCHERS_SHEET_NAME,
            PROMOTIONS_SHEET_NAME,
            ORDER_BENEFITS_SHEET_NAME,
            USERS_SHEET_NAME,
            REOPEN_REQUESTS_SHEET_NAME,
            SYSTEM_LOGS_SHEET_NAME,
            ORDER_FEEDBACK_SHEET_NAME,
            INVOICE_GROUPS_SHEET_NAME,
            SYSTEM_CATALOGS_SHEET_NAME,
            CSKH_SHIFT_REPORTS_SHEET_NAME,
            CALENDAR_VEHICLE_ORDER_SHEET_NAME,
            CONTRACT_PRICING_SHEET_NAME,
        }
        if managed_sheet and worksheet.col_count < len(headers):
            worksheet.resize(cols=len(headers))
        end_col = gspread.utils.rowcol_to_a1(1, len(headers)).replace("1", "")
        worksheet.update(f"A1:{end_col}1", [headers], value_input_option="RAW")
        if managed_sheet and worksheet.col_count != len(headers):
            worksheet.resize(cols=len(headers))
    _HEADER_CHECKED_SHEETS.add(sheet_name)
    return worksheet


def worksheet_cache_key(worksheet: Any) -> str:
    return str(getattr(worksheet, "title", "") or id(worksheet))


def invalidate_worksheet_cache(worksheet: Any) -> None:
    key = worksheet_cache_key(worksheet)
    with _SHEET_VALUES_CACHE_LOCK:
        cached = _SHEET_VALUES_CACHE.get(key)
        if cached:
            cached["expires_at"] = 0.0


def worksheet_refresh_lock(key: str) -> threading.Lock:
    with _SHEET_VALUES_CACHE_LOCK:
        return _SHEET_REFRESH_LOCKS.setdefault(key, threading.Lock())


def worksheet_values(worksheet: Any, force_refresh: bool = False) -> list[list[Any]]:
    key = worksheet_cache_key(worksheet)
    now = time.monotonic()
    with _SHEET_VALUES_CACHE_LOCK:
        cached = _SHEET_VALUES_CACHE.get(key)
        if not force_refresh and cached and cached["expires_at"] > now:
            return copy.deepcopy(cached["values"])

    # Chỉ tuần tự hóa việc đọc cùng một sheet. Không giữ khóa cache chung trong lúc
    # chờ Google, vì một sheet chậm sẽ làm tất cả API khác bị xếp hàng và hết timeout.
    refresh_lock = worksheet_refresh_lock(key)
    with refresh_lock:
        with _SHEET_VALUES_CACHE_LOCK:
            cached = _SHEET_VALUES_CACHE.get(key)
            if not force_refresh and cached and cached["expires_at"] > time.monotonic():
                return copy.deepcopy(cached["values"])
        try:
            values = worksheet.get_all_values()
        except Exception as exc:
            # Khi Google tạm chậm, lỗi mạng hoặc giới hạn quota, dữ liệu gần nhất
            # vẫn tốt hơn việc làm toàn bộ màn hình trả lỗi hoặc trắng dữ liệu.
            if cached and is_google_transient_error(exc):
                return copy.deepcopy(cached["values"])
            if is_google_quota_error(exc):
                raise google_quota_exception() from exc
            raise
        with _SHEET_VALUES_CACHE_LOCK:
            _SHEET_VALUES_CACHE[key] = {
                "expires_at": time.monotonic() + SHEET_VALUES_CACHE_TTL_SECONDS,
                "values": copy.deepcopy(values),
            }
        return values


def worksheet_records(worksheet: Any, expected_headers: list[str] | None = None) -> list[dict[str, Any]]:
    values = worksheet_values(worksheet)
    if len(values) < 2:
        return []
    headers = expected_headers or values[0]
    records = []
    for row in values[1:]:
        padded = row + [""] * max(len(headers) - len(row), 0)
        record = {header: padded[index] for index, header in enumerate(headers)}
        if is_deleted_row(record):
            continue
        if any(str(value).strip() for value in record.values()):
            records.append(record)
    return records


def is_deleted_row(row: dict[str, Any]) -> bool:
    return bool(str(row.get("deletedAt") or "").strip()) or normalize_text(row.get("trangThai")) == "da xoa" or normalize_text(row.get("status")) == "da xoa"


def soft_delete_row(worksheet: Any, row_number: int, headers: list[str], row: dict[str, Any], request: Request) -> dict[str, Any]:
    deleted = dict(row)
    deleted["deletedAt"] = now_iso()
    deleted["deletedBy"] = current_user_display_name(request)
    if "trangThai" in headers:
        deleted["trangThai"] = "Đã xóa"
    if "status" in headers:
        deleted["status"] = "Đã xóa"
    update_row_by_headers(worksheet, row_number, headers, deleted)
    return deleted


def customers_worksheet() -> Any:
    return get_worksheet(CUSTOMERS_SHEET_NAME, CUSTOMER_HEADERS)


def franchise_vehicles_worksheet() -> Any:
    worksheet = get_worksheet(FRANCHISE_VEHICLES_SHEET_NAME, FRANCHISE_VEHICLE_HEADERS)
    migrate_franchise_vehicle_rows(worksheet)
    return worksheet


def migrate_franchise_vehicle_rows(worksheet: Any) -> None:
    values = worksheet_values(worksheet)
    if not values:
        return
    headers = values[0]
    if "diaChiLaiXe" in headers:
        address_index = headers.index("diaChiLaiXe")
        for row_number, row in enumerate(values[1:], start=2):
            padded = row + [""] * max(len(headers) - len(row), 0)
            maybe_status = normalize_text(padded[address_index] if address_index < len(padded) else "")
            if maybe_status in {"dang hop tac", "tam ngung", "ngung hop tac", "da xoa"}:
                repaired = (row[:address_index] + [""] + row[address_index:])[: len(headers)]
                update_row_by_headers(worksheet, row_number, headers, dict(zip(headers, repaired)))
        return
    if "soDienThoaiLaiXe" not in headers:
        return
    insert_at = headers.index("soDienThoaiLaiXe") + 2
    worksheet.insert_cols([["diaChiLaiXe"]] + [[""] for _ in values[1:]], col=insert_at, value_input_option="RAW")
    invalidate_worksheet_cache(worksheet)
    _HEADER_CHECKED_SHEETS.discard(FRANCHISE_VEHICLES_SHEET_NAME)


def migrate_tour_rows(worksheet: Any) -> None:
    values = worksheet_values(worksheet)
    if len(values) < 2:
        return
    for row_number, row in enumerate(values[1:], start=2):
        padded = row + [""] * max(len(TOUR_HEADERS) - len(row), 0)
        route_candidate = str(padded[1] or "").strip()
        destination = str(padded[2] or "").strip()
        current_route = str(padded[3] or "").strip()
        if not route_candidate or destination:
            continue
        start, end = split_route(route_candidate)
        if not end:
            continue
        updated = {
            "id": padded[0],
            "diemDi": start,
            "diemDen": end,
            "tuyen": route_text(start, end),
            "ghiChu": "",
            "createdAt": current_route if "T" in current_route else now_iso(),
        }
        update_row_by_headers(worksheet, row_number, TOUR_HEADERS, updated)


def tours_worksheet() -> Any:
    worksheet = get_worksheet(TOURS_SHEET_NAME, TOUR_HEADERS)
    migrate_tour_rows(worksheet)
    return worksheet


def contract_pricing_worksheet() -> Any:
    return get_worksheet(CONTRACT_PRICING_SHEET_NAME, CONTRACT_PRICING_HEADERS)


def contract_pricing_config() -> dict[str, Any]:
    rows = worksheet_records(contract_pricing_worksheet(), CONTRACT_PRICING_HEADERS)
    if not rows:
        return copy.deepcopy(DEFAULT_CONTRACT_PRICING)
    try:
        loaded = json.loads(str(rows[0].get("configJson") or "{}"))
        if not isinstance(loaded, dict):
            return copy.deepcopy(DEFAULT_CONTRACT_PRICING)
        round_trip = loaded.get("roundTrip")
        if isinstance(round_trip, list):
            migrated: list[dict[str, Any]] = []
            for tier in round_trip:
                if not isinstance(tier, dict):
                    continue
                minimum = int(tier.get("minKm") or 0)
                maximum = tier.get("maxKm")
                if minimum <= 1 and (maximum is None or int(maximum) >= 19):
                    migrated.append({"minKm": 1, "maxKm": 19, "percentages": {"4": 100, "7": 100, "16": 100}})
                    if maximum is None or int(maximum) >= 20:
                        remainder = copy.deepcopy(tier)
                        remainder["minKm"] = 20
                        migrated.append(remainder)
                else:
                    migrated.append(tier)
            loaded["roundTrip"] = migrated
        return loaded
    except (TypeError, ValueError, json.JSONDecodeError):
        return copy.deepcopy(DEFAULT_CONTRACT_PRICING)


def tour_records() -> list[dict[str, Any]]:
    records = worksheet_records(tours_worksheet(), TOUR_HEADERS)
    for row in records:
        if not row.get("diemDi") and row.get("tuyen"):
            row["diemDi"], row["diemDen"] = split_route(row.get("tuyen"))
        row["tuyen"] = route_text(row.get("diemDi"), row.get("diemDen"), row.get("tuyen"))
    return records


def tour_payload_values(payload: TourInput) -> tuple[str, str, str]:
    start = payload.diemDi.strip()
    end = payload.diemDen.strip()
    if not start or not end:
        start, end = split_route(payload.tuyen)
    if not start or not end:
        raise HTTPException(status_code=422, detail="Vui lÃ²ng nháº­p Ä‘iá»ƒm Ä‘i vÃ  Ä‘iá»ƒm Ä‘áº¿n.")
    return start, end, route_text(start, end)


def orders_worksheet() -> Any:
    return get_worksheet(ORDERS_SHEET_NAME, ORDER_HEADERS)


def shared_ride_worksheet() -> Any:
    return get_worksheet(SHARED_RIDE_SHEET_NAME, SHARED_RIDE_HEADERS)


def all_order_records() -> list[dict[str, Any]]:
    current = worksheet_records(orders_worksheet(), ORDER_HEADERS)
    records = merge_records_by_id(current, legacy_worksheet_records(ORDERS_SHEET_NAME, ORDER_HEADERS))
    customers_by_id = {
        str(customer.get("id") or "").strip(): customer
        for customer in customer_records()
        if str(customer.get("id") or "").strip()
    }
    for row in records:
        customer = customers_by_id.get(str(row.get("khachHangId") or "").strip())
        if not customer:
            continue
        # Thông tin nhận diện khách luôn lấy theo hồ sơ khách hàng mới nhất.
        # Các trường tài chính, hóa đơn và hành trình của đơn vẫn là dữ liệu lịch sử.
        row["tenKhach"] = customer.get("tenKhach", "")
        row["soDienThoai"] = customer.get("soDienThoai", "")
    return records


def all_shared_ride_records() -> list[dict[str, Any]]:
    current = worksheet_records(shared_ride_worksheet(), SHARED_RIDE_HEADERS)
    return merge_records_by_id(current, legacy_worksheet_records(SHARED_RIDE_SHEET_NAME, SHARED_RIDE_HEADERS))


def sync_customer_profile_to_current_orders(customer_id: str, before: dict[str, Any], after: dict[str, Any]) -> int:
    """Đồng bộ hồ sơ khách sang các dòng đơn hiện hành, không đụng dữ liệu nghiệp vụ."""
    updated_rows = 0
    old_phone = normalize_phone(before.get("soDienThoai"))

    order_sheet = orders_worksheet()
    for row_number, row in enumerate(worksheet_records(order_sheet, ORDER_HEADERS), start=2):
        linked_by_id = str(row.get("khachHangId") or "").strip() == str(customer_id).strip()
        linked_by_old_phone = not str(row.get("khachHangId") or "").strip() and old_phone and normalize_phone(row.get("soDienThoai")) == old_phone
        if not (linked_by_id or linked_by_old_phone):
            continue
        row["khachHangId"] = customer_id
        row["tenKhach"] = after.get("tenKhach", "")
        row["soDienThoai"] = after.get("soDienThoai", "")
        update_row_by_headers(order_sheet, row_number, ORDER_HEADERS, row)
        updated_rows += 1

    # Khách xe ghép chưa có cột khachHangId trong cấu trúc cũ, nên đồng bộ các
    # dòng hiện hành bằng SĐT cũ ngay trong chính thao tác sửa khách hàng.
    if old_phone:
        shared_sheet = shared_ride_worksheet()
        for row_number, row in enumerate(worksheet_records(shared_sheet, SHARED_RIDE_HEADERS), start=2):
            if normalize_phone(row.get("soDienThoai")) != old_phone:
                continue
            row["hoTen"] = after.get("tenKhach", "")
            row["soDienThoai"] = after.get("soDienThoai", "")
            row["soCCCD"] = after.get("soCCCD", "")
            row["diaChi"] = after.get("diaChi", "")
            row["gioiTinh"] = after.get("gioiTinh", "")
            row["namSinh"] = after.get("namSinh", "")
            row["nguonKhach"] = after.get("nguonKhach", "")
            update_row_by_headers(shared_sheet, row_number, SHARED_RIDE_HEADERS, row)
            updated_rows += 1
    return updated_rows


def all_system_log_records() -> list[dict[str, Any]]:
    current = worksheet_records(system_logs_worksheet(), SYSTEM_LOG_HEADERS)
    return merge_records_by_id(current, legacy_worksheet_records(SYSTEM_LOGS_SHEET_NAME, SYSTEM_LOG_HEADERS))


def vouchers_worksheet() -> Any:
    return get_worksheet(VOUCHERS_SHEET_NAME, VOUCHER_HEADERS)


def promotions_worksheet() -> Any:
    return get_worksheet(PROMOTIONS_SHEET_NAME, PROMOTION_HEADERS)


def order_benefits_worksheet() -> Any:
    return get_worksheet(ORDER_BENEFITS_SHEET_NAME, ORDER_BENEFIT_HEADERS)


def users_worksheet() -> Any:
    return get_worksheet(USERS_SHEET_NAME, USER_HEADERS)


def system_catalogs_worksheet() -> Any:
    return get_worksheet(SYSTEM_CATALOGS_SHEET_NAME, SYSTEM_CATALOG_HEADERS)


def cskh_shift_reports_worksheet() -> Any:
    return get_worksheet(CSKH_SHIFT_REPORTS_SHEET_NAME, CSKH_SHIFT_REPORT_HEADERS)


def calendar_vehicle_order_worksheet() -> Any:
    return get_worksheet(CALENDAR_VEHICLE_ORDER_SHEET_NAME, CALENDAR_VEHICLE_ORDER_HEADERS)


def cskh_shift_report_records(include_deleted: bool = False) -> list[dict[str, Any]]:
    rows = worksheet_records(cskh_shift_reports_worksheet(), CSKH_SHIFT_REPORT_HEADERS)
    if include_deleted:
        return rows
    return [row for row in rows if normalize_text(row.get("Trạng Thái")) != "da xoa"]


def reopen_requests_worksheet() -> Any:
    return get_worksheet(REOPEN_REQUESTS_SHEET_NAME, REOPEN_REQUEST_HEADERS)


def system_logs_worksheet() -> Any:
    return get_worksheet(SYSTEM_LOGS_SHEET_NAME, SYSTEM_LOG_HEADERS)


def order_feedback_worksheet() -> Any:
    return get_worksheet(ORDER_FEEDBACK_SHEET_NAME, ORDER_FEEDBACK_HEADERS)


def invoice_groups_worksheet() -> Any:
    return get_worksheet(INVOICE_GROUPS_SHEET_NAME, INVOICE_GROUP_HEADERS)


def make_id(prefix: str) -> str:
    return f"{prefix}-{datetime.now().strftime('%Y%m%d-%H%M%S%f')}-{secrets.token_hex(3).upper()}"


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def normalize_text(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    return text.replace("đ", "d")


def normalize_phone(value: Any) -> str:
    phone = "".join(char for char in str(value or "") if char.isdigit())
    # Google Sheets may store a Vietnamese mobile number as a number and drop
    # its leading zero. Restore it only for the unambiguous 9-digit mobile
    # prefixes; leave international and unusual values untouched.
    if len(phone) == 9 and phone[0] in "35789":
        return f"0{phone}"
    return phone


def normalize_customer_segment(value: Any) -> str:
    normalized = normalize_text(value)
    if normalized == "b2b" or "doanh nghiep" in normalized:
        return "B2B"
    if normalized == "b2c" or "ca nhan" in normalized:
        return "B2C"
    return ""


def validate_customer_phone(value: Any, label: str = "Số điện thoại") -> str:
    phone = normalize_phone(value)
    if not re.fullmatch(r"0\d{9}", phone):
        raise HTTPException(
            status_code=422,
            detail=f"{label} phải gồm đúng 10 chữ số và bắt đầu bằng số 0.",
        )
    return phone


FRANCHISE_PLATE_RE = re.compile(r"^\d{2}[A-Z]-\d{3}\.\d{2}$")


def normalize_franchise_plate(value: Any) -> str:
    plate = str(value or "").strip().upper()
    if not FRANCHISE_PLATE_RE.fullmatch(plate):
        raise HTTPException(status_code=422, detail="Biển số xe phải đúng định dạng 68A-123.45.")
    return plate


def split_route(value: Any) -> tuple[str, str]:
    text = str(value or "").strip()
    for separator in ("->", " - ", "-", "→"):
        if separator in text:
            start, end = text.split(separator, 1)
            return start.strip(), end.strip()
    return text, ""


def route_text(start: Any, end: Any, fallback: Any = "") -> str:
    start_text = str(start or "").strip()
    end_text = str(end or "").strip()
    if start_text and end_text:
        return f"{start_text} - {end_text}"
    fallback_text = str(fallback or "").strip()
    return fallback_text or start_text or end_text


def parse_datetime(value: str) -> datetime:
    text = str(value or "").strip()
    for fmt in ("%d/%m/%Y %H:%M", "%d/%m/%Y %I:%M %p"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    try:
        return datetime.fromisoformat(text.replace(" ", "T"))
    except ValueError as exc:
        raise HTTPException(status_code=422, detail="Ngày giờ không hợp lệ. Vui lòng nhập theo dạng dd/MM/yyyy HH:mm.") from exc


def parse_existing_datetime(value: Any) -> datetime | None:
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%d/%m/%Y %H:%M", "%d/%m/%Y %I:%M %p", "%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    try:
        parsed = datetime.fromisoformat(text.replace(" ", "T"))
        if parsed.tzinfo is not None:
            vietnam_timezone = timezone(timedelta(hours=7))
            parsed = parsed.astimezone(vietnam_timezone).replace(tzinfo=None)
        return parsed
    except ValueError:
        return None


def parse_existing_date(value: Any) -> datetime | None:
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    parsed = parse_existing_datetime(text)
    return parsed


def selected_report_date(value: Any) -> datetime:
    text = str(value or "").strip()
    if not text:
        return datetime.now()
    parsed = parse_existing_date(text)
    if parsed is None:
        raise HTTPException(status_code=422, detail="Ngày xuất báo cáo không hợp lệ.")
    return parsed


def rows_for_departure_date(rows: list[dict[str, Any]], selected_date: datetime) -> list[dict[str, Any]]:
    target = selected_date.date()
    return [
        row
        for row in rows
        if (departure := parse_existing_datetime(row.get("ngayGioDi"))) is not None
        and departure.date() == target
    ]


def selected_report_range(tu_ngay: Any = "", den_ngay: Any = "", ngay: Any = "") -> tuple[datetime, datetime]:
    start = selected_report_date(tu_ngay or ngay)
    end = selected_report_date(den_ngay or tu_ngay or ngay)
    if start.date() > end.date():
        raise HTTPException(status_code=422, detail="Từ ngày không được lớn hơn đến ngày.")
    return start, end


def rows_for_departure_range(rows: list[dict[str, Any]], start: datetime, end: datetime) -> list[dict[str, Any]]:
    start_date, end_date = start.date(), end.date()
    return [
        row for row in rows
        if (departure := parse_existing_datetime(row.get("ngayGioDi"))) is not None
        and start_date <= departure.date() <= end_date
    ]


def benefit_is_active(row: dict[str, Any]) -> bool:
    status = normalize_text(row.get("trangThai"))
    if "ngung" in status or "tam ngung" in status or "het han" in status:
        return False
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    start = parse_existing_date(row.get("ngayBatDau"))
    end = parse_existing_date(row.get("ngayHetHan"))
    if start and today < start:
        return False
    if end and today > end:
        return False
    return True


def benefit_status(row: dict[str, Any], used: bool = False) -> str:
    if used:
        return "Đã sử dụng"
    status = normalize_text(row.get("trangThai"))
    if "ngung" in status or "tam ngung" in status:
        return "Tạm ngưng"
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    start = parse_existing_date(row.get("ngayBatDau"))
    end = parse_existing_date(row.get("ngayHetHan"))
    if start and today < start:
        return "Chưa đến hạn"
    if end and today > end:
        return "Hết hạn"
    return "Còn hạn"


def benefit_discount(row: dict[str, Any], base_amount: float) -> float:
    value = money_value(row.get("giaTri"))
    if normalize_text(row.get("loaiGiaTri")) in {"percent", "phan tram", "%"}:
        return round(max(base_amount, 0) * min(value, 100) / 100)
    return min(value, max(base_amount, 0))


def benefit_key(row: dict[str, Any], kind: str) -> str:
    if kind == "voucher":
        return str(row.get("maVoucher") or row.get("id") or "").strip()
    return str(row.get("id") or row.get("tenChuongTrinh") or "").strip()


def benefit_row_by_key(rows: list[dict[str, Any]], key: str, kind: str) -> dict[str, Any] | None:
    key_text = str(key or "").strip()
    if not key_text:
        return None
    return next(
        (
            row
            for row in rows
            if key_text
            in {
                str(row.get("id") or "").strip(),
                str(row.get("maVoucher") or "").strip(),
                str(row.get("tenVoucher") or "").strip(),
                str(row.get("tenChuongTrinh") or "").strip(),
                benefit_key(row, kind),
            }
        ),
        None,
    )


def benefit_usage_keys(row: dict[str, Any]) -> set[str]:
    return {
        str(row.get("uuDaiId") or "").strip(),
        str(row.get("maUuDai") or "").strip(),
        str(row.get("tenUuDai") or "").strip(),
    }


def build_benefit_rows(
    order_id: str,
    customer_id: str,
    customer_name: str,
    voucher_ids: list[str],
    promotion_ids: list[str],
    amount: float,
    vouchers: list[dict[str, Any]],
    promotions: list[dict[str, Any]],
    benefit_usage: list[dict[str, Any]],
    used_voucher_ids: set[str],
    percent_base_amount: float | None = None,
) -> tuple[list[list[Any]], float, list[dict[str, Any]], list[dict[str, Any]]]:
    rows: list[list[Any]] = []
    remaining_amount = max(amount, 0)
    percent_base = max(percent_base_amount if percent_base_amount is not None else amount, 0)
    selected_vouchers = []
    selected_promotions = []
    total_discount = 0.0

    for voucher_id in dict.fromkeys(voucher_ids):
        voucher = benefit_row_by_key(vouchers, voucher_id, "voucher")
        if voucher is None:
            raise HTTPException(status_code=404, detail="Không tìm thấy voucher đã chọn.")
        voucher_key = benefit_key(voucher, "voucher")
        if voucher_key in used_voucher_ids or any(item.get("loaiUuDai") == "voucher" and voucher_key in benefit_usage_keys(item) for item in benefit_usage):
            raise HTTPException(status_code=409, detail=f"Voucher {voucher.get('maVoucher')} đã được sử dụng.")
        if not benefit_is_active(voucher):
            raise HTTPException(status_code=422, detail=f"Voucher {voucher.get('maVoucher')} không còn hiệu lực.")
        discount = min(remaining_amount, benefit_discount(voucher, percent_base))
        total_discount += discount
        remaining_amount = max(remaining_amount - discount, 0)
        selected_vouchers.append(voucher)
        used_voucher_ids.add(voucher_key)
        rows.append(
            [
                make_id("UD"),
                order_id,
                customer_id,
                customer_name,
                "voucher",
                voucher.get("id") or voucher_key,
                voucher.get("maVoucher", ""),
                voucher.get("tenVoucher", ""),
                voucher.get("loaiGiaTri", ""),
                voucher.get("giaTri", ""),
                discount,
                now_iso(),
            ]
        )

    for promotion_id in dict.fromkeys(promotion_ids):
        promotion = benefit_row_by_key(promotions, promotion_id, "promotion")
        if promotion is None:
            raise HTTPException(status_code=404, detail="Không tìm thấy chương trình khuyến mãi đã chọn.")
        if not benefit_is_active(promotion):
            raise HTTPException(status_code=422, detail=f"Chương trình {promotion.get('tenChuongTrinh')} không còn hiệu lực.")
        discount = min(remaining_amount, benefit_discount(promotion, percent_base))
        total_discount += discount
        remaining_amount = max(remaining_amount - discount, 0)
        selected_promotions.append(promotion)
        promotion_key = benefit_key(promotion, "promotion")
        rows.append(
            [
                make_id("UD"),
                order_id,
                customer_id,
                customer_name,
                "promotion",
                promotion.get("id") or promotion_key,
                "",
                promotion.get("tenChuongTrinh", ""),
                promotion.get("loaiGiaTri", ""),
                promotion.get("giaTri", ""),
                discount,
                now_iso(),
            ]
        )

    return rows, total_discount, selected_vouchers, selected_promotions


def order_benefit_records() -> list[dict[str, Any]]:
    return worksheet_records(order_benefits_worksheet(), ORDER_BENEFIT_HEADERS)


def replace_order_benefits(order_id: str, rows: list[list[Any]]) -> None:
    worksheet = order_benefits_worksheet()
    values = worksheet_values(worksheet)
    order_id_column = ORDER_BENEFIT_HEADERS.index("donHangId")
    end_column = re.sub(r"\d+$", "", gspread.utils.rowcol_to_a1(1, len(ORDER_BENEFIT_HEADERS)))
    ranges = [
        f"A{row_number}:{end_column}{row_number}"
        for row_number, row in enumerate(values[1:], start=2)
        if len(row) > order_id_column and str(row[order_id_column] or "") == str(order_id)
    ]
    for start in range(0, len(ranges), 200):
        worksheet.batch_clear(ranges[start : start + 200])
    if rows:
        append_worksheet_rows(worksheet, rows)
    invalidate_worksheet_cache(worksheet)


def voucher_campaign_name(value: Any) -> str:
    return re.sub(r"\s+#\d{3,}$", "", str(value or "").strip())


def stored_voucher_usage() -> dict[str, dict[str, Any]]:
    """Find voucher use from saved orders, including legacy rows without benefit details."""
    result: dict[str, dict[str, Any]] = {}
    for order in all_order_records():
        for code in [value.strip() for value in str(order.get("voucherCodes") or "").split(",") if value.strip()]:
            result.setdefault(
                normalize_text(code),
                {
                    "donHangId": order.get("id") or "",
                    "khachHangId": order.get("khachHangId") or "",
                    "tenKhach": order.get("tenKhach") or "",
                },
            )
    for passenger in all_shared_ride_records():
        for code in [value.strip() for value in str(passenger.get("voucherCodes") or "").split(",") if value.strip()]:
            result.setdefault(
                normalize_text(code),
                {
                    "donHangId": passenger.get("donHangId") or "",
                    "khachHangId": passenger.get("khachHangId") or "",
                    "tenKhach": passenger.get("hoTen") or "",
                },
            )
    return result


def voucher_records() -> list[dict[str, Any]]:
    rows = worksheet_records(vouchers_worksheet(), VOUCHER_HEADERS)
    usage = order_benefit_records()
    direct_usage = stored_voucher_usage()
    for row in rows:
        row["tenVoucher"] = voucher_campaign_name(row.get("tenVoucher"))
        key = benefit_key(row, "voucher")
        used = next((item for item in usage if item.get("loaiUuDai") == "voucher" and key and key in benefit_usage_keys(item)), None)
        saved_usage = direct_usage.get(normalize_text(row.get("maVoucher"))) or {}
        usage_info = used or saved_usage
        row["trangThaiSuDung"] = benefit_status(row, bool(usage_info))
        row["donHangId"] = usage_info.get("donHangId", "") if usage_info else ""
        row["khachHangId"] = usage_info.get("khachHangId", "") if usage_info else ""
        row["tenKhach"] = usage_info.get("tenKhach", "") if usage_info else ""
    return rows


def promotion_records() -> list[dict[str, Any]]:
    rows = worksheet_records(promotions_worksheet(), PROMOTION_HEADERS)
    for row in rows:
        row["trangThaiHieuLuc"] = benefit_status(row)
    return rows


def validate_unique_promotion_name(name: Any, exclude_id: str = "") -> str:
    cleaned_name = " ".join(str(name or "").split())
    normalized_name = " ".join(normalize_text(cleaned_name).split())
    if not normalized_name:
        raise HTTPException(status_code=422, detail="Vui lòng nhập tên chương trình khuyến mãi.")
    for row in promotion_records():
        if exclude_id and str(row.get("id") or "").strip() == str(exclude_id).strip():
            continue
        existing_name = " ".join(normalize_text(row.get("tenChuongTrinh")).split())
        if existing_name == normalized_name:
            raise HTTPException(
                status_code=409,
                detail=f'Tên chương trình khuyến mãi "{cleaned_name}" đã tồn tại.',
            )
    return cleaned_name
    for fmt in ("%d/%m/%Y %H:%M", "%d/%m/%Y %I:%M %p"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    try:
        return datetime.fromisoformat(text.replace(" ", "T"))
    except ValueError:
        return None


def order_is_done(row: dict[str, Any]) -> bool:
    status = normalize_text(row.get("trangThai"))
    return bool(row.get("ngayGioHoanThanh")) or status in {"da hoan thanh", "hoan thanh"}


def order_is_cancelled(row: dict[str, Any]) -> bool:
    return "huy" in normalize_text(row.get("trangThai"))


def order_range(row: dict[str, Any]) -> tuple[datetime, datetime] | None:
    start_at = parse_existing_datetime(row.get("ngayGioDi"))
    if start_at is None:
        return None
    end_at = (
        parse_existing_datetime(row.get("ngayGioHoanThanh"))
        or parse_existing_datetime(row.get("ngayGioDuKienKetThuc"))
        or start_at + timedelta(hours=4)
    )
    return start_at, end_at


def ranges_overlap(left_start: datetime, left_end: datetime, right_start: datetime, right_end: datetime) -> bool:
    return left_start < right_end and left_end > right_start


def conflicting_order(
    orders: list[dict[str, Any]],
    plate: str,
    start_at: datetime,
    end_at: datetime,
    exclude_order_id: str = "",
) -> dict[str, Any] | None:
    normalized_plate = normalize_text(plate)
    for row in orders:
        if exclude_order_id and str(row.get("id")) == exclude_order_id:
            continue
        if order_is_cancelled(row) or normalize_text(row.get("bienKiemSoat")) != normalized_plate:
            continue
        existing_range = order_range(row)
        if existing_range is None or ranges_overlap(existing_range[0], existing_range[1], start_at, end_at):
            return row
    return None


def find_row_by_id(worksheet: Any, row_id: str) -> int | None:
    values = worksheet_values(worksheet)
    if not values:
        return None
    headers = values[0]
    try:
        id_column = headers.index("id")
    except ValueError:
        return None
    for row_number, row in enumerate(values[1:], start=2):
        if id_column < len(row) and str(row[id_column]) == str(row_id):
            return row_number
    return None


def find_rows_by_id(worksheet: Any, row_id: str, force_refresh: bool = False) -> list[int]:
    values = worksheet_values(worksheet, force_refresh=force_refresh)
    if not values:
        return []
    try:
        id_column = values[0].index("id")
    except ValueError:
        return []
    return [
        row_number
        for row_number, row in enumerate(values[1:], start=2)
        if id_column < len(row) and str(row[id_column]).strip() == str(row_id).strip()
    ]


def update_row_by_headers(worksheet: Any, row_number: int, headers: list[str], row: dict[str, Any]) -> None:
    values = [[row.get(header, "") for header in headers]]
    end_col = gspread.utils.rowcol_to_a1(row_number, len(headers)).replace(str(row_number), "")
    worksheet.update(f"A{row_number}:{end_col}{row_number}", values, value_input_option="RAW")
    invalidate_worksheet_cache(worksheet)


def append_worksheet_row(worksheet: Any, values: list[Any]) -> None:
    append_worksheet_rows(worksheet, [values])


def worksheet_write_lock(worksheet: Any) -> threading.Lock:
    key = str(getattr(worksheet, "id", "") or getattr(worksheet, "title", "") or id(worksheet))
    with _WORKSHEET_WRITE_LOCKS_LOCK:
        return _WORKSHEET_WRITE_LOCKS.setdefault(key, threading.Lock())


def append_worksheet_rows(worksheet: Any, rows: list[list[Any]]) -> None:
    if not rows:
        return
    with worksheet_write_lock(worksheet):
        # Việc tìm dòng trống và ghi phải nằm trong cùng một khóa; nếu không hai
        # request đồng thời có thể chọn cùng một dòng và ghi đè dữ liệu của nhau.
        existing_values = worksheet.get_all_values()
        start_row = max(len(existing_values) + 1, 2)
        required_rows = start_row + len(rows) - 1
        if required_rows > worksheet.row_count:
            worksheet.add_rows(required_rows - worksheet.row_count)
        width = max(len(row) for row in rows)
        normalized_rows = [list(row) + [""] * (width - len(row)) for row in rows]
        end_column = re.sub(r"\d+$", "", gspread.utils.rowcol_to_a1(1, width))
        worksheet.update(
            f"A{start_row}:{end_column}{required_rows}",
            normalized_rows,
            value_input_option="RAW",
        )
        invalidate_worksheet_cache(worksheet)


def row_by_id(rows: list[dict[str, Any]], row_id: str) -> dict[str, Any] | None:
    return next((row for row in rows if str(row.get("id")) == str(row_id)), None)


def is_pending_reopen_status(value: Any) -> bool:
    return normalize_text(value) in {"cho duyet", "pending"}


def append_system_log(
    username: str,
    role: str,
    action: str,
    target_type: str,
    target_id: str,
    note: str = "",
    before: Any = None,
    after: Any = None,
) -> None:
    try:
        system_logs_worksheet().append_row(
            [
                make_id("LOG"),
                now_iso(),
                username,
                role,
                action,
                target_type,
                target_id,
                note,
                json.dumps(before or {}, ensure_ascii=False),
                json.dumps(after or {}, ensure_ascii=False),
            ],
            value_input_option="RAW",
        )
    except Exception:
        pass


def log_action(
    request: Request,
    action: str,
    target_type: str,
    target_id: str,
    note: str = "",
    before: Any = None,
    after: Any = None,
) -> None:
    user = current_user(request)
    display_name = str(user.get("username") or user.get("displayName") or "").strip()
    append_system_log(
        display_name,
        str(user.get("role") or ""),
        action,
        target_type,
        target_id,
        note,
        before,
        after,
    )


def customer_records() -> list[dict[str, Any]]:
    worksheet = customers_worksheet()
    records = worksheet_records(worksheet, CUSTOMER_HEADERS)
    values = worksheet.get_all_values()
    actual_headers = values[0] if values else []

    def header_key(value: Any) -> str:
        return re.sub(r"[^a-z0-9]", "", normalize_text(value))

    header_aliases = {
        "tenkhach": "tenKhach",
        "tenkhachhang": "tenKhach",
        "sodienthoai": "soDienThoai",
        "sdt": "soDienThoai",
        "socccd": "soCCCD",
        "cccd": "soCCCD",
        "diachi": "diaChi",
        "diachikhach": "diaChi",
        "diachikhachhang": "diaChi",
        "loaikhach": "loaiKhachHang",
        "loaikhachhang": "loaiKhachHang",
        "namsinh": "namSinh",
        "gioitinh": "gioiTinh",
        "nguonkhach": "nguonKhach",
        "nhanviennhap": "nhanVienNhap",
        "ngaytao": "createdAt",
        "createdat": "createdAt",
    }

    for index, row in enumerate(records, start=1):
        raw = values[index] if index < len(values) else []
        for column_index, header in enumerate(actual_headers):
            canonical = header_aliases.get(header_key(header))
            if canonical and column_index < len(raw) and not str(row.get(canonical) or "").strip():
                row[canonical] = raw[column_index]
        source = str(row.get("nguonKhach") or "")
        if "T" in source and row.get("createdAt", "") == "":
            row["createdAt"] = source
            row["nguonKhach"] = ""
            row["nhanVienNhap"] = ""
    # Một số khách dữ liệu cũ từng bị copy-on-write nhiều lần khi cache sheet
    # chưa kịp làm mới. Giữ bản nằm sau cùng (bản sửa mới nhất) cho mỗi ID.
    seen_customer_ids: set[str] = set()
    deduplicated_records: list[dict[str, Any]] = []
    for row in reversed(records):
        customer_id = str(row.get("id") or "").strip()
        if customer_id and customer_id in seen_customer_ids:
            continue
        if customer_id:
            seen_customer_ids.add(customer_id)
        deduplicated_records.append(row)
    records = list(reversed(deduplicated_records))
    id_column = actual_headers.index("id") if "id" in actual_headers else 0
    # Kể cả dòng hiện hành đã xóa cũng phải chặn bản cùng ID trong kho dữ liệu
    # cũ; nếu chỉ xét records (đã lọc deletedAt), khách cũ sẽ xuất hiện trở lại.
    current_ids = {
        str(row[id_column]).strip()
        for row in values[1:]
        if id_column < len(row) and str(row[id_column]).strip()
    }
    archived = legacy_worksheet_records(CUSTOMERS_SHEET_NAME, CUSTOMER_HEADERS)
    combined = records + [
        row for row in archived
        if not str(row.get("id") or "").strip() or str(row.get("id") or "").strip() not in current_ids
    ]
    for row in combined:
        normalized_phone = normalize_phone(row.get("soDienThoai"))
        if normalized_phone:
            row["soDienThoai"] = normalized_phone
    return combined


def find_customer_by_phone(rows: list[dict[str, Any]], phone: str) -> dict[str, Any] | None:
    normalized = normalize_phone(phone)
    if not normalized:
        return None
    return next((row for row in rows if normalize_phone(row.get("soDienThoai")) == normalized), None)


def next_customer_id(rows: list[dict[str, Any]]) -> str:
    highest = 0
    for row in rows:
        match = re.fullmatch(r"DX(\d+)", str(row.get("id") or "").strip(), flags=re.IGNORECASE)
        if match:
            highest = max(highest, int(match.group(1)))
    return f"DX{highest + 1:02d}"


def create_customer_from_order(payload: OrderInput) -> dict[str, Any]:
    phone = validate_customer_phone(payload.soDienThoai, "Số điện thoại khách hàng")
    if not payload.tenKhach.strip():
        raise HTTPException(status_code=422, detail="Vui lòng nhập tên khách hàng.")
    required_customer_fields = {
        "loại khách": payload.loaiKhachHang,
        "giới tính": payload.gioiTinh,
        "nguồn khách": payload.nguonKhach,
        "nhân viên nhập": payload.nhanVienNhap,
    }
    missing_fields = [label for label, value in required_customer_fields.items() if not str(value or "").strip()]
    if missing_fields:
        raise HTTPException(status_code=422, detail=f"Vui lòng nhập đầy đủ thông tin khách hàng: {', '.join(missing_fields)}.")

    customer_id = next_customer_id(customer_records())
    row = {
        "id": customer_id,
        "tenKhach": payload.tenKhach,
        "soDienThoai": phone,
        "soCCCD": payload.soCCCD,
        "diaChi": payload.diaChi,
        "loaiKhachHang": payload.loaiKhachHang,
        "namSinh": payload.namSinh,
        "gioiTinh": payload.gioiTinh,
        "nguonKhach": payload.nguonKhach,
        "nhanVienNhap": payload.nhanVienNhap,
        "createdAt": now_iso(),
    }
    customer_worksheet = customers_worksheet()
    append_worksheet_row(customer_worksheet, [row.get(header, "") for header in CUSTOMER_HEADERS])
    return row


def roster_driver_name(value: Any) -> str:
    text = str(value or "").strip()
    return text.split(" - ", 1)[0].strip()


def roster_driver_code(value: Any) -> str:
    text = str(value or "").strip()
    return text.split(" - ", 1)[1].strip() if " - " in text else ""


def roster_driver_text(row: dict[str, Any]) -> str:
    return str(row.get("hoTenMSNVLaiXe") or row.get("hoTenNhanVienLaiXe") or "").strip()


def parse_roster_date(value: Any) -> datetime:
    text = str(value or "").strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return datetime.min


def date_key(value: datetime) -> str:
    return value.strftime("%Y-%m-%d")


def roster_date_key(row: dict[str, Any]) -> str:
    parsed = parse_roster_date(row.get("thoiGianTao"))
    return "" if parsed == datetime.min else parsed.strftime("%Y-%m-%d")


def roster_is_on_shift(row: dict[str, Any]) -> bool:
    return "len ca" in normalize_text(row.get("trangThaiLenXuongCa"))


def better_roster_row(current: dict[str, Any] | None, candidate: dict[str, Any]) -> dict[str, Any]:
    if current is None:
        return candidate
    current_on_shift = roster_is_on_shift(current)
    candidate_on_shift = roster_is_on_shift(candidate)
    if candidate_on_shift != current_on_shift:
        return candidate if candidate_on_shift else current
    return candidate if parse_roster_date(candidate.get("thoiGianTao")) >= parse_roster_date(current.get("thoiGianTao")) else current


def roster_rows() -> list[dict[str, Any]]:
    return read_public_sheet(ROSTER_SHEET_NAME)


def roster_vehicle_by_plate(plate: str, shift_date: str = "") -> dict[str, Any] | None:
    normalized_plate = normalize_text(plate)
    best_row: dict[str, Any] | None = None
    for row in roster_rows():
        if shift_date and roster_date_key(row) != shift_date:
            continue
        if normalize_text(row.get("bienKiemSoat")) == normalized_plate and roster_is_on_shift(row) and roster_driver_text(row):
            best_row = better_roster_row(best_row, row)
    return best_row


def franchise_vehicle_records() -> list[dict[str, Any]]:
    return worksheet_records(franchise_vehicles_worksheet(), FRANCHISE_VEHICLE_HEADERS)


def franchise_vehicle_by_plate(plate: str) -> dict[str, Any] | None:
    normalized_plate = normalize_text(plate)
    for row in franchise_vehicle_records():
        if normalize_text(row.get("bienKiemSoat")) == normalized_plate and "ngung" not in normalize_text(row.get("trangThai")):
            return row
    return None


@app.get("/")
def index() -> FileResponse:
    return FileResponse(PUBLIC_DIR / "index.html")


@app.get("/api/app-version")
def api_app_version() -> dict[str, str]:
    tracked_files = [Path(__file__), PUBLIC_DIR / "index.html", PUBLIC_DIR / "app.js", PUBLIC_DIR / "styles.css"]
    version = max((path.stat().st_mtime_ns for path in tracked_files if path.exists()), default=0)
    return {"version": str(version)}


@app.get("/api/me")
def api_me(request: Request) -> dict[str, Any]:
    user = current_user(request)
    return {
        "user": public_user(user),
        "permissions": user_permissions(str(user.get("role") or ""), user.get("extraPermissions")),
        "roles": ROLE_LABELS,
    }


@app.post("/api/login")
def api_login(payload: LoginInput) -> Response:
    user = authenticate_user(payload.username, payload.password)
    if not user:
        raise HTTPException(status_code=401, detail="Tên đăng nhập hoặc mật khẩu không đúng.")
    token = secrets.token_urlsafe(32)
    session_user = public_user(user)
    SESSIONS[token] = {"user": session_user, "lastSeen": datetime.now(timezone.utc)}
    response = JSONResponse(
        {
            "token": token,
            "user": session_user,
            "permissions": user_permissions(str(session_user.get("role") or ""), session_user.get("extraPermissions")),
            "roles": ROLE_LABELS,
        }
    )
    response.set_cookie(
        "dx_session",
        token,
        httponly=True,
        samesite="lax",
        max_age=7 * 24 * 60 * 60,
    )
    return response


@app.post("/api/logout")
def api_logout(request: Request) -> Response:
    token = request.cookies.get("dx_session", "")
    auth_header = request.headers.get("authorization", "")
    if auth_header.lower().startswith("bearer "):
        token = auth_header.split(" ", 1)[1].strip()
    if token:
        SESSIONS.pop(token, None)
    response = JSONResponse({"ok": True})
    response.delete_cookie("dx_session")
    return response


@app.post("/api/me/password")
def change_my_password(payload: ChangePasswordInput, request: Request) -> dict[str, Any]:
    user = current_user(request)
    if str(user.get("id") or "") == "default-admin":
        if not compare_secret(payload.currentPassword, DEFAULT_ADMIN_PASSWORD):
            raise HTTPException(status_code=401, detail="Mật khẩu hiện tại không đúng.")
        raise HTTPException(status_code=409, detail="Tài khoản admin mặc định cần đổi trong cấu hình hệ thống.")
    worksheet = users_worksheet()
    rows = user_rows_cached(force=True)
    user_id = str(user.get("id") or "")
    row = row_by_id(rows, user_id)
    if row is None:
        raise HTTPException(status_code=404, detail="Không tìm thấy tài khoản.")
    if not compare_secret(row.get("password") or "", payload.currentPassword):
        raise HTTPException(status_code=401, detail="Mật khẩu hiện tại không đúng.")
    row_number = find_row_by_id(worksheet, user_id)
    if row_number is None:
        raise HTTPException(status_code=404, detail="Không tìm thấy tài khoản.")
    before = public_user(row)
    updated = dict(row)
    updated["password"] = payload.newPassword
    update_row_by_headers(worksheet, row_number, USER_HEADERS, updated)
    _USER_CACHE["expires"] = datetime.now(timezone.utc) + timedelta(seconds=USER_CACHE_TTL_SECONDS)
    _USER_CACHE["rows"] = [updated if str(item.get("id")) == user_id else item for item in rows]
    log_action(request, "change_password", "user", user_id, before=before, after=public_user(updated))
    for token, session in list(SESSIONS.items()):
        if str((session.get("user") or {}).get("id") or "") == user_id:
            SESSIONS.pop(token, None)
    return {"ok": True}


@app.get("/api/users")
def list_users(request: Request) -> dict[str, Any]:
    require_admin(request)
    rows = user_rows_cached()
    return {"sheetName": USERS_SHEET_NAME, "rows": [public_user(row) for row in rows], "roles": ROLE_LABELS}


@app.post("/api/users")
def create_user(payload: UserInput, request: Request) -> dict[str, Any]:
    admin = require_admin(request)
    role = payload.role.strip()
    if role not in ROLE_LABELS:
        raise HTTPException(status_code=422, detail="Vai trò không hợp lệ.")
    username = payload.username.strip()
    rows = user_rows_cached()
    if any(normalize_text(row.get("username")) == normalize_text(username) for row in rows):
        raise HTTPException(status_code=409, detail="Tên đăng nhập đã tồn tại.")
    row = {
        "id": make_id("USER"),
        "username": username,
        "password": payload.password,
        "displayName": payload.displayName.strip(),
        "role": role,
        "status": payload.status or "active",
        "createdAt": now_iso(),
        "createdBy": admin.get("username", ""),
        "extraPermissions": ",".join(payload.extraPermissions),
    }
    try:
        append_worksheet_row(users_worksheet(), [row.get(header, "") for header in USER_HEADERS])
    except Exception as exc:
        if is_google_quota_error(exc):
            raise google_quota_exception() from exc
        raise
    _USER_CACHE["expires"] = datetime.now(timezone.utc) + timedelta(seconds=USER_CACHE_TTL_SECONDS)
    _USER_CACHE["rows"] = [*rows, row]
    log_action(request, "create_user", "user", username, note=f"Tạo tài khoản {username}", after=public_user(row))
    return {"ok": True, "id": row["id"]}


@app.put("/api/users/{user_id}")
def update_user(user_id: str, payload: UserUpdateInput, request: Request) -> dict[str, Any]:
    admin = require_admin(request)
    worksheet = users_worksheet()
    rows = user_rows_cached(force=True)
    current = row_by_id(rows, user_id)
    if current is None:
        raise HTTPException(status_code=404, detail="Không tìm thấy tài khoản.")

    username = payload.username.strip()
    role = payload.role.strip()
    status = payload.status.strip()
    if role not in ROLE_LABELS:
        raise HTTPException(status_code=422, detail="Vai trò không hợp lệ.")
    if any(
        str(row.get("id") or "") != str(user_id)
        and normalize_text(row.get("username")) == normalize_text(username)
        for row in rows
    ):
        raise HTTPException(status_code=409, detail="Tên đăng nhập đã tồn tại.")

    is_current_account = str(admin.get("id") or "") == str(user_id)
    if is_current_account and (status != "active" or role != "admin"):
        raise HTTPException(
            status_code=409,
            detail="Không thể ngừng kích hoạt hoặc bỏ quyền Admin của tài khoản đang đăng nhập.",
        )

    was_active_admin = (
        str(current.get("role") or "") == "admin"
        and normalize_text(current.get("status")) in {"", "active", "dang hoat dong", "hoat dong"}
    )
    if was_active_admin and (role != "admin" or status != "active"):
        other_active_admins = [
            row
            for row in rows
            if str(row.get("id") or "") != str(user_id)
            and str(row.get("role") or "") == "admin"
            and normalize_text(row.get("status")) in {"", "active", "dang hoat dong", "hoat dong"}
        ]
        if not other_active_admins:
            raise HTTPException(status_code=409, detail="Hệ thống phải còn ít nhất một tài khoản Admin đang hoạt động.")

    row_number = find_row_by_id(worksheet, user_id)
    if row_number is None:
        raise HTTPException(status_code=404, detail="Không tìm thấy tài khoản.")
    updated = {
        **current,
        "username": username,
        "displayName": payload.displayName.strip(),
        "role": role,
        "status": status,
        "extraPermissions": ",".join(payload.extraPermissions),
    }
    update_row_by_headers(worksheet, row_number, USER_HEADERS, updated)
    _USER_CACHE["expires"] = datetime.now(timezone.utc) + timedelta(seconds=USER_CACHE_TTL_SECONDS)
    _USER_CACHE["rows"] = [updated if str(item.get("id")) == str(user_id) else item for item in rows]

    for token, session in list(SESSIONS.items()):
        session_user = session.get("user") or {}
        if str(session_user.get("id") or "") != str(user_id):
            continue
        if status != "active":
            SESSIONS.pop(token, None)
        else:
            session["user"] = public_user(updated)

    log_action(
        request,
        "update_user",
        "user",
        user_id,
        note=f"Cập nhật tài khoản {username}",
        before=public_user(current),
        after=public_user(updated),
    )
    return {"ok": True, "id": user_id}


@app.post("/api/users/{user_id}/reset-password")
def reset_user_password(user_id: str, payload: ResetPasswordInput, request: Request) -> dict[str, Any]:
    require_admin(request)
    worksheet = users_worksheet()
    rows = user_rows_cached(force=True)
    row = row_by_id(rows, user_id)
    if row is None:
        raise HTTPException(status_code=404, detail="Không tìm thấy tài khoản.")
    row_number = find_row_by_id(worksheet, user_id)
    if row_number is None:
        raise HTTPException(status_code=404, detail="Không tìm thấy tài khoản.")
    before = public_user(row)
    updated = dict(row)
    updated["password"] = payload.newPassword
    update_row_by_headers(worksheet, row_number, USER_HEADERS, updated)
    _USER_CACHE["expires"] = datetime.now(timezone.utc) + timedelta(seconds=USER_CACHE_TTL_SECONDS)
    _USER_CACHE["rows"] = [updated if str(item.get("id")) == user_id else item for item in rows]
    for token, session in list(SESSIONS.items()):
        if str((session.get("user") or {}).get("id") or "") == str(user_id):
            SESSIONS.pop(token, None)
    log_action(request, "reset_password", "user", user_id, note=f"Reset mật khẩu cho {row.get('username') or user_id}", before=before, after=public_user(updated))
    return {"ok": True, "id": user_id}


def system_catalog_rows(seed_defaults: bool = True) -> list[dict[str, Any]]:
    worksheet = system_catalogs_worksheet()
    rows = worksheet_records(worksheet, SYSTEM_CATALOG_HEADERS)
    if not seed_defaults:
        return rows
    created_at = now_iso()
    existing_types = {str(row.get("loaiDanhMuc") or "") for row in rows}
    defaults: list[dict[str, Any]] = list(rows)
    for catalog_type, values in DEFAULT_SYSTEM_CATALOGS.items():
        if catalog_type in existing_types:
            continue
        for index, value in enumerate(values, start=1):
            row = {
                "id": make_id("DM"),
                "loaiDanhMuc": catalog_type,
                "giaTri": value,
                "thuTu": index,
                "trangThai": "active",
                "createdAt": created_at,
                "createdBy": "system",
                "deletedAt": "",
                "deletedBy": "",
            }
            append_worksheet_row(worksheet, [row.get(header, "") for header in SYSTEM_CATALOG_HEADERS])
            defaults.append(row)
    return defaults


@app.get("/api/system-catalogs")
def list_system_catalogs(request: Request) -> dict[str, Any]:
    current_user(request)
    rows = system_catalog_rows()
    active_rows = [
        row for row in rows
        if normalize_text(row.get("trangThai")) in {"", "active", "dang hoat dong", "hoat dong"}
        and not str(row.get("deletedAt") or "").strip()
    ]
    def catalog_order(row: dict[str, Any]) -> tuple[str, float]:
        try:
            order = float(row.get("thuTu") or 9999)
        except (TypeError, ValueError):
            order = 9999
        return str(row.get("loaiDanhMuc") or ""), order

    active_rows.sort(key=catalog_order)
    return {"sheetName": SYSTEM_CATALOGS_SHEET_NAME, "types": SYSTEM_CATALOG_TYPES, "rows": active_rows}


@app.post("/api/system-catalogs")
def create_system_catalog(payload: SystemCatalogInput, request: Request) -> dict[str, Any]:
    user = current_user(request)
    if not has_action(user, "manage_system_catalogs"):
        raise HTTPException(status_code=403, detail="Bạn không có quyền quản trị danh mục hệ thống.")
    catalog_type = payload.loaiDanhMuc.strip()
    value = payload.giaTri.strip()
    if catalog_type not in SYSTEM_CATALOG_TYPES:
        raise HTTPException(status_code=422, detail="Loại danh mục không hợp lệ.")
    rows = system_catalog_rows()
    active_same_type = [
        row for row in rows
        if str(row.get("loaiDanhMuc") or "") == catalog_type
        and normalize_text(row.get("trangThai")) in {"", "active", "dang hoat dong", "hoat dong"}
        and not str(row.get("deletedAt") or "").strip()
    ]
    if any(normalize_text(row.get("giaTri")) == normalize_text(value) for row in active_same_type):
        raise HTTPException(status_code=409, detail="Giá trị này đã có trong danh mục.")
    row = {
        "id": make_id("DM"),
        "loaiDanhMuc": catalog_type,
        "giaTri": value,
        "thuTu": len(active_same_type) + 1,
        "trangThai": "active",
        "createdAt": now_iso(),
        "createdBy": user.get("username", ""),
        "deletedAt": "",
        "deletedBy": "",
    }
    append_worksheet_row(system_catalogs_worksheet(), [row.get(header, "") for header in SYSTEM_CATALOG_HEADERS])
    log_action(request, "create_system_catalog", "system_catalog", row["id"], note=f"Thêm {SYSTEM_CATALOG_TYPES[catalog_type]}: {value}", after=row)
    return {"ok": True, "id": row["id"]}


@app.delete("/api/system-catalogs/{catalog_id}")
def delete_system_catalog(catalog_id: str, request: Request) -> dict[str, Any]:
    user = current_user(request)
    if not has_action(user, "manage_system_catalogs"):
        raise HTTPException(status_code=403, detail="Bạn không có quyền quản trị danh mục hệ thống.")
    worksheet = system_catalogs_worksheet()
    rows = system_catalog_rows(seed_defaults=False)
    current = row_by_id(rows, catalog_id)
    if current is None or str(current.get("deletedAt") or "").strip():
        raise HTTPException(status_code=404, detail="Không tìm thấy giá trị danh mục.")
    row_number = find_row_by_id(worksheet, catalog_id)
    if row_number is None:
        raise HTTPException(status_code=404, detail="Không tìm thấy giá trị danh mục.")
    updated = {
        **current,
        "trangThai": "inactive",
        "deletedAt": now_iso(),
        "deletedBy": user.get("username", ""),
    }
    update_row_by_headers(worksheet, row_number, SYSTEM_CATALOG_HEADERS, updated)
    log_action(request, "delete_system_catalog", "system_catalog", catalog_id, note=f"Xóa {current.get('giaTri') or ''}", before=current, after=updated)
    return {"ok": True, "id": catalog_id}


@app.get("/api/reopen-requests")
def list_reopen_requests(request: Request) -> dict[str, Any]:
    user = current_user(request)
    rows = worksheet_records(reopen_requests_worksheet(), REOPEN_REQUEST_HEADERS)
    if str(user.get("role") or "") != "admin":
        rows = [row for row in rows if str(row.get("requestedBy") or "") == str(user.get("username") or "")]
    return {"sheetName": REOPEN_REQUESTS_SHEET_NAME, "rows": rows}


@app.post("/api/orders/{order_id}/reopen-requests")
def create_reopen_request(order_id: str, payload: ReopenRequestInput, request: Request) -> dict[str, Any]:
    user = current_user(request)
    if str(user.get("role") or "") != "cskh" or not has_action(user, "request_reopen"):
        raise HTTPException(status_code=403, detail="Bạn không có quyền gửi yêu cầu mở lại đơn.")
    order = row_by_id(all_order_records(), order_id)
    if order is None:
        raise HTTPException(status_code=404, detail="Không tìm thấy đơn hàng.")
    if not order_is_done(order):
        raise HTTPException(status_code=409, detail="Chỉ đơn đã hoàn thành mới cần yêu cầu mở lại.")
    requests = worksheet_records(reopen_requests_worksheet(), REOPEN_REQUEST_HEADERS)
    if any(str(row.get("orderId") or "") == order_id and is_pending_reopen_status(row.get("status")) for row in requests):
        raise HTTPException(status_code=409, detail="Đơn này đã có yêu cầu chờ admin duyệt.")
    row = {
        "id": make_id("REQ"),
        "orderId": order_id,
        "orderCode": order.get("id", ""),
        "requestedBy": user.get("username", ""),
        "requestedRole": user.get("role", ""),
        "reason": payload.reason.strip(),
        "status": "Chờ duyệt",
        "adminNote": "",
        "createdAt": now_iso(),
        "reviewedAt": "",
        "reviewedBy": "",
    }
    append_worksheet_row(reopen_requests_worksheet(), [row.get(header, "") for header in REOPEN_REQUEST_HEADERS])
    log_action(request, "request_reopen", "order", order_id, note=payload.reason.strip(), before=order)
    return {"ok": True, "id": row["id"]}


@app.post("/api/reopen-requests/{request_id}/approve")
def approve_reopen_request(request_id: str, payload: ReopenReviewInput, request: Request) -> dict[str, Any]:
    admin = require_admin(request)
    request_worksheet = reopen_requests_worksheet()
    request_row_number = find_row_by_id(request_worksheet, request_id)
    if request_row_number is None:
        raise HTTPException(status_code=404, detail="Không tìm thấy yêu cầu.")
    reopen_row = row_by_id(worksheet_records(request_worksheet, REOPEN_REQUEST_HEADERS), request_id)
    if reopen_row is None or not is_pending_reopen_status(reopen_row.get("status")):
        raise HTTPException(status_code=409, detail="Yêu cầu này không còn ở trạng thái chờ duyệt.")

    order_id = str(reopen_row.get("orderId") or "")
    order_worksheet = orders_worksheet()
    order_row_number = find_row_by_id(order_worksheet, order_id)
    if order_row_number is None:
        raise HTTPException(status_code=404, detail="Không tìm thấy đơn hàng.")
    order = row_by_id(worksheet_records(order_worksheet, ORDER_HEADERS), order_id)
    updated_order = dict(order or {})
    updated_order["trangThai"] = "Chưa hoàn thành"
    updated_order["ngayGioHoanThanh"] = ""
    update_row_by_headers(order_worksheet, order_row_number, ORDER_HEADERS, updated_order)

    reopen_row["status"] = "Đã duyệt"
    reopen_row["adminNote"] = payload.adminNote
    reopen_row["reviewedAt"] = now_iso()
    reopen_row["reviewedBy"] = admin.get("displayName") or admin.get("username", "")
    update_row_by_headers(request_worksheet, request_row_number, REOPEN_REQUEST_HEADERS, reopen_row)
    log_action(request, "approve_reopen", "order", order_id, note=payload.adminNote, before=order, after=updated_order)
    return {"ok": True, "id": request_id}


@app.post("/api/reopen-requests/{request_id}/reject")
def reject_reopen_request(request_id: str, payload: ReopenReviewInput, request: Request) -> dict[str, Any]:
    admin = require_admin(request)
    worksheet = reopen_requests_worksheet()
    row_number = find_row_by_id(worksheet, request_id)
    if row_number is None:
        raise HTTPException(status_code=404, detail="Không tìm thấy yêu cầu.")
    reopen_row = row_by_id(worksheet_records(worksheet, REOPEN_REQUEST_HEADERS), request_id)
    if reopen_row is None or not is_pending_reopen_status(reopen_row.get("status")):
        raise HTTPException(status_code=409, detail="Yêu cầu này không còn ở trạng thái chờ duyệt.")
    reopen_row["status"] = "Từ chối"
    reopen_row["adminNote"] = payload.adminNote
    reopen_row["reviewedAt"] = now_iso()
    reopen_row["reviewedBy"] = admin.get("displayName") or admin.get("username", "")
    update_row_by_headers(worksheet, row_number, REOPEN_REQUEST_HEADERS, reopen_row)
    log_action(request, "reject_reopen", "order", str(reopen_row.get("orderId") or ""), note=payload.adminNote, before=reopen_row)
    return {"ok": True, "id": request_id}


@app.get("/api/logs")
def list_logs(request: Request) -> dict[str, Any]:
    require_admin(request)
    rows = all_system_log_records()

    def comparable_log_value(value: Any) -> Any:
        if value is None or value == "":
            return ""
        if isinstance(value, bool):
            return ("boolean", value)
        if isinstance(value, (int, float)):
            return ("number", float(value))
        if isinstance(value, str):
            text = value.strip()
            if re.fullmatch(r"-?(?:0|[1-9]\d*)(?:\.\d+)?", text):
                return ("number", float(text))
            return ("string", text)
        if isinstance(value, list):
            return tuple(comparable_log_value(item) for item in value)
        if isinstance(value, dict):
            return tuple(sorted((key, comparable_log_value(item)) for key, item in value.items()))
        return str(value)

    ignored_fields = {"id", "createdAt", "updatedAt", "createdBy", "updatedBy"}
    result: list[dict[str, Any]] = []
    for source in reversed(rows[-500:]):
        row = dict(source)
        try:
            before = json.loads(str(row.get("before") or "{}"))
        except (TypeError, ValueError, json.JSONDecodeError):
            before = {}
        try:
            after = json.loads(str(row.get("after") or "{}"))
        except (TypeError, ValueError, json.JSONDecodeError):
            after = {}
        if isinstance(before, dict) and isinstance(after, dict):
            changed_keys = [
                key
                for key in dict.fromkeys([*before.keys(), *after.keys()])
                if key not in ignored_fields
                and comparable_log_value(before.get(key)) != comparable_log_value(after.get(key))
            ]
            if (
                str(row.get("action") or "").startswith("update_")
                and before
                and after
                and not changed_keys
            ):
                continue
            row["before"] = json.dumps({key: before.get(key) for key in changed_keys}, ensure_ascii=False)
            row["after"] = json.dumps({key: after.get(key) for key in changed_keys}, ensure_ascii=False)
        result.append(row)
    return {"sheetName": SYSTEM_LOGS_SHEET_NAME, "rows": result}


@app.get("/api/roster")
def list_roster() -> dict[str, Any]:
    rows = roster_rows()
    return {"sheetName": ROSTER_SHEET_NAME, "rows": rows, "fetchedAt": now_iso()}


@app.get("/api/calendar-vehicle-order")
def list_calendar_vehicle_order(request: Request) -> dict[str, Any]:
    current_user(request)
    rows = worksheet_records(calendar_vehicle_order_worksheet(), CALENDAR_VEHICLE_ORDER_HEADERS)
    def order_value(row: dict[str, Any]) -> float:
        try:
            return float(row.get("thuTu") or 999999)
        except (TypeError, ValueError):
            return 999999
    rows.sort(key=order_value)
    return {
        "sheetName": CALENDAR_VEHICLE_ORDER_SHEET_NAME,
        "rows": rows,
        "bienKiemSoat": [str(row.get("bienKiemSoat") or "").strip() for row in rows if str(row.get("bienKiemSoat") or "").strip()],
    }


@app.put("/api/calendar-vehicle-order")
def update_calendar_vehicle_order(payload: CalendarVehicleOrderInput, request: Request) -> dict[str, Any]:
    user = current_user(request)
    if "calendar" not in user_permissions(str(user.get("role") or ""), user.get("extraPermissions")).get("views", []):
        raise HTTPException(status_code=403, detail="Tài khoản không có quyền xem lịch điều xe.")
    plates: list[str] = []
    seen: set[str] = set()
    for raw_plate in payload.bienKiemSoat:
        plate = str(raw_plate or "").strip()
        key = normalize_text(plate)
        if plate and key not in seen:
            plates.append(plate)
            seen.add(key)
    worksheet = calendar_vehicle_order_worksheet()
    existing = worksheet_values(worksheet)
    if len(existing) > 1:
        worksheet.batch_clear([f"A2:D{len(existing)}"])
        invalidate_worksheet_cache(worksheet)
    now = now_iso()
    append_worksheet_rows(
        worksheet,
        [[plate, index, now, user.get("username", "")] for index, plate in enumerate(plates, start=1)],
    )
    log_action(request, "update_calendar_vehicle_order", "calendar", "vehicle_order", note="Cập nhật thứ tự xe trong lịch điều xe", after={"bienKiemSoat": plates})
    return {"ok": True, "bienKiemSoat": plates}


@app.delete("/api/calendar-vehicle-order")
def reset_calendar_vehicle_order(request: Request) -> dict[str, Any]:
    user = current_user(request)
    if "calendar" not in user_permissions(str(user.get("role") or ""), user.get("extraPermissions")).get("views", []):
        raise HTTPException(status_code=403, detail="Tài khoản không có quyền xem lịch điều xe.")
    worksheet = calendar_vehicle_order_worksheet()
    existing = worksheet_values(worksheet)
    if len(existing) > 1:
        worksheet.batch_clear([f"A2:D{len(existing)}"])
        invalidate_worksheet_cache(worksheet)
    log_action(request, "reset_calendar_vehicle_order", "calendar", "vehicle_order", note="Khôi phục thứ tự mặc định của lịch điều xe")
    return {"ok": True, "bienKiemSoat": []}


@app.get("/api/franchise-vehicles")
def list_franchise_vehicles() -> dict[str, Any]:
    return {
        "sheetName": FRANCHISE_VEHICLES_SHEET_NAME,
        "rows": franchise_vehicle_records(),
        "fetchedAt": now_iso(),
    }


@app.post("/api/franchise-vehicles")
def create_franchise_vehicle(payload: FranchiseVehicleInput) -> dict[str, Any]:
    worksheet = franchise_vehicles_worksheet()
    rows = worksheet_records(worksheet, FRANCHISE_VEHICLE_HEADERS)
    plate = normalize_franchise_plate(payload.bienKiemSoat)
    vehicle_line = payload.dongXe.strip()
    vehicle_make = payload.hieuXe.strip()
    if not vehicle_line or not vehicle_make:
        raise HTTPException(status_code=422, detail="Dòng xe và Hiệu xe là thông tin bắt buộc.")
    if any(normalize_text(row.get("bienKiemSoat")) == normalize_text(plate) for row in rows):
        raise HTTPException(status_code=409, detail="Biển số xe thương quyền đã tồn tại.")
    vehicle_id = make_id("XHQ")
    row = [
        vehicle_id,
        plate,
        vehicle_line,
        vehicle_make,
        payload.soCho,
        payload.tenChuXe,
        payload.soDienThoaiChuXe,
        payload.hoTenLaiXe,
        payload.soDienThoaiLaiXe,
        payload.diaChiLaiXe,
        payload.trangThai,
        payload.ghiChu,
        now_iso(),
    ]
    append_worksheet_row(worksheet, row)
    return {"ok": True, "id": vehicle_id}


@app.put("/api/franchise-vehicles/{vehicle_id}")
def update_franchise_vehicle(vehicle_id: str, payload: FranchiseVehicleInput, request: Request) -> dict[str, Any]:
    worksheet = franchise_vehicles_worksheet()
    rows = worksheet_records(worksheet, FRANCHISE_VEHICLE_HEADERS)
    row_number = find_row_by_id(worksheet, vehicle_id)
    plate = normalize_franchise_plate(payload.bienKiemSoat)
    vehicle_line = payload.dongXe.strip()
    vehicle_make = payload.hieuXe.strip()
    if not vehicle_line or not vehicle_make:
        raise HTTPException(status_code=422, detail="Dòng xe và Hiệu xe là thông tin bắt buộc.")
    if row_number is None:
        raise HTTPException(status_code=404, detail="Không tìm thấy xe thương quyền.")
    if any(
        str(row.get("id")) != vehicle_id and normalize_text(row.get("bienKiemSoat")) == normalize_text(plate)
        for row in rows
    ):
        raise HTTPException(status_code=409, detail="Biển số xe thương quyền đã tồn tại.")
    existing = row_by_id(rows, vehicle_id) or {}
    updated = {
        "id": vehicle_id,
        "bienKiemSoat": plate,
        "dongXe": vehicle_line,
        "hieuXe": vehicle_make,
        "soCho": payload.soCho,
        "tenChuXe": payload.tenChuXe,
        "soDienThoaiChuXe": payload.soDienThoaiChuXe,
        "hoTenLaiXe": payload.hoTenLaiXe,
        "soDienThoaiLaiXe": payload.soDienThoaiLaiXe,
        "diaChiLaiXe": payload.diaChiLaiXe,
        "trangThai": payload.trangThai,
        "ghiChu": payload.ghiChu,
        "createdAt": existing.get("createdAt") or now_iso(),
    }
    update_row_by_headers(worksheet, row_number, FRANCHISE_VEHICLE_HEADERS, updated)
    log_action(request, "update_franchise_vehicle", "franchiseVehicle", vehicle_id, before=existing, after=updated)
    return {"ok": True, "id": vehicle_id}


@app.delete("/api/franchise-vehicles/{vehicle_id}")
def delete_franchise_vehicle(vehicle_id: str, request: Request) -> dict[str, Any]:
    worksheet = franchise_vehicles_worksheet()
    row_number = find_row_by_id(worksheet, vehicle_id)
    if row_number is None:
        raise HTTPException(status_code=404, detail="Không tìm thấy xe thương quyền.")
    vehicle = row_by_id(worksheet_records(worksheet, FRANCHISE_VEHICLE_HEADERS), vehicle_id) or {}
    orders = all_order_records()
    if any(normalize_text(row.get("bienKiemSoat")) == normalize_text(vehicle.get("bienKiemSoat")) for row in orders):
        raise HTTPException(status_code=409, detail="Xe đã có đơn hàng, không thể xóa.")
    deleted = soft_delete_row(worksheet, row_number, FRANCHISE_VEHICLE_HEADERS, vehicle, request)
    log_action(request, "delete_franchise_vehicle", "franchiseVehicle", vehicle_id, before=vehicle, after=deleted)
    return {"ok": True, "id": vehicle_id}


@app.get("/api/customers")
def list_customers() -> dict[str, Any]:
    return {"sheetName": CUSTOMERS_SHEET_NAME, "rows": customer_records(), "fetchedAt": now_iso()}


@app.post("/api/customers")
def create_customer(request: Request, payload: CustomerInput) -> dict[str, Any]:
    worksheet = customers_worksheet()
    rows = worksheet_records(worksheet, CUSTOMER_HEADERS)
    phone = validate_customer_phone(payload.soDienThoai)
    if any(normalize_phone(row.get("soDienThoai")) == phone for row in rows):
        raise HTTPException(status_code=409, detail="Số điện thoại này đã được khai báo.")

    staff_name = current_user_display_name(request) or payload.nhanVienNhap.strip()
    if not staff_name:
        raise HTTPException(status_code=422, detail="Khong xac dinh duoc nhan vien nhap.")

    customer_id = next_customer_id(rows)
    worksheet.append_row(
        [
            customer_id,
        payload.tenKhach,
        phone,
        payload.soCCCD,
        payload.diaChi,
        payload.loaiKhachHang,
        payload.namSinh,
            payload.gioiTinh,
            payload.nguonKhach,
            staff_name,
            now_iso(),
        ],
        value_input_option="RAW",
    )
    invalidate_worksheet_cache(worksheet)
    return {"ok": True, "id": customer_id}


@app.put("/api/customers/{customer_id}")
def update_customer(request: Request, customer_id: str, payload: CustomerInput) -> dict[str, Any]:
    worksheet = customers_worksheet()
    rows = customer_records()
    row_numbers = find_rows_by_id(worksheet, customer_id, force_refresh=True)
    row_number = row_numbers[-1] if row_numbers else None
    phone = validate_customer_phone(payload.soDienThoai)
    if any(row.get("id") != customer_id and normalize_phone(row.get("soDienThoai")) == phone for row in rows):
        raise HTTPException(status_code=409, detail="Số điện thoại này đã được khai báo.")
    current = row_by_id(rows, customer_id) or {}
    if not current:
        raise HTTPException(status_code=404, detail="Không tìm thấy khách hàng.")
    staff_name = str(current.get("nhanVienNhap") or "").strip() or current_user_display_name(request) or payload.nhanVienNhap.strip()
    updated = {
        "id": customer_id,
        "tenKhach": payload.tenKhach,
        "soDienThoai": phone,
        "soCCCD": payload.soCCCD,
        "diaChi": payload.diaChi,
        "loaiKhachHang": payload.loaiKhachHang,
        "namSinh": payload.namSinh,
        "gioiTinh": payload.gioiTinh,
        "nguonKhach": payload.nguonKhach,
        "nhanVienNhap": staff_name,
        "createdAt": current.get("createdAt") or now_iso(),
    }
    if row_number is None:
        # Khách dữ liệu cũ được đọc từ file lưu trữ riêng. Khi người dùng sửa,
        # tạo bản ghi hiện hành cùng ID trong sheet chính (copy-on-write).
        # customer_records() ưu tiên bản ghi hiện hành theo ID nên không hiển
        # thị trùng, trong khi file lưu trữ vẫn được giữ nguyên để đối soát.
        append_worksheet_row(worksheet, [updated.get(header, "") for header in CUSTOMER_HEADERS])
    else:
        update_row_by_headers(worksheet, row_number, CUSTOMER_HEADERS, updated)
        for duplicate_row_number in row_numbers[:-1]:
            duplicate = dict(current)
            duplicate["deletedAt"] = now_iso()
            duplicate["deletedBy"] = current_user_display_name(request)
            update_row_by_headers(worksheet, duplicate_row_number, CUSTOMER_HEADERS, duplicate)
    synced_orders = sync_customer_profile_to_current_orders(customer_id, current, updated)
    log_action(
        request,
        "update_customer",
        "customer",
        customer_id,
        note=f"Đã đồng bộ {synced_orders} dòng đơn hàng/khách xe ghép hiện hành.",
        before=current,
        after=updated,
    )
    return {"ok": True, "id": customer_id, "syncedOrders": synced_orders}


@app.delete("/api/customers/{customer_id}")
def delete_customer(customer_id: str, request: Request) -> dict[str, Any]:
    worksheet = customers_worksheet()
    row_numbers = find_rows_by_id(worksheet, customer_id, force_refresh=True)
    customer = row_by_id(customer_records(), customer_id) or {}
    if not customer:
        raise HTTPException(status_code=404, detail="Không tìm thấy khách hàng.")
    orders = all_order_records()
    if any(str(row.get("khachHangId")) == customer_id for row in orders):
        raise HTTPException(status_code=409, detail="Khách hàng đã có đơn hàng, không thể xóa.")
    if not row_numbers:
        # Khách từ kho dữ liệu cũ không có dòng để sửa trong sheet hiện hành.
        # Ghi một tombstone cùng ID để ẩn bản lưu trữ mà không sửa dữ liệu gốc.
        deleted = {header: customer.get(header, "") for header in CUSTOMER_HEADERS}
        deleted["deletedAt"] = now_iso()
        deleted["deletedBy"] = current_user_display_name(request)
        append_worksheet_row(worksheet, [deleted.get(header, "") for header in CUSTOMER_HEADERS])
    else:
        deleted = dict(customer)
        for row_number in row_numbers:
            deleted = soft_delete_row(worksheet, row_number, CUSTOMER_HEADERS, customer, request)
    log_action(request, "delete_customer", "customer", customer_id, before=customer, after=deleted)
    return {"ok": True, "id": customer_id}


@app.get("/api/tours")
def list_tours() -> dict[str, Any]:
    return {"sheetName": TOURS_SHEET_NAME, "rows": tour_records(), "fetchedAt": now_iso()}


@app.get("/api/contract-pricing")
def get_contract_pricing(request: Request) -> dict[str, Any]:
    current_user(request)
    return {"sheetName": CONTRACT_PRICING_SHEET_NAME, "config": contract_pricing_config(), "fetchedAt": now_iso()}


@app.put("/api/contract-pricing")
def save_contract_pricing(payload: ContractPricingInput, request: Request) -> dict[str, Any]:
    user = current_user(request)
    if str(user.get("role") or "").strip().lower() != "admin":
        raise HTTPException(status_code=403, detail="Bạn không có quyền cập nhật bảng giá hợp đồng.")
    config = payload.model_dump()
    if not config["oneWay"] or not config["roundTrip"] or not config["waiting"]:
        raise HTTPException(status_code=422, detail="Bảng giá hợp đồng chưa đầy đủ.")
    worksheet = contract_pricing_worksheet()
    rows = worksheet_records(worksheet, CONTRACT_PRICING_HEADERS)
    row = {"id": "CONTRACT-PRICING", "configJson": json.dumps(config, ensure_ascii=False), "updatedAt": now_iso(), "updatedBy": user.get("username", "")}
    if rows:
        update_row_by_headers(worksheet, 2, CONTRACT_PRICING_HEADERS, row)
    else:
        append_worksheet_row(worksheet, [row.get(header, "") for header in CONTRACT_PRICING_HEADERS])
    log_action(request, "update_contract_pricing", "contract_pricing", row["id"], note="Cập nhật bảng giá hợp đồng", after=config)
    return {"ok": True, "config": config}


@app.post("/api/tours")
def create_tour(payload: TourInput) -> dict[str, Any]:
    worksheet = tours_worksheet()
    rows = tour_records()
    start, end, route = tour_payload_values(payload)
    if any(normalize_text(row.get("tuyen")) == normalize_text(route) for row in rows):
        raise HTTPException(status_code=409, detail="Tuyến này đã tồn tại.")
    tour_id = make_id("HD")
    append_worksheet_row(worksheet, [tour_id, start, end, route, payload.ghiChu, now_iso()])
    return {"ok": True, "id": tour_id}


@app.put("/api/tours/{tour_id}")
def update_tour(tour_id: str, payload: TourInput, request: Request) -> dict[str, Any]:
    worksheet = tours_worksheet()
    rows = tour_records()
    row_number = find_row_by_id(worksheet, tour_id)
    if row_number is None:
        raise HTTPException(status_code=404, detail="Không tìm thấy hợp đồng/tuyến.")
    start, end, route = tour_payload_values(payload)
    if any(row.get("id") != tour_id and normalize_text(row.get("tuyen")) == normalize_text(route) for row in rows):
        raise HTTPException(status_code=409, detail="Tuyến này đã tồn tại.")
    current = row_by_id(rows, tour_id) or {}
    updated = {
        "id": tour_id,
        "diemDi": start,
        "diemDen": end,
        "tuyen": route,
        "ghiChu": payload.ghiChu,
        "createdAt": current.get("createdAt") or now_iso(),
    }
    update_row_by_headers(worksheet, row_number, TOUR_HEADERS, updated)
    log_action(request, "update_tour", "tour", tour_id, before=current, after=updated)
    return {"ok": True, "id": tour_id}


@app.delete("/api/tours/{tour_id}")
def delete_tour(tour_id: str, request: Request) -> dict[str, Any]:
    worksheet = tours_worksheet()
    row_number = find_row_by_id(worksheet, tour_id)
    if row_number is None:
        raise HTTPException(status_code=404, detail="Không tìm thấy hợp đồng/tuyến.")
    orders = all_order_records()
    if any(str(row.get("hopDongTourId")) == tour_id for row in orders):
        raise HTTPException(status_code=409, detail="Tuyến đã có đơn hàng, không thể xóa.")
    tour = row_by_id(worksheet_records(worksheet, TOUR_HEADERS), tour_id) or {}
    deleted = soft_delete_row(worksheet, row_number, TOUR_HEADERS, tour, request)
    log_action(request, "delete_tour", "tour", tour_id, before=tour, after=deleted)
    return {"ok": True, "id": tour_id}


@app.get("/api/vouchers")
def list_vouchers() -> dict[str, Any]:
    return {"sheetName": VOUCHERS_SHEET_NAME, "rows": voucher_records(), "fetchedAt": now_iso()}


@app.get("/api/vouchers/print.pdf")
def print_vouchers_pdf(request: Request, voucherIds: str = "") -> Response:
    user = current_user(request)
    if not has_action(user, "manage_benefits"):
        raise HTTPException(status_code=403, detail="Không có quyền in voucher.")

    requested_ids = [value.strip() for value in voucherIds.split(",") if value.strip()]
    if not requested_ids:
        raise HTTPException(status_code=422, detail="Vui lòng chọn ít nhất một voucher để in.")
    if len(requested_ids) > 200:
        raise HTTPException(status_code=422, detail="Mỗi lần chỉ được in tối đa 200 voucher.")

    rows_by_id = {str(row.get("id") or ""): row for row in voucher_records()}
    selected = [rows_by_id[voucher_id] for voucher_id in requested_ids if voucher_id in rows_by_id]
    if len(selected) != len(requested_ids):
        raise HTTPException(status_code=404, detail="Có voucher không còn tồn tại hoặc đã bị xóa.")

    front_template = BASE_DIR / "Mặt trước.pdf"
    back_template = BASE_DIR / "Mặt sau.pdf"
    if not front_template.exists() or not back_template.exists():
        raise HTTPException(status_code=500, detail="Chưa cấu hình đủ mẫu mặt trước và mặt sau voucher.")

    try:
        from pypdf import PdfReader, PdfWriter
        from reportlab.pdfgen import canvas
    except ImportError as exc:  # pragma: no cover
        raise HTTPException(status_code=500, detail="Thiếu thư viện tạo PDF voucher.") from exc

    front_source = PdfReader(str(front_template)).pages[0]
    back_source = PdfReader(str(back_template)).pages[0]
    writer = PdfWriter()

    for voucher in selected:
        code = str(voucher.get("maVoucher") or "").strip().upper()
        if not code:
            continue
        expiry = parse_existing_date(voucher.get("ngayHetHan"))
        expiry_text = expiry.strftime("%d/%m/%Y") if expiry else "KHONG GIOI HAN"
        front_page = copy.deepcopy(front_source)
        back_page = copy.deepcopy(back_source)
        width = float(back_page.mediabox.width)
        height = float(back_page.mediabox.height)

        overlay_buffer = BytesIO()
        overlay = canvas.Canvas(overlay_buffer, pagesize=(width, height))
        overlay.setFillColorRGB(0.08, 0.34, 0.24)
        overlay.setFont("Helvetica-Bold", 9)
        overlay.drawCentredString(234, 93.3, expiry_text)
        box_x, box_y, box_width, box_height = 190, 70, 88, 16
        overlay.setFillColorRGB(1, 1, 1)
        overlay.roundRect(box_x, box_y, box_width, box_height, 2.5, fill=1, stroke=0)
        overlay.setFillColorRGB(0.08, 0.34, 0.24)
        overlay.setFont("Helvetica-Bold", 9.5)
        overlay.drawCentredString(box_x + box_width / 2, box_y + 4.6, code)
        overlay.save()
        overlay_buffer.seek(0)
        back_page.merge_page(PdfReader(overlay_buffer).pages[0])

        writer.add_page(front_page)
        writer.add_page(back_page)

    output = BytesIO()
    writer.write(output)
    output.seek(0)
    filename = f"voucher-in-hang-loat-{datetime.now().strftime('%Y%m%d-%H%M%S')}.pdf"
    return Response(
        content=output.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/api/vouchers")
def create_voucher(payload: VoucherInput) -> dict[str, Any]:
    worksheet = vouchers_worksheet()
    rows = worksheet_records(worksheet, VOUCHER_HEADERS)
    if any(normalize_text(row.get("maVoucher")) == normalize_text(payload.maVoucher) for row in rows):
        raise HTTPException(status_code=409, detail="Mã voucher đã tồn tại.")
    voucher_id = make_id("VC")
    worksheet.append_row(
        [
            voucher_id,
            payload.maVoucher,
            payload.tenVoucher,
            payload.loaiGiaTri,
            payload.giaTri,
            payload.ngayBatDau,
            payload.ngayHetHan,
            payload.trangThai,
            payload.ghiChu,
            now_iso(),
        ],
        value_input_option="RAW",
    )
    invalidate_worksheet_cache(worksheet)
    return {"ok": True, "id": voucher_id}


def generate_voucher_code(existing_codes: set[str]) -> str:
    alphabet = "ABCDEFGHJKLMNPQRSTUVWXYZ23456789"
    while True:
        code = "DX" + "".join(secrets.choice(alphabet) for _ in range(8))
        if normalize_text(code) not in existing_codes:
            existing_codes.add(normalize_text(code))
            return code


def generate_campaign_voucher_codes(
    rows: list[dict[str, Any]],
    campaign_id: str,
    quantity: int,
) -> list[str]:
    normalized_campaign_id = campaign_id.strip().upper()
    campaign_prefix = f"DX-{normalized_campaign_id}-"
    alphabet = "ABCDEFGHJKLMNPQRSTUVWXYZ23456789"
    existing_codes = {normalize_text(row.get("maVoucher")) for row in rows if row.get("maVoucher")}
    generated: list[str] = []
    while len(generated) < quantity:
        short_id = "".join(secrets.choice(alphabet) for _ in range(6))
        voucher_code = f"{campaign_prefix}{short_id}"
        normalized_code = normalize_text(voucher_code)
        if normalized_code in existing_codes:
            continue
        existing_codes.add(normalized_code)
        generated.append(voucher_code)
    return generated


@app.get("/api/vouchers/suggest-code")
def suggest_voucher_code() -> dict[str, Any]:
    rows = worksheet_records(vouchers_worksheet(), VOUCHER_HEADERS)
    existing_codes = {normalize_text(row.get("maVoucher")) for row in rows if row.get("maVoucher")}
    return {"maVoucher": generate_voucher_code(existing_codes)}


@app.post("/api/vouchers/batch")
def create_voucher_batch(payload: VoucherBatchInput) -> dict[str, Any]:
    worksheet = vouchers_worksheet()
    rows = worksheet_records(worksheet, VOUCHER_HEADERS)
    campaign_id = payload.idChienDich.strip().upper()
    voucher_codes = generate_campaign_voucher_codes(rows, campaign_id, payload.soLuong)
    created_at = now_iso()
    batch_note = f"ID chiến dịch: {campaign_id} - Phát hành hàng loạt: {payload.tenLoPhatHanh}"
    if payload.ghiChu.strip():
        batch_note = f"{batch_note} - {payload.ghiChu.strip()}"
    new_rows: list[list[Any]] = []
    ids: list[str] = []
    for voucher_code in voucher_codes:
        voucher_id = make_id("VC")
        ids.append(voucher_id)
        new_rows.append(
            [
                voucher_id,
                voucher_code,
                payload.tenLoPhatHanh,
                payload.loaiGiaTri,
                payload.menhGia,
                payload.ngayBatDau,
                payload.ngayHetHan,
                "Đang áp dụng",
                batch_note,
                created_at,
            ]
        )
    append_worksheet_rows(worksheet, new_rows)
    return {
        "ok": True,
        "count": len(new_rows),
        "ids": ids,
        "idChienDich": campaign_id,
        "maDauTien": new_rows[0][1],
        "maCuoiCung": new_rows[-1][1],
    }


@app.delete("/api/vouchers/campaign")
def delete_voucher_campaign(campaignName: str, request: Request) -> dict[str, Any]:
    user = current_user(request)
    if not has_action(user, "manage_benefits"):
        raise HTTPException(status_code=403, detail="Không có quyền xóa chiến dịch voucher.")

    campaign_name = voucher_campaign_name(campaignName)
    if not campaign_name:
        raise HTTPException(status_code=422, detail="Vui lòng chọn chiến dịch voucher cần xóa.")

    worksheet = vouchers_worksheet()
    values = worksheet.get_all_values()
    campaign_rows: list[tuple[int, dict[str, Any]]] = []
    for row_number, values_row in enumerate(values[1:], start=2):
        padded = values_row + [""] * max(len(VOUCHER_HEADERS) - len(values_row), 0)
        row = {header: padded[index] for index, header in enumerate(VOUCHER_HEADERS)}
        if is_deleted_row(row):
            continue
        if normalize_text(voucher_campaign_name(row.get("tenVoucher"))) == normalize_text(campaign_name):
            campaign_rows.append((row_number, row))

    if not campaign_rows:
        raise HTTPException(status_code=404, detail="Không tìm thấy voucher thuộc chiến dịch đã chọn.")

    usage = [item for item in order_benefit_records() if item.get("loaiUuDai") == "voucher"]
    direct_usage = stored_voucher_usage()
    used_vouchers = [
        row
        for _, row in campaign_rows
        if any(
            str(row.get("id") or "") == str(item.get("uuDaiId") or "")
            or normalize_text(row.get("maVoucher")) in {normalize_text(key) for key in benefit_usage_keys(item)}
            for item in usage
        )
        or normalize_text(row.get("maVoucher")) in direct_usage
    ]
    if used_vouchers:
        raise HTTPException(
            status_code=409,
            detail=f"Chiến dịch có {len(used_vouchers)} voucher đã được sử dụng nên không thể xóa toàn bộ.",
        )

    deleted_at = now_iso()
    deleted_by = current_user_display_name(request)
    end_column = re.sub(r"\d+$", "", gspread.utils.rowcol_to_a1(1, len(VOUCHER_HEADERS)))
    updates: list[dict[str, Any]] = []
    deleted_ids: list[str] = []
    for row_number, row in campaign_rows:
        deleted = dict(row)
        deleted["deletedAt"] = deleted_at
        deleted["deletedBy"] = deleted_by
        deleted["trangThai"] = "Đã xóa"
        updates.append(
            {
                "range": f"A{row_number}:{end_column}{row_number}",
                "values": [[deleted.get(header, "") for header in VOUCHER_HEADERS]],
            }
        )
        deleted_ids.append(str(row.get("id") or ""))

    for start in range(0, len(updates), 200):
        worksheet.batch_update(updates[start : start + 200], value_input_option="RAW")
    invalidate_worksheet_cache(worksheet)

    log_action(
        request,
        "delete_voucher_campaign",
        "voucher_campaign",
        campaign_name,
        before={"campaignName": campaign_name, "count": len(campaign_rows)},
        after={"deletedAt": deleted_at, "deletedBy": deleted_by},
    )
    return {"ok": True, "campaignName": campaign_name, "count": len(campaign_rows), "ids": deleted_ids}


@app.put("/api/vouchers/{voucher_id}")
def update_voucher(voucher_id: str, payload: VoucherInput, request: Request) -> dict[str, Any]:
    worksheet = vouchers_worksheet()
    rows = worksheet_records(worksheet, VOUCHER_HEADERS)
    row_number = find_row_by_id(worksheet, voucher_id)
    if row_number is None:
        raise HTTPException(status_code=404, detail="Không tìm thấy voucher.")
    if any(row.get("id") != voucher_id and normalize_text(row.get("maVoucher")) == normalize_text(payload.maVoucher) for row in rows):
        raise HTTPException(status_code=409, detail="Mã voucher đã tồn tại.")
    current = row_by_id(rows, voucher_id) or {}
    updated = {
            "id": voucher_id,
            "maVoucher": payload.maVoucher,
            "tenVoucher": payload.tenVoucher,
            "loaiGiaTri": payload.loaiGiaTri,
            "giaTri": payload.giaTri,
            "ngayBatDau": payload.ngayBatDau,
            "ngayHetHan": payload.ngayHetHan,
            "trangThai": payload.trangThai,
            "ghiChu": payload.ghiChu,
            "createdAt": current.get("createdAt") or now_iso(),
        }
    update_row_by_headers(worksheet, row_number, VOUCHER_HEADERS, updated)
    log_action(request, "update_voucher", "voucher", voucher_id, before=current, after=updated)
    return {"ok": True, "id": voucher_id}


@app.delete("/api/vouchers/{voucher_id}")
def delete_voucher(voucher_id: str, request: Request) -> dict[str, Any]:
    worksheet = vouchers_worksheet()
    row_number = find_row_by_id(worksheet, voucher_id)
    if row_number is None:
        raise HTTPException(status_code=404, detail="Không tìm thấy voucher.")
    voucher = row_by_id(worksheet_records(worksheet, VOUCHER_HEADERS), voucher_id) or {}
    usage = order_benefit_records()
    voucher_keys = {
        str(voucher.get("id") or "").strip(),
        str(voucher.get("maVoucher") or "").strip(),
    }
    used_in_benefits = any(
        item.get("loaiUuDai") == "voucher" and bool(voucher_keys & benefit_usage_keys(item))
        for item in usage
    )
    used_in_orders = normalize_text(voucher.get("maVoucher")) in stored_voucher_usage()
    if used_in_benefits or used_in_orders:
        raise HTTPException(status_code=409, detail="Voucher đã được áp dụng vào đơn hàng nên không thể xóa.")
    deleted = soft_delete_row(worksheet, row_number, VOUCHER_HEADERS, voucher, request)
    log_action(request, "delete_voucher", "voucher", voucher_id, before=voucher, after=deleted)
    return {"ok": True, "id": voucher_id}


@app.get("/api/promotions")
def list_promotions() -> dict[str, Any]:
    return {"sheetName": PROMOTIONS_SHEET_NAME, "rows": promotion_records(), "fetchedAt": now_iso()}


@app.post("/api/promotions")
def create_promotion(payload: PromotionInput) -> dict[str, Any]:
    promotion_name = validate_unique_promotion_name(payload.tenChuongTrinh)
    promo_id = make_id("KM")
    worksheet = promotions_worksheet()
    worksheet.append_row(
        [
            promo_id,
            promotion_name,
            payload.loaiGiaTri,
            payload.giaTri,
            payload.ngayBatDau,
            payload.ngayHetHan,
            payload.trangThai,
            payload.ghiChu,
            now_iso(),
        ],
        value_input_option="RAW",
    )
    invalidate_worksheet_cache(worksheet)
    return {"ok": True, "id": promo_id}


@app.put("/api/promotions/{promotion_id}")
def update_promotion(promotion_id: str, payload: PromotionInput, request: Request) -> dict[str, Any]:
    worksheet = promotions_worksheet()
    rows = worksheet_records(worksheet, PROMOTION_HEADERS)
    row_number = find_row_by_id(worksheet, promotion_id)
    if row_number is None:
        raise HTTPException(status_code=404, detail="Không tìm thấy chương trình khuyến mãi.")
    promotion_name = validate_unique_promotion_name(payload.tenChuongTrinh, exclude_id=promotion_id)
    current = row_by_id(rows, promotion_id) or {}
    updated = {
            "id": promotion_id,
            "tenChuongTrinh": promotion_name,
            "loaiGiaTri": payload.loaiGiaTri,
            "giaTri": payload.giaTri,
            "ngayBatDau": payload.ngayBatDau,
            "ngayHetHan": payload.ngayHetHan,
            "trangThai": payload.trangThai,
            "ghiChu": payload.ghiChu,
            "createdAt": current.get("createdAt") or now_iso(),
        }
    update_row_by_headers(worksheet, row_number, PROMOTION_HEADERS, updated)
    log_action(request, "update_promotion", "promotion", promotion_id, before=current, after=updated)
    return {"ok": True, "id": promotion_id}


@app.delete("/api/promotions/{promotion_id}")
def delete_promotion(promotion_id: str, request: Request) -> dict[str, Any]:
    worksheet = promotions_worksheet()
    row_number = find_row_by_id(worksheet, promotion_id)
    if row_number is None:
        raise HTTPException(status_code=404, detail="Không tìm thấy chương trình khuyến mãi.")
    usage = order_benefit_records()
    if any(item.get("loaiUuDai") == "promotion" and item.get("uuDaiId") == promotion_id for item in usage):
        raise HTTPException(status_code=409, detail="Chương trình đã được áp vào đơn hàng, không thể xóa.")
    promotion = row_by_id(worksheet_records(worksheet, PROMOTION_HEADERS), promotion_id) or {}
    deleted = soft_delete_row(worksheet, row_number, PROMOTION_HEADERS, promotion, request)
    log_action(request, "delete_promotion", "promotion", promotion_id, before=promotion, after=deleted)
    return {"ok": True, "id": promotion_id}


@app.get("/api/orders")
def list_orders() -> dict[str, Any]:
    rows = all_order_records()
    benefit_usage = order_benefit_records()
    shared_by_order: dict[str, list[dict[str, Any]]] = {}
    for passenger in all_shared_ride_records():
        order_id = str(passenger.get("donHangId") or "")
        if not order_id:
            continue
        passenger["loaiKhach"] = str(passenger.get("loaiKhach") or "B2C").strip().upper()
        passenger["voucherIds"] = [
            value.strip()
            for value in str(passenger.get("voucherCodes") or passenger.get("voucherIds") or "").split(",")
            if value.strip()
        ]
        passenger["promotionIds"] = [
            value.strip() for value in str(passenger.get("promotionIds") or "").split(",") if value.strip()
        ]
        passenger["yeuCauHoaDon"] = normalize_text(passenger.get("yeuCauHoaDon")) == "co"
        passenger["congNo"] = normalize_text(passenger.get("congNo")) == "co"
        shared_by_order.setdefault(order_id, []).append(passenger)
    promotions_by_name = {
        normalize_text(item.get("tenChuongTrinh")): str(item.get("id") or "")
        for item in promotion_records()
    }
    for row in rows:
        related = [item for item in benefit_usage if str(item.get("donHangId") or "") == str(row.get("id") or "")]
        voucher_ids = [
            str(item.get("maUuDai") or item.get("uuDaiId") or "")
            for item in related
            if item.get("loaiUuDai") == "voucher"
        ]
        promotion_ids = [str(item.get("uuDaiId") or "") for item in related if item.get("loaiUuDai") == "promotion"]
        # Backward-compatible recovery for orders whose discount summary was saved
        # before the benefit-detail rows were written consistently.
        if not voucher_ids and str(row.get("voucherCodes") or "").strip():
            voucher_ids = [value.strip() for value in str(row.get("voucherCodes") or "").split(",") if value.strip()]
        if not promotion_ids and str(row.get("khuyenMai") or "").strip():
            promotion_names = [value.strip() for value in str(row.get("khuyenMai") or "").split(",") if value.strip()]
            promotion_ids = [
                promotions_by_name.get(normalize_text(name), "")
                for name in promotion_names
                if promotions_by_name.get(normalize_text(name), "")
            ]
        row["voucherIds"] = list(dict.fromkeys(voucher_ids))
        row["promotionIds"] = list(dict.fromkeys(promotion_ids))
        row["khachXeGhep"] = shared_by_order.get(str(row.get("id") or ""), [])
    return {"sheetName": ORDERS_SHEET_NAME, "rows": rows, "fetchedAt": now_iso()}


@app.get("/api/order-feedback")
def list_order_feedback() -> dict[str, Any]:
    rows = worksheet_records(order_feedback_worksheet(), ORDER_FEEDBACK_HEADERS)
    return {"sheetName": ORDER_FEEDBACK_SHEET_NAME, "rows": rows, "fetchedAt": now_iso()}


@app.get("/api/cskh-shift-reports")
def list_cskh_shift_reports(request: Request) -> dict[str, Any]:
    rows = cskh_shift_report_records()
    return {"sheetName": CSKH_SHIFT_REPORTS_SHEET_NAME, "rows": rows, "fetchedAt": now_iso()}


@app.post("/api/cskh-shift-reports")
def save_cskh_shift_report(payload: CskhShiftReportInput, request: Request) -> dict[str, Any]:
    try:
        report_date = datetime.strptime(payload.ngay, "%Y-%m-%d")
    except ValueError as exc:
        raise HTTPException(status_code=422, detail="Ngày báo cáo không hợp lệ.") from exc

    employee = current_user_display_name(request)
    shift_label = f"Ca {payload.caLamViec}"
    shift_time = "07:00 - 14:30" if payload.caLamViec == 1 else "14:30 - 22:00"
    row = {
        "Ngày": report_date.strftime("%d/%m/%Y"),
        "Nhân Viên Trực": employee,
        "Ca Làm Việc": shift_label,
        "Thời Gian": shift_time,
        "Số lượng tin nhắn meta": payload.soLuongTinNhanMeta,
        "Số lượng khách phản hồi": payload.soLuongKhachPhanHoi,
        "Số lượng cuộc gọi": payload.soLuongCuocGoi,
        "Số lượng chat zalo": payload.soLuongChatZalo,
        "Số lượng khách từ website": payload.soLuongKhachTuWebsite,
        "Số lượng khách từ Email": payload.soLuongKhachTuEmail,
        "Số lượng tin nhắn khách vãng lai": payload.soLuongTinNhanKhachVangLai,
        "Số lượng khách phản hồi từ tiktok": payload.soLuongKhachPhanHoiTuTiktok,
        "Số lượng đơn chốt từ tiktok": payload.soLuongDonChotTuTiktok,
        "Tổng số lượng đơn chốt": payload.tongSoLuongDonChot,
    }
    worksheet = cskh_shift_reports_worksheet()
    values = worksheet_values(worksheet, force_refresh=True)
    existing_row_number = None
    for row_number, values_row in enumerate(values[1:], start=2):
        padded = values_row + [""] * max(len(CSKH_SHIFT_REPORT_HEADERS) - len(values_row), 0)
        current = dict(zip(CSKH_SHIFT_REPORT_HEADERS, padded))
        if (
            normalize_text(current.get("Trạng Thái")) != "da xoa"
            and
            str(current.get("Ngày") or "").strip() == row["Ngày"]
            and str(current.get("Nhân Viên Trực") or "").strip() == employee
            and str(current.get("Ca Làm Việc") or "").strip() == shift_label
        ):
            existing_row_number = row_number
            break
    if existing_row_number:
        raise HTTPException(
            status_code=409,
            detail=f"Báo cáo ngày {row['Ngày']} - {shift_label} của {employee} đã được khai báo.",
        )
    append_worksheet_row(worksheet, [row.get(header, "") for header in CSKH_SHIFT_REPORT_HEADERS])
    return {"ok": True, "row": row, "updated": False}


@app.post("/api/cskh-shift-reports/delete")
def delete_cskh_shift_report(payload: CskhShiftReportDeleteInput, request: Request) -> dict[str, Any]:
    try:
        selected_date = datetime.strptime(payload.ngay, "%Y-%m-%d").strftime("%d/%m/%Y")
    except ValueError as exc:
        raise HTTPException(status_code=422, detail="Ngày báo cáo không hợp lệ.") from exc
    user = current_user(request)
    current_employee = current_user_display_name(request)
    requested_employee = payload.nhanVienTruc.strip()
    if (
        user.get("role") != "admin"
        and requested_employee
        and normalize_text(requested_employee) != normalize_text(current_employee)
    ):
        raise HTTPException(status_code=403, detail="Bạn chỉ được xóa báo cáo ca do chính mình tạo.")
    target_employee = requested_employee if user.get("role") == "admin" else current_employee
    if not target_employee:
        target_employee = current_employee
    shift_label = f"Ca {payload.caLamViec}"
    worksheet = cskh_shift_reports_worksheet()
    values = worksheet_values(worksheet, force_refresh=True)
    for row_number, values_row in enumerate(values[1:], start=2):
        padded = values_row + [""] * max(len(CSKH_SHIFT_REPORT_HEADERS) - len(values_row), 0)
        row = dict(zip(CSKH_SHIFT_REPORT_HEADERS, padded))
        if (
            normalize_text(row.get("Trạng Thái")) != "da xoa"
            and str(row.get("Ngày") or "").strip() == selected_date
            and str(row.get("Nhân Viên Trực") or "").strip() == target_employee
            and str(row.get("Ca Làm Việc") or "").strip() == shift_label
        ):
            deleted = dict(row)
            deleted["Trạng Thái"] = "Đã xóa"
            deleted["Ngày Xóa"] = now_iso()
            deleted["Nhân Viên Xóa"] = current_employee
            update_row_by_headers(worksheet, row_number, CSKH_SHIFT_REPORT_HEADERS, deleted)
            target_id = f"{selected_date}-{shift_label}-{target_employee}"
            log_action(
                request,
                "delete_cskh_shift_report",
                "cskh_shift_report",
                target_id,
                note=f"Xóa mềm báo cáo {selected_date} - {shift_label} - {target_employee}",
                before=row,
                after=deleted,
            )
            return {"ok": True}
    raise HTTPException(status_code=404, detail="Không tìm thấy báo cáo ca cần xóa.")


@app.get("/api/cskh-shift-reports/export.xlsx")
def export_cskh_shift_reports(
    request: Request,
    tuNgay: str = "",
    denNgay: str = "",
    ngay: str = "",
) -> Response:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover
        raise HTTPException(status_code=500, detail="Thiếu thư viện xuất Excel.") from exc

    start_date, end_date = selected_report_range(tuNgay, denNgay, ngay)
    rows = [
        row for row in cskh_shift_report_records()
        if (report_date := parse_existing_date(row.get("Ngày"))) is not None
        and start_date.date() <= report_date.date() <= end_date.date()
    ]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Bao cao ca CSKH"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A4"
    export_headers = [
        header for header in CSKH_SHIFT_REPORT_HEADERS[:14]
        if header != "Số lượng đơn chốt từ tiktok"
    ]
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(export_headers) + 1)
    range_label = (
        start_date.strftime("%d/%m/%Y")
        if start_date.date() == end_date.date()
        else f"TỪ {start_date.strftime('%d/%m/%Y')} ĐẾN {end_date.strftime('%d/%m/%Y')}"
    )
    title = sheet.cell(1, 1, f"BÁO CÁO CA CSKH - {range_label}")
    title.font = Font(bold=True, size=15, color="FFFFFF")
    title.fill = PatternFill("solid", fgColor="0F172A")
    title.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 28
    display_headers = [
        "Tổng số lượng đơn chốt (B2C)" if header == "Tổng số lượng đơn chốt" else header
        for header in export_headers
    ]
    headers = ["STT", *display_headers]
    header_fill = PatternFill("solid", fgColor="DDEBFF")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(3, column, header)
        cell.font = Font(bold=True, color="0F172A")
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for index, row in enumerate(rows, start=1):
        values = [index, *[row.get(header, "") for header in export_headers]]
        for column, value in enumerate(values, start=1):
            if column >= 6:
                value = money_value(value)
            cell = sheet.cell(index + 3, column, value)
            cell.border = border
            cell.alignment = Alignment(vertical="center", horizontal="center" if column != 3 else "left")
            if column >= 6:
                cell.number_format = "#,##0"
    total_row = len(rows) + 4
    sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=5)
    sheet.cell(total_row, 1, "TỔNG CỘNG").alignment = Alignment(horizontal="center", vertical="center")
    for column in range(6, len(headers) + 1):
        sheet.cell(total_row, column, f"=SUM({get_column_letter(column)}4:{get_column_letter(column)}{total_row - 1})" if rows else 0)
        sheet.cell(total_row, column).number_format = "#,##0"
    for column in range(1, len(headers) + 1):
        cell = sheet.cell(total_row, column)
        cell.font = Font(bold=True, color="0F172A")
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    widths = [7, 14, 24, 14, 18, 21, 22, 18, 18, 21, 20, 26, 28, 22]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{max(3, len(rows) + 3)}"
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"bao-cao-ca-cskh-{start_date.strftime('%Y%m%d')}-{end_date.strftime('%Y%m%d')}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.put("/api/order-feedback/{order_id}")
def save_order_feedback(order_id: str, payload: OrderFeedbackInput, request: Request) -> dict[str, Any]:
    order = row_by_id(all_order_records(), order_id)
    if not order:
        raise HTTPException(status_code=404, detail="Không tìm thấy đơn hàng.")
    if not order_is_done(order):
        raise HTTPException(status_code=409, detail="Chỉ có thể ghi nhận phản hồi sau khi đơn hàng hoàn thành.")

    worksheet = order_feedback_worksheet()
    rows = worksheet_records(worksheet, ORDER_FEEDBACK_HEADERS)
    existing = next((row for row in rows if str(row.get("donHangId") or "") == str(order_id)), None)
    username = current_user_display_name(request) or str(current_user(request).get("username") or "")
    timestamp = now_iso()
    feedback = {
        "id": str(existing.get("id") or "") if existing else make_id("PH"),
        "donHangId": order_id,
        "maDon": order.get("id") or order_id,
        "khachHangId": order.get("khachHangId") or "",
        "tenKhach": order.get("tenKhach") or "",
        "kenhChamSoc": payload.kenhChamSoc.strip(),
        "diemDanhGia": payload.diemDanhGia,
        "noiDungPhanHoi": payload.noiDungPhanHoi.strip(),
        "hinhThucXuLy": payload.hinhThucXuLy.strip(),
        "ketQuaXuLy": payload.ketQuaXuLy.strip(),
        "chuThich": payload.chuThich.strip(),
        "createdAt": (existing.get("createdAt") or timestamp) if existing else timestamp,
        "createdBy": (existing.get("createdBy") or username) if existing else username,
        "updatedAt": timestamp,
        "updatedBy": username,
    }
    if existing:
        row_number = find_row_by_id(worksheet, str(existing.get("id") or ""))
        if row_number is None:
            raise HTTPException(status_code=404, detail="Không tìm thấy phản hồi cần cập nhật.")
        update_row_by_headers(worksheet, row_number, ORDER_FEEDBACK_HEADERS, feedback)
        action = "update_order_feedback"
    else:
        append_worksheet_row(worksheet, [feedback.get(header, "") for header in ORDER_FEEDBACK_HEADERS])
        action = "create_order_feedback"
    log_action(request, action, "order_feedback", feedback["id"], note=f"Đơn hàng {order_id}", before=existing, after=feedback)
    return {"ok": True, "row": feedback}


def order_requires_invoice(row: dict[str, Any]) -> bool:
    value = normalize_text(row.get("yeuCauHoaDon"))
    return bool(value) and value not in {"khong", "false", "0", "no"}


def invoice_status(row: dict[str, Any]) -> str:
    status = str(row.get("trangThaiHoaDon") or "").strip()
    return status if status else "Chưa xuất"


def invoice_rows() -> list[dict[str, Any]]:
    orders = all_order_records()
    groups = worksheet_records(invoice_groups_worksheet(), INVOICE_GROUP_HEADERS)
    grouped_order_ids = {
        order_id.strip()
        for group in groups
        for order_id in str(group.get("orderIds") or "").split(",")
        if order_id.strip()
    }
    orders_by_id = {str(order.get("id") or ""): order for order in orders}
    result: list[dict[str, Any]] = []
    shared_orders: dict[str, dict[str, Any]] = {}
    for order in orders:
        if "ghep" in normalize_text(order.get("loaiHopDong")):
            shared_orders[str(order.get("id") or "")] = order
        elif (
            order_requires_invoice(order)
            and order_is_done(order)
            and str(order.get("id") or "") not in grouped_order_ids
        ):
            item = dict(order)
            item["orderCode"] = item.get("id") or ""
            item["invoiceEntityType"] = "order"
            result.append(item)

    for group in groups:
        member_ids = [value.strip() for value in str(group.get("orderIds") or "").split(",") if value.strip()]
        members = [orders_by_id[value] for value in member_ids if value in orders_by_id]
        routes = list(dict.fromkeys(str(row.get("tuyen") or "").strip() for row in members if row.get("tuyen")))
        departure_dates = [row.get("ngayGioDi") for row in members if row.get("ngayGioDi")]
        result.append({
            **group,
            "orderCode": group.get("id") or "",
            "invoiceEntityType": "invoiceGroup",
            "trangThaiHoaDon": group.get("trangThai") or "Chưa xuất",
            "giaTien": group.get("tongThanhToan") or 0,
            "thucThu": group.get("tongThanhToan") or 0,
            "tuyen": " • ".join(routes),
            "diemDon": "",
            "diemTra": "",
            "ngayGioDi": min(departure_dates) if departure_dates else "",
            "yeuCauHoaDon": "Có",
            "nhomHoaDonId": group.get("id") or "",
            "soDonTrongNhom": len(member_ids),
        })

    for passenger in all_shared_ride_records():
        if not order_requires_invoice(passenger):
            continue
        order = shared_orders.get(str(passenger.get("donHangId") or ""), {})
        if not order or not order_is_done(order):
            continue
        item = dict(order)
        item.update(
            {
                "id": passenger.get("id") or "",
                "orderCode": passenger.get("donHangId") or order.get("id") or "",
                "invoiceEntityType": "sharedPassenger",
                "tenKhach": passenger.get("hoTen") or "",
                "soDienThoai": passenger.get("soDienThoai") or "",
                "soCCCD": passenger.get("soCCCD") or "",
                "diaChi": passenger.get("diaChi") or "",
                "diemDon": passenger.get("diemDon") or order.get("diemDon") or "",
                "diemTra": passenger.get("diemTra") or order.get("diemTra") or "",
                "giaTien": passenger.get("soTien") or 0,
                "giamGia": passenger.get("giamGia") or 0,
                "daCoc": passenger.get("daCoc") or 0,
                "thucThu": passenger.get("thucThu") or 0,
                "tongUuDai": passenger.get("tongUuDai") or 0,
                "voucherCodes": passenger.get("voucherCodes") or "",
                "khuyenMai": passenger.get("khuyenMai") or "",
                "yeuCauHoaDon": passenger.get("yeuCauHoaDon") or "Có",
                "tenCongTy": passenger.get("tenCongTy") or "",
                "maSoThue": passenger.get("maSoThue") or "",
                "diaChiHoaDon": passenger.get("diaChiHoaDon") or "",
                "emailHoaDon": passenger.get("emailHoaDon") or "",
                "trangThaiHoaDon": invoice_status(passenger),
                "ngayXuatHoaDon": passenger.get("ngayXuatHoaDon") or "",
                "nguoiXuatHoaDon": passenger.get("nguoiXuatHoaDon") or "",
                "loaiKhach": passenger.get("loaiKhach") or "B2C",
            }
        )
        result.append(item)
    return result


@app.get("/api/invoice-orders")
def list_invoice_orders(request: Request) -> dict[str, Any]:
    user = current_user(request)
    if not any(has_action(user, action) for action in ["view_invoices", "manage_invoices", "export_invoices", "reports_all"]):
        raise HTTPException(status_code=403, detail="Không có quyền xem danh sách hóa đơn.")
    rows = invoice_rows()
    for row in rows:
        row["trangThaiHoaDon"] = invoice_status(row)
    return {"sheetName": ORDERS_SHEET_NAME, "rows": rows, "fetchedAt": now_iso()}


@app.get("/api/invoice-groups/candidates")
def list_invoice_group_candidates(request: Request) -> dict[str, Any]:
    user = current_user(request)
    if not any(has_action(user, action) for action in ["create_invoice_groups", "manage_invoices", "export_invoices", "reports_all"]):
        raise HTTPException(status_code=403, detail="Không có quyền tạo hóa đơn gộp.")
    rows = []
    for order in all_order_records():
        if order.get("deletedAt") or not order_is_done(order):
            continue
        if "ghep" in normalize_text(order.get("loaiHopDong")):
            continue
        if normalize_text(invoice_status(order)) == "da xuat" or str(order.get("nhomHoaDonId") or "").strip():
            continue
        item = dict(order)
        item["tienTruocVAT"] = max(
            money_value(order.get("giaTien")) + money_value(order.get("phuThu"))
            - money_value(order.get("giamGia")) - money_value(order.get("tongUuDai")),
            0,
        )
        rows.append(item)
    return {"sheetName": ORDERS_SHEET_NAME, "rows": rows, "fetchedAt": now_iso()}


@app.post("/api/invoice-groups")
def create_invoice_group(payload: InvoiceGroupInput, request: Request) -> dict[str, Any]:
    user = current_user(request)
    if not any(has_action(user, action) for action in ["create_invoice_groups", "manage_invoices"]):
        raise HTTPException(status_code=403, detail="Không có quyền xuất hóa đơn gộp.")
    selected_ids = list(dict.fromkeys(str(value).strip() for value in payload.orderIds if str(value).strip()))
    if len(selected_ids) < 2:
        raise HTTPException(status_code=422, detail="Vui lòng chọn ít nhất hai đơn hàng.")
    worksheet = orders_worksheet()
    all_orders = worksheet_records(worksheet, ORDER_HEADERS)
    selected = [row for row in all_orders if str(row.get("id") or "") in selected_ids]
    if len(selected) != len(selected_ids):
        raise HTTPException(status_code=404, detail="Có đơn hàng không còn tồn tại.")
    customer_keys = {
        str(row.get("khachHangId") or "").strip() or normalize_phone(row.get("soDienThoai"))
        for row in selected
    }
    if len(customer_keys) != 1:
        raise HTTPException(status_code=409, detail="Chỉ được gộp các đơn của cùng một khách hàng.")
    for order in selected:
        if not order_is_done(order) or "ghep" in normalize_text(order.get("loaiHopDong")):
            raise HTTPException(status_code=409, detail="Chỉ gộp các đơn nguyên chuyến đã hoàn thành.")
        if normalize_text(invoice_status(order)) == "da xuat" or str(order.get("nhomHoaDonId") or "").strip():
            raise HTTPException(status_code=409, detail=f"Đơn {order.get('id')} đã được xuất hóa đơn.")

    group_id = make_id("HDG")
    timestamp = now_iso()
    actor = current_user_display_name(request)
    total_before_vat = 0.0
    total_vat = 0.0
    for order in selected:
        before_vat = max(
            money_value(order.get("giaTien")) + money_value(order.get("phuThu"))
            - money_value(order.get("giamGia")) - money_value(order.get("tongUuDai")),
            0,
        )
        vat = round(before_vat * 0.08)
        total_before_vat += before_vat
        total_vat += vat
        updated = dict(order)
        updated.update({
            "yeuCauHoaDon": "Có", "tenCongTy": payload.tenCongTy.strip(), "maSoThue": payload.maSoThue.strip(),
            "diaChiHoaDon": payload.diaChiHoaDon.strip(), "emailHoaDon": payload.emailHoaDon.strip(),
            "thueVAT": vat, "tongThanhToan": before_vat + vat,
            "thucThu": max(before_vat + vat - money_value(order.get("daCoc")), 0),
            "trangThaiHoaDon": "Chưa xuất",
            "ngayXuatHoaDon": "", "nguoiXuatHoaDon": "", "nhomHoaDonId": group_id,
        })
        row_number = find_row_by_id(worksheet, str(order.get("id") or ""))
        if row_number is None:
            raise HTTPException(status_code=404, detail=f"Không tìm thấy đơn {order.get('id')}.")
        update_row_by_headers(worksheet, row_number, ORDER_HEADERS, updated)

    first = selected[0]
    group = {
        "id": group_id, "khachHangId": first.get("khachHangId") or "", "tenKhach": first.get("tenKhach") or "",
        "soDienThoai": first.get("soDienThoai") or "", "orderIds": ", ".join(selected_ids),
        "tenCongTy": payload.tenCongTy.strip(), "maSoThue": payload.maSoThue.strip(),
        "diaChiHoaDon": payload.diaChiHoaDon.strip(), "emailHoaDon": payload.emailHoaDon.strip(),
        "tongTruocVAT": total_before_vat, "tongVAT": total_vat,
        "tongThanhToan": total_before_vat + total_vat, "trangThai": "Chưa xuất",
        "createdAt": timestamp, "createdBy": actor,
        "ngayXuatHoaDon": "", "nguoiXuatHoaDon": "",
    }
    append_worksheet_row(invoice_groups_worksheet(), [group.get(header, "") for header in INVOICE_GROUP_HEADERS])
    log_action(request, "create_invoice_group", "invoice_group", group_id, note=group["orderIds"], after=group)
    return {"ok": True, "row": group}


@app.post("/api/invoice-groups/{group_id}/invoice-status")
def update_invoice_group_status(group_id: str, payload: InvoiceStatusInput, request: Request) -> dict[str, Any]:
    user = current_user(request)
    if not has_action(user, "manage_invoices"):
        raise HTTPException(status_code=403, detail="Không có quyền cập nhật trạng thái hóa đơn gộp.")
    group_worksheet = invoice_groups_worksheet()
    groups = worksheet_records(group_worksheet, INVOICE_GROUP_HEADERS)
    group_row_number = find_row_by_id(group_worksheet, group_id)
    if group_row_number is None:
        raise HTTPException(status_code=404, detail="Không tìm thấy hóa đơn gộp.")
    group = row_by_id(groups, group_id) or {}
    before = dict(group)
    timestamp = now_iso() if payload.trangThaiHoaDon == "Đã xuất" else ""
    actor = current_user_display_name(request) if timestamp else ""
    group["trangThai"] = payload.trangThaiHoaDon
    group["ngayXuatHoaDon"] = timestamp
    group["nguoiXuatHoaDon"] = actor
    update_row_by_headers(group_worksheet, group_row_number, INVOICE_GROUP_HEADERS, group)

    order_worksheet = orders_worksheet()
    orders = worksheet_records(order_worksheet, ORDER_HEADERS)
    member_ids = [value.strip() for value in str(group.get("orderIds") or "").split(",") if value.strip()]
    for order_id in member_ids:
        row_number = find_row_by_id(order_worksheet, order_id)
        order = row_by_id(orders, order_id) or {}
        if row_number is None or not order:
            continue
        order["trangThaiHoaDon"] = payload.trangThaiHoaDon
        order["ngayXuatHoaDon"] = timestamp
        order["nguoiXuatHoaDon"] = actor
        update_row_by_headers(order_worksheet, row_number, ORDER_HEADERS, order)
    log_action(request, "update_invoice_group_status", "invoice_group", group_id, before=before, after=group)
    return {"ok": True, "id": group_id, "trangThaiHoaDon": payload.trangThaiHoaDon}


def debt_status(row: dict[str, Any]) -> str:
    status = str(row.get("trangThaiCongNo") or "").strip()
    return status if status else "Chưa thu hồi"


def debt_order_rows() -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    orders = all_order_records()
    orders_by_id = {str(order.get("id") or ""): order for order in orders}
    for order in orders:
        if normalize_text(order.get("congNo")) not in {"co", "yes", "true", "1"}:
            continue
        if "huy" in normalize_text(order.get("trangThai")):
            continue
        item = dict(order)
        gross = money_value(item.get("giaTien"))
        manual = money_value(item.get("giamGia"))
        benefits = money_value(item.get("tongUuDai"))
        vat = money_value(item.get("thueVAT"))
        deposit = money_value(item.get("daCoc"))
        item["soTienCongNo"] = max(gross + money_value(item.get("phuThu")) - manual - benefits + vat - deposit, 0)
        item["trangThaiCongNo"] = debt_status(item)
        item["debtEntityType"] = "order"
        item["orderCode"] = item.get("id") or ""
        rows.append(item)
    for passenger in all_shared_ride_records():
        if normalize_text(passenger.get("congNo")) not in {"co", "yes", "true", "1"}:
            continue
        parent = orders_by_id.get(str(passenger.get("donHangId") or ""), {})
        if not parent or "huy" in normalize_text(parent.get("trangThai")):
            continue
        item = dict(parent)
        item.update(passenger)
        item["id"] = passenger.get("id") or ""
        item["orderCode"] = parent.get("id") or passenger.get("donHangId") or ""
        item["debtEntityType"] = "sharedPassenger"
        item["tenKhach"] = passenger.get("hoTen") or ""
        item["soDienThoai"] = passenger.get("soDienThoai") or ""
        item["giaTien"] = passenger.get("soTien") or 0
        item["tuyen"] = passenger.get("tuyen") or parent.get("tuyen") or ""
        item["ngayGioDi"] = passenger.get("ngayGioDi") or parent.get("ngayGioDi") or ""
        item["soTienCongNo"] = max(
            money_value(passenger.get("soTien"))
            + money_value(passenger.get("phuThu"))
            - money_value(passenger.get("giamGia"))
            - money_value(passenger.get("tongUuDai"))
            + money_value(passenger.get("thueVAT"))
            - money_value(passenger.get("daCoc")),
            0,
        )
        item["trangThaiCongNo"] = debt_status(passenger)
        rows.append(item)
    return rows


@app.get("/api/debt-orders")
def list_debt_orders(request: Request) -> dict[str, Any]:
    user = current_user(request)
    if not any(has_action(user, action) for action in ["manage_debts", "export_debts", "reports_all"]):
        raise HTTPException(status_code=403, detail="Không có quyền xem danh sách công nợ.")
    return {"sheetName": ORDERS_SHEET_NAME, "rows": debt_order_rows(), "fetchedAt": now_iso()}


@app.post("/api/debt-orders/{order_id}/status")
def update_debt_order_status(
    order_id: str,
    payload: DebtStatusInput,
    request: Request,
    entityType: str = "order",
) -> dict[str, Any]:
    user = current_user(request)
    if not has_action(user, "manage_debts"):
        raise HTTPException(status_code=403, detail="Không có quyền xác nhận thu hồi công nợ.")
    is_shared_passenger = entityType == "sharedPassenger"
    worksheet = shared_ride_worksheet() if is_shared_passenger else orders_worksheet()
    headers = SHARED_RIDE_HEADERS if is_shared_passenger else ORDER_HEADERS
    rows = worksheet_records(worksheet, headers)
    row_number = find_row_by_id(worksheet, order_id)
    if row_number is None:
        raise HTTPException(status_code=404, detail="Không tìm thấy đơn hàng.")
    order = row_by_id(rows, order_id) or {}
    if normalize_text(order.get("congNo")) not in {"co", "yes", "true", "1"}:
        raise HTTPException(status_code=409, detail="Đơn hàng này không được ghi nhận công nợ.")
    before = dict(order)
    order["trangThaiCongNo"] = payload.trangThaiCongNo
    if payload.trangThaiCongNo == "Đã thu hồi":
        order["ngayThuHoiCongNo"] = now_iso()
        order["nguoiThuHoiCongNo"] = current_user_display_name(request)
    else:
        order["ngayThuHoiCongNo"] = ""
        order["nguoiThuHoiCongNo"] = ""
    update_row_by_headers(worksheet, row_number, headers, order)
    entity_name = "shared_passenger" if is_shared_passenger else "order"
    log_action(request, "update_debt_status", entity_name, order_id, before=before, after=order)
    return {
        "ok": True,
        "id": order_id,
        "entityType": entityType,
        "trangThaiCongNo": order["trangThaiCongNo"],
    }


def commission_status(row: dict[str, Any]) -> str:
    status = str(row.get("trangThaiHoaHong") or "").strip()
    return status if status else "Chưa thu"


def commission_order_rows() -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for order in all_order_records():
        if order.get("deletedAt") or "huy" in normalize_text(order.get("trangThai")):
            continue
        commission = money_value(order.get("soTienNopLai"))
        if commission <= 0 or "thuong quyen" not in normalize_text(order.get("loaiXeDieuDong")):
            continue
        item = dict(order)
        item["trangThaiHoaHong"] = commission_status(item)
        rows.append(item)
    return rows


@app.get("/api/commission-orders")
def list_commission_orders(request: Request) -> dict[str, Any]:
    user = current_user(request)
    if not any(has_action(user, action) for action in ["manage_commissions", "export_commissions", "reports_all"]):
        raise HTTPException(status_code=403, detail="Không có quyền xem danh sách hoa hồng xe thương quyền.")
    return {"sheetName": ORDERS_SHEET_NAME, "rows": commission_order_rows(), "fetchedAt": now_iso()}


@app.post("/api/commission-orders/{order_id}/status")
def update_commission_order_status(order_id: str, payload: CommissionStatusInput, request: Request) -> dict[str, Any]:
    user = current_user(request)
    if not has_action(user, "manage_commissions"):
        raise HTTPException(status_code=403, detail="Không có quyền xác nhận thu hoa hồng.")
    worksheet = orders_worksheet()
    row_number = find_row_by_id(worksheet, order_id)
    if row_number is None:
        raise HTTPException(status_code=404, detail="Không tìm thấy đơn hàng.")
    order = row_by_id(worksheet_records(worksheet, ORDER_HEADERS), order_id) or {}
    if money_value(order.get("soTienNopLai")) <= 0 or "thuong quyen" not in normalize_text(order.get("loaiXeDieuDong")):
        raise HTTPException(status_code=409, detail="Đơn hàng không phát sinh hoa hồng xe thương quyền.")
    before = dict(order)
    order["trangThaiHoaHong"] = payload.trangThaiHoaHong
    if payload.trangThaiHoaHong == "Đã thu":
        order["ngayThuHoaHong"] = now_iso()
        order["nguoiThuHoaHong"] = current_user_display_name(request)
    else:
        order["ngayThuHoaHong"] = ""
        order["nguoiThuHoaHong"] = ""
    update_row_by_headers(worksheet, row_number, ORDER_HEADERS, order)
    log_action(request, "update_commission_status", "order", order_id, before=before, after=order)
    return {"ok": True, "id": order_id, "trangThaiHoaHong": order["trangThaiHoaHong"]}


@app.post("/api/shared-passengers/{passenger_id}/invoice-status")
def update_shared_passenger_invoice_status(passenger_id: str, payload: InvoiceStatusInput, request: Request) -> dict[str, Any]:
    worksheet = shared_ride_worksheet()
    rows = worksheet_records(worksheet, SHARED_RIDE_HEADERS)
    row_number = find_row_by_id(worksheet, passenger_id)
    if row_number is None:
        raise HTTPException(status_code=404, detail="Không tìm thấy khách xe ghép.")
    passenger = row_by_id(rows, passenger_id) or {}
    if not order_requires_invoice(passenger):
        raise HTTPException(status_code=409, detail="Khách này không có yêu cầu xuất hóa đơn.")
    passenger["trangThaiHoaDon"] = payload.trangThaiHoaDon
    if payload.trangThaiHoaDon == "Đã xuất":
        passenger["ngayXuatHoaDon"] = now_iso()
        passenger["nguoiXuatHoaDon"] = current_user_display_name(request)
    else:
        passenger["ngayXuatHoaDon"] = ""
        passenger["nguoiXuatHoaDon"] = ""
    update_row_by_headers(worksheet, row_number, SHARED_RIDE_HEADERS, passenger)
    return {"ok": True, "id": passenger_id}


@app.post("/api/orders/{order_id}/invoice-status")
def update_invoice_status(order_id: str, payload: InvoiceStatusInput, request: Request) -> dict[str, Any]:
    worksheet = orders_worksheet()
    rows = worksheet_records(worksheet, ORDER_HEADERS)
    row_number = find_row_by_id(worksheet, order_id)
    if row_number is None:
        raise HTTPException(status_code=404, detail="Không tìm thấy đơn hàng.")
    order = row_by_id(rows, order_id) or {}
    if not order_requires_invoice(order):
        raise HTTPException(status_code=409, detail="Đơn hàng này không có yêu cầu xuất hóa đơn.")
    before = dict(order)
    order["trangThaiHoaDon"] = payload.trangThaiHoaDon
    if payload.trangThaiHoaDon == "Đã xuất":
        order["ngayXuatHoaDon"] = now_iso()
        order["nguoiXuatHoaDon"] = current_user_display_name(request)
    else:
        order["ngayXuatHoaDon"] = ""
        order["nguoiXuatHoaDon"] = ""
    update_row_by_headers(worksheet, row_number, ORDER_HEADERS, order)
    log_action(request, "update_invoice_status", "order", order_id, before=before, after=order)
    return {"ok": True, "id": order_id, "trangThaiHoaDon": order["trangThaiHoaDon"]}


def money_value(value: Any) -> float:
    text = str(value or "").strip().replace(" ", "").replace("%", "")
    if not text:
        return 0.0
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        parts = text.split(",")
        text = "".join(parts) if len(parts[-1]) == 3 else text.replace(",", ".")
    elif "." in text:
        parts = text.split(".")
        if len(parts) > 2 or len(parts[-1]) == 3:
            text = "".join(parts)
    try:
        return float(text)
    except ValueError:
        return 0.0


def excel_safe_value(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    # Excel/OpenXML không chấp nhận các ký tự điều khiển và giới hạn mỗi ô 32.767 ký tự.
    cleaned = re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", value)
    return cleaned[:32767]


def round_up_ten_thousand(value: Any) -> float:
    amount = max(money_value(value), 0)
    return math.ceil(amount / 10000) * 10000 if amount else 0


def driver_revenue_amount(
    gross: Any,
    manual_discount: Any = 0,
    benefit_discount: Any = 0,
    vat: Any = 0,
    surcharge: Any = 0,
) -> float:
    return max(
        money_value(gross)
        + money_value(surcharge)
        - money_value(manual_discount)
        - money_value(benefit_discount),
        0,
    )


def order_driver_revenue(row: dict[str, Any]) -> float:
    return driver_revenue_amount(
        row.get("giaTien"),
        row.get("giamGia"),
        row.get("tongUuDai"),
        row.get("thueVAT"),
        row.get("phuThu"),
    )


def order_driver_remittance(row: dict[str, Any]) -> float:
    if normalize_text(row.get("congNo")) in {"co", "yes", "true", "1"}:
        return 0
    deposit = money_value(row.get("daCoc"))
    normal_due = max(order_driver_revenue(row) + money_value(row.get("thueVAT")) - deposit, 0)
    commission = money_value(row.get("soTienNopLai"))
    if "thuong quyen" in normalize_text(row.get("loaiXeDieuDong")) and commission > 0:
        return commission
    return normal_due


def order_remittance_status(row: dict[str, Any]) -> str:
    if normalize_text(row.get("congNo")) in {"co", "yes", "true", "1"}:
        return "Công nợ"
    status = str(row.get("trangThaiNopTien") or "").strip()
    if status in {"Chưa nộp tiền", "Đã nộp tiền"}:
        return status
    return "Chưa nộp tiền"


@app.get("/api/reports/driver-remittance.xlsx")
def export_driver_remittance_report(ngay: str = "") -> Response:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover
        raise HTTPException(status_code=500, detail="Thiếu thư viện xuất Excel.") from exc

    def same_report_date(value: Any) -> bool:
        if not report_date:
            return True
        parsed = parse_existing_datetime(value)
        return bool(parsed and parsed.date() == report_date.date())

    def route_label(start_point: Any, end_point: Any, fallback: Any = "") -> str:
        start_text = str(start_point or "").strip()
        end_text = str(end_point or "").strip()
        if start_text and end_text:
            return f"{start_text} - {end_text}"
        return str(fallback or "").strip()

    def normal_due(gross: float, discount: float, deposit: float) -> float:
        return max(gross - discount - deposit, 0)

    def is_shared_contract(value: Any) -> bool:
        normalized = normalize_text(value).replace("_", " ")
        return normalized == "xe ghep"

    def is_company_dispatched(row: dict[str, Any]) -> bool:
        vehicle_type = normalize_text(row.get("loaiXeDieuDong"))
        return bool(vehicle_type) and "thuong quyen" not in vehicle_type

    report_date = parse_existing_date(ngay)

    def roster_is_remittance_eligible(row: dict[str, Any]) -> bool:
        status = normalize_text(row.get("trangThaiLenXuongCa"))
        return "len ca" in status or "xuong ca" in status

    all_orders = all_order_records()
    shared_rows = all_shared_ride_records()
    orders_by_id = {str(row.get("id") or ""): row for row in all_orders}
    roster = roster_rows()
    roster_by_plate_date: dict[tuple[str, str], dict[str, Any]] = {}
    roster_by_plate: dict[str, dict[str, Any]] = {}
    roster_by_driver_date: dict[tuple[str, str], dict[str, Any]] = {}
    driver_first_seen: dict[tuple[str, str], tuple[datetime, int]] = {}
    area_first_seen: dict[str, tuple[datetime, int]] = {}
    for roster_index, roster_row in enumerate(roster):
        plate = normalize_text(roster_row.get("bienKiemSoat"))
        driver = normalize_text(roster_driver_name(roster_driver_text(roster_row)))
        roster_day = roster_date_key(roster_row)
        area = str(roster_row.get("khuVucHoatDong") or "Chưa xác định khu vực").strip() or "Chưa xác định khu vực"
        area_key = normalize_text(area)
        first_seen_at = parse_roster_date(roster_row.get("thoiGianTao"))
        if first_seen_at == datetime.min:
            first_seen_at = datetime.max
        area_candidate = (first_seen_at, roster_index)
        if area_key not in area_first_seen or area_candidate < area_first_seen[area_key]:
            area_first_seen[area_key] = area_candidate
        if driver:
            driver_key = (area_key, driver)
            if driver_key not in driver_first_seen or area_candidate < driver_first_seen[driver_key]:
                driver_first_seen[driver_key] = area_candidate
        if plate:
            roster_by_plate[plate] = better_roster_row(roster_by_plate.get(plate), roster_row)
            if roster_day:
                key = (plate, roster_day)
                roster_by_plate_date[key] = better_roster_row(roster_by_plate_date.get(key), roster_row)
        if driver and roster_day:
            key = (driver, roster_day)
            roster_by_driver_date[key] = better_roster_row(roster_by_driver_date.get(key), roster_row)

    area_order = {
        area_key: index
        for index, (area_key, _) in enumerate(
            sorted(area_first_seen.items(), key=lambda item: (item[1], item[0]))
        )
    }
    driver_area_order = {
        driver_key: index
        for index, (driver_key, _) in enumerate(
            sorted(
                driver_first_seen.items(),
                key=lambda item: (
                    area_order.get(item[0][0], len(area_order)),
                    item[1],
                    item[0][1],
                ),
            )
        )
    }

    def operating_area(row: dict[str, Any], date_source: Any) -> str:
        parsed = parse_existing_datetime(date_source)
        day = parsed.strftime("%Y-%m-%d") if parsed else ""
        plate = normalize_text(row.get("bienKiemSoat"))
        driver = normalize_text(row.get("hoTenLaiXe"))
        roster_row = (
            roster_by_plate_date.get((plate, day))
            or roster_by_driver_date.get((driver, day))
            or roster_by_plate.get(plate)
            or {}
        )
        return str(roster_row.get("khuVucHoatDong") or "Chưa xác định khu vực").strip() or "Chưa xác định khu vực"

    def is_on_report_roster(row: dict[str, Any], date_source: Any) -> bool:
        if not report_date:
            return True
        parsed = parse_existing_datetime(date_source)
        day = parsed.strftime("%Y-%m-%d") if parsed else report_date.strftime("%Y-%m-%d")
        plate = normalize_text(row.get("bienKiemSoat"))
        driver = normalize_text(row.get("hoTenLaiXe"))
        roster_row = roster_by_plate_date.get((plate, day)) or roster_by_driver_date.get((driver, day))
        return bool(roster_row and roster_is_remittance_eligible(roster_row))

    total_shared_gross_by_order: dict[str, float] = {}
    for passenger in shared_rows:
        order_id = str(passenger.get("donHangId") or "")
        total_shared_gross_by_order[order_id] = total_shared_gross_by_order.get(order_id, 0) + money_value(passenger.get("soTien"))

    report_items: list[dict[str, Any]] = []
    for row in all_orders:
        if "huy" in normalize_text(row.get("trangThai")):
            continue
        if is_shared_contract(row.get("loaiHopDong")):
            continue
        if not is_company_dispatched(row):
            continue
        if not same_report_date(row.get("ngayGioDi")):
            continue
        if not is_on_report_roster(row, row.get("ngayGioDi")):
            continue

        start_at = parse_existing_datetime(row.get("ngayGioDi"))
        gross = money_value(row.get("giaTien"))
        surcharge = money_value(row.get("phuThu"))
        discount = money_value(row.get("giamGia")) + money_value(row.get("tongUuDai"))
        revenue = order_driver_revenue(row)
        vat = money_value(row.get("thueVAT"))
        deposit = money_value(row.get("daCoc"))
        debt = max(revenue + vat - deposit, 0) if normalize_text(row.get("congNo")) in {"co", "yes", "true", "1"} else 0
        actual_receipt = max(revenue + vat - deposit - debt, 0)
        note = str(row.get("ghiChu") or "")
        if normalize_text(row.get("congNo")) in {"co", "yes", "true", "1"}:
            note = f"Công nợ: {row.get('congNoChoAi') or 'Chưa ghi người chịu công nợ'}"
        if "thuong quyen" in normalize_text(row.get("loaiXeDieuDong")):
            commission_rate = money_value(row.get("tyLeNopLai"))
            note = f"Xe thương quyền nộp {commission_rate:g}%" if commission_rate else "Xe thương quyền"

        report_items.append(
            {
                "driver": str(row.get("hoTenLaiXe") or "Chưa gán lái xe").strip(),
                "area": operating_area(row, row.get("ngayGioDi")),
                "sort_at": start_at or datetime.min,
                "date": start_at.strftime("%d/%m/%Y") if start_at else "",
                "customer": row.get("tenKhach") or "",
                "route": row.get("tuyen") or route_label(row.get("diemDon"), row.get("diemTra")),
                "base_before_vat": gross,
                "surcharge": surcharge,
                "deposit": deposit or "",
                "discount": discount or "",
                "revenue": revenue,
                "vat": vat or "",
                "actual_receipt": actual_receipt,
                "remittance_status": order_remittance_status(row),
                "note": note,
            }
        )

    for passenger in shared_rows:
        parent = orders_by_id.get(str(passenger.get("donHangId") or ""), {})
        if parent and "huy" in normalize_text(parent.get("trangThai")):
            continue
        if not parent or not is_company_dispatched(parent):
            continue
        date_source = passenger.get("ngayGioDi") or parent.get("ngayGioDi")
        if not same_report_date(date_source):
            continue
        if not is_on_report_roster(parent, date_source):
            continue

        start_at = parse_existing_datetime(date_source)
        gross = money_value(passenger.get("soTien"))
        discount = money_value(passenger.get("giamGia")) + money_value(passenger.get("tongUuDai"))
        deposit = money_value(passenger.get("daCoc"))
        passenger_revenue = driver_revenue_amount(
            gross,
            passenger.get("giamGia"),
            passenger.get("tongUuDai"),
            passenger.get("thueVAT"),
            passenger.get("phuThu"),
        )
        surcharge = money_value(passenger.get("phuThu"))
        vat = money_value(passenger.get("thueVAT"))
        is_debt = normalize_text(passenger.get("congNo")) in {"co", "yes", "true", "1"}
        debt = max(passenger_revenue + vat - deposit, 0) if is_debt else 0
        actual_receipt = max(passenger_revenue + vat - deposit - debt, 0)
        note = str(parent.get("ghiChu") or "")
        if normalize_text(passenger.get("congNo")) in {"co", "yes", "true", "1"}:
            note = f"Công nợ: {passenger.get('congNoChoAi') or 'Chưa ghi người chịu công nợ'}"
        if "thuong quyen" in normalize_text(parent.get("loaiXeDieuDong")):
            commission = money_value(parent.get("soTienNopLai"))
            order_gross = total_shared_gross_by_order.get(str(parent.get("id") or ""), 0)
            if commission > 0 and order_gross > 0:
                actual_receipt = round(commission * gross / order_gross)
            commission_rate = money_value(parent.get("tyLeNopLai"))
            note = f"Xe thương quyền nộp {commission_rate:g}%" if commission_rate else "Xe thương quyền"

        report_items.append(
            {
                "driver": str(parent.get("hoTenLaiXe") or "Chưa gán lái xe").strip(),
                "area": operating_area(parent, date_source),
                "sort_at": start_at or datetime.min,
                "date": start_at.strftime("%d/%m/%Y") if start_at else "",
                "customer": passenger.get("hoTen") or passenger.get("soDienThoai") or "Khách lẻ",
                "route": route_label(passenger.get("diemDon"), passenger.get("diemTra"), passenger.get("tuyen") or parent.get("tuyen")),
                "base_before_vat": gross,
                "surcharge": surcharge,
                "deposit": deposit or "",
                "discount": discount or "",
                "revenue": passenger_revenue,
                "vat": vat or "",
                "actual_receipt": actual_receipt,
                "remittance_status": "Công nợ" if is_debt else order_remittance_status(parent),
                "note": note,
            }
        )

    if report_date:
        reported_drivers = {normalize_text(item.get("driver")) for item in report_items}
        master_driver_rows: dict[str, dict[str, Any]] = {}
        for roster_row in roster:
            if roster_date_key(roster_row) != report_date.strftime("%Y-%m-%d") or not roster_is_remittance_eligible(roster_row):
                continue
            driver_key = normalize_text(roster_driver_name(roster_driver_text(roster_row)))
            if not driver_key:
                continue
            current = master_driver_rows.get(driver_key)
            if current is None or parse_roster_date(roster_row.get("thoiGianTao")) >= parse_roster_date(current.get("thoiGianTao")):
                master_driver_rows[driver_key] = roster_row
        ordered_master_rows = sorted(
            master_driver_rows.values(),
            key=lambda roster_row: (
                area_order.get(normalize_text(roster_row.get("khuVucHoatDong")), len(area_order)),
                driver_area_order.get(
                    (
                        normalize_text(roster_row.get("khuVucHoatDong") or "Chưa xác định khu vực"),
                        normalize_text(roster_driver_name(roster_driver_text(roster_row))),
                    ),
                    len(roster),
                ),
            ),
        )
        for roster_row in ordered_master_rows:
            vehicle_description = normalize_text(
                " ".join(
                    str(roster_row.get(field) or "")
                    for field in ["soHieuXe", "loai_xe", "loaiXe", "so_cho", "soCho"]
                )
            )
            if "tai van 945 kg" in vehicle_description:
                continue
            driver = roster_driver_name(roster_driver_text(roster_row)).strip()
            if not driver or normalize_text(driver) in reported_drivers:
                continue
            reported_drivers.add(normalize_text(driver))
            report_items.append(
                {
                    "driver": driver,
                    "area": str(roster_row.get("khuVucHoatDong") or "Chưa xác định khu vực").strip() or "Chưa xác định khu vực",
                    "sort_at": report_date,
                    "date": report_date.strftime("%d/%m/%Y"),
                    "customer": "—",
                    "route": "—",
                    "base_before_vat": 0,
                    "surcharge": 0,
                    "deposit": 0,
                    "discount": 0,
                    "revenue": 0,
                    "vat": 0,
                    "actual_receipt": 0,
                    "remittance_status": "Chưa nộp tiền",
                    "note": "Không phát sinh doanh thu",
                }
            )

    unknown_order = len(roster) + len(report_items) + 1
    report_items.sort(
        key=lambda item: (
            area_order.get(normalize_text(item.get("area")), len(area_order)),
            driver_area_order.get(
                (normalize_text(item.get("area")), normalize_text(item.get("driver"))),
                unknown_order,
            ),
            item.get("sort_at") or datetime.min,
        )
    )
    grouped: dict[tuple[str, str], list[dict[str, Any]]] = {}
    for item in report_items:
        area = str(item.get("area") or "Chưa xác định khu vực")
        driver = str(item.get("driver") or "Chưa gán lái xe")
        grouped.setdefault((area, driver), []).append(item)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "LENH_NOP_TIEN"
    sheet.sheet_view.showGridLines = False

    thin = Side(style="thin", color="CBD5E1")
    medium = Side(style="medium", color="0F172A")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    section_border = Border(left=medium, right=medium, top=medium, bottom=medium)
    green_fill = PatternFill("solid", fgColor="D9EAD3")
    title_fill = PatternFill("solid", fgColor="EAF3F8")
    driver_fill = PatternFill("solid", fgColor="F8FAFC")
    area_fill = PatternFill("solid", fgColor="DDEBFF")
    total_fill = PatternFill("solid", fgColor="FEF3C7")
    money_format = '#,##0'
    headers = ["STT", "NGÀY", "KHÁCH HÀNG", "TUYẾN", "GIÁ CHƯA VAT", "TỔNG PHỤ THU", "TỔNG GIẢM GIÁ", "DOANH THU", "THUẾ VAT", "ĐÃ CỌC", "SỐ TIỀN THỰC THU", "TRẠNG THÁI NỘP TIỀN", "GHI CHÚ"]
    widths = [7, 13, 28, 24, 16, 15, 17, 16, 14, 14, 19, 20, 34]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=13)
    title_cell = sheet.cell(1, 1, "LỆNH NỘP TIỀN LÁI XE")
    title_cell.font = Font(bold=True, size=18, color="0F172A")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = title_fill
    sheet.row_dimensions[1].height = 28

    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=13)
    subtitle = f"Ngày báo cáo: {report_date.strftime('%d/%m/%Y')}" if report_date else "Tất cả đơn hàng"
    subtitle_cell = sheet.cell(2, 1, subtitle)
    subtitle_cell.font = Font(italic=True, color="475569")
    subtitle_cell.alignment = Alignment(horizontal="center")

    grand_total = sum(money_value(item.get("actual_receipt")) for item in report_items)
    sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=10)
    sheet.cell(3, 1, "TỔNG SỐ TIỀN PHẢI NỘP")
    sheet.cell(3, 1).font = Font(bold=True, size=13, color="0F172A")
    sheet.cell(3, 1).alignment = Alignment(horizontal="right")
    sheet.cell(3, 11, grand_total)
    sheet.cell(3, 11).font = Font(bold=True, size=14, color="C00000")
    sheet.cell(3, 11).number_format = money_format
    sheet.cell(3, 11).alignment = Alignment(horizontal="right")
    for col in range(1, 14):
        cell = sheet.cell(3, col)
        cell.fill = total_fill
        cell.border = border

    current_row = 5
    current_area = ""
    for (area, driver), driver_items in grouped.items():
        if area != current_area:
            if current_area:
                current_row += 1
            current_area = area
            area_total = sum(
                money_value(item.get("actual_receipt"))
                for item in report_items
                if str(item.get("area") or "Chưa xác định khu vực") == area
            )
            sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
            area_cell = sheet.cell(current_row, 1, f"KHU VỰC HOẠT ĐỘNG: {area.upper()}")
            area_cell.font = Font(bold=True, size=15, color="1E3A8A")
            area_cell.alignment = Alignment(horizontal="left", vertical="center")
            sheet.cell(current_row, 11, area_total).number_format = money_format
            sheet.cell(current_row, 11).font = Font(bold=True, size=14, color="C00000")
            for col in range(1, 14):
                sheet.cell(current_row, col).fill = area_fill
                sheet.cell(current_row, col).border = section_border
            sheet.row_dimensions[current_row].height = 27
            current_row += 1
        total_due = sum(money_value(item.get("actual_receipt")) for item in driver_items)
        sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
        title_cell = sheet.cell(current_row, 1, f"{driver.upper()} - TỔNG SỐ TIỀN PHẢI NỘP")
        title_cell.font = Font(bold=True, size=14, color="0F172A")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.fill = driver_fill
        total_cell = sheet.cell(current_row, 11, total_due)
        total_cell.font = Font(bold=True, size=14, color="C00000")
        total_cell.number_format = money_format
        total_cell.alignment = Alignment(horizontal="right", vertical="center")
        total_cell.fill = driver_fill
        sheet.row_dimensions[current_row].height = 24
        for col in range(1, 14):
            cell = sheet.cell(current_row, col)
            cell.border = section_border
            cell.fill = driver_fill

        current_row += 1
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(current_row, col, header)
            cell.fill = green_fill
            cell.font = Font(bold=True, color="0F172A")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        sheet.row_dimensions[current_row].height = 22

        current_row += 1
        fixed_items = list(driver_items[:4])
        if len(driver_items) > 4:
            overflow = driver_items[4:]
            combined = {
                "date": "\n".join(str(item.get("date") or "") for item in overflow),
                "customer": "\n".join(str(item.get("customer") or "") for item in overflow),
                "route": "\n".join(str(item.get("route") or "") for item in overflow),
                "base_before_vat": sum(money_value(item.get("base_before_vat")) for item in overflow),
                "surcharge": sum(money_value(item.get("surcharge")) for item in overflow),
                "discount": sum(money_value(item.get("discount")) for item in overflow),
                "revenue": sum(money_value(item.get("revenue")) for item in overflow),
                "vat": sum(money_value(item.get("vat")) for item in overflow),
                "deposit": sum(money_value(item.get("deposit")) for item in overflow),
                "actual_receipt": sum(money_value(item.get("actual_receipt")) for item in overflow),
                "remittance_status": "\n".join(dict.fromkeys(str(item.get("remittance_status") or "Chưa nộp tiền") for item in overflow)),
                "note": f"Gộp {len(overflow)} cuốc từ cuốc thứ 5\n" + "\n".join(str(item.get("note") or "") for item in overflow if item.get("note")),
            }
            fixed_items.append(combined)
        while len(fixed_items) < 5:
            fixed_items.append(None)

        for item_index, item in enumerate(fixed_items, start=1):
            item = item or {}
            values = [
                item_index,
                item.get("date") or "",
                item.get("customer") or "",
                item.get("route") or "",
                item.get("base_before_vat") or "",
                item.get("surcharge") or "",
                item.get("discount") or "",
                item.get("revenue") or "",
                item.get("vat") or "",
                item.get("deposit") or "",
                item.get("actual_receipt") if item.get("actual_receipt") is not None else "",
                item.get("remittance_status") or "Chưa nộp tiền",
                item.get("note") or "",
            ]
            for col, value in enumerate(values, start=1):
                cell = sheet.cell(current_row, col, value)
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.font = Font(color="111827")
                if 5 <= col <= 11:
                    cell.number_format = money_format
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                if col == 11:
                    cell.font = Font(bold=True, color="111827")
            sheet.row_dimensions[current_row].height = 24
            current_row += 1

        subtotal_row = current_row
        sheet.merge_cells(start_row=subtotal_row, start_column=1, end_row=subtotal_row, end_column=10)
        sheet.cell(subtotal_row, 1, "Tổng cộng")
        sheet.cell(subtotal_row, 1).font = Font(bold=True, color="0F172A")
        sheet.cell(subtotal_row, 1).alignment = Alignment(horizontal="right")
        sheet.cell(subtotal_row, 11, total_due)
        sheet.cell(subtotal_row, 11).font = Font(bold=True, color="C00000")
        sheet.cell(subtotal_row, 11).number_format = money_format
        sheet.cell(subtotal_row, 11).alignment = Alignment(horizontal="right")
        for col in range(1, 14):
            cell = sheet.cell(subtotal_row, col)
            cell.fill = total_fill
            cell.border = border
        current_row += 2

        sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
        sheet.merge_cells(start_row=current_row, start_column=8, end_row=current_row, end_column=11)
        sheet.cell(current_row, 1, "Lái xe xác nhận")
        sheet.cell(current_row, 8, "Người lập phiếu")
        for col in (1, 8):
            sheet.cell(current_row, col).font = Font(italic=True, color="64748B")
            sheet.cell(current_row, col).alignment = Alignment(horizontal="center")
        current_row += 3

    sheet.freeze_panes = "A5"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.4
    sheet.page_margins.bottom = 0.4

    if not grouped:
        message = "Chưa có đơn hàng để xuất lệnh nộp tiền."
        if report_date:
            message = f"Chưa có đơn hàng trong ngày {report_date.strftime('%d/%m/%Y')} để xuất lệnh nộp tiền."
        sheet.cell(5, 1, message)

    final_total_row = sheet.max_row + 2
    sheet.merge_cells(start_row=final_total_row, start_column=1, end_row=final_total_row, end_column=10)
    sheet.cell(final_total_row, 1, "TỔNG CỘNG TOÀN BỘ").alignment = Alignment(horizontal="right")
    sheet.cell(final_total_row, 11, grand_total).number_format = money_format
    for col in range(1, 14):
        cell = sheet.cell(final_total_row, col)
        cell.fill = total_fill
        cell.border = section_border
        cell.font = Font(bold=True, color="C00000" if col == 11 else "0F172A")

    transfer_sheet = workbook.create_sheet("NOI_DUNG_NOP_TIEN")
    transfer_sheet.sheet_view.showGridLines = False
    transfer_sheet.column_dimensions["A"].width = 72
    transfer_sheet.column_dimensions["B"].width = 20
    transfer_headers = ["Nội Dung", "Số Tiền"]
    for column, header in enumerate(transfer_headers, start=1):
        cell = transfer_sheet.cell(1, column, header)
        cell.fill = green_fill
        cell.font = Font(bold=True, color="0F172A")
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    transfer_sheet.freeze_panes = "A2"
    transfer_date = report_date or datetime.now()
    transfer_date_text = transfer_date.strftime("%d/%m/%y")
    for row_index, ((_, driver), driver_items) in enumerate(grouped.items(), start=2):
        amount = sum(money_value(item.get("actual_receipt")) for item in driver_items)
        content_cell = transfer_sheet.cell(
            row_index,
            1,
            f"Bán hàng vận chuyển hành khách ngày {transfer_date_text} - {driver}",
        )
        amount_cell = transfer_sheet.cell(row_index, 2, amount)
        content_cell.alignment = Alignment(vertical="center", wrap_text=True)
        amount_cell.alignment = Alignment(horizontal="right", vertical="center")
        amount_cell.number_format = money_format
        for cell in (content_cell, amount_cell):
            cell.border = border
        transfer_sheet.row_dimensions[row_index].height = 22
    transfer_total_row = len(grouped) + 2
    transfer_sheet.cell(transfer_total_row, 1, "TỔNG CỘNG")
    transfer_sheet.cell(transfer_total_row, 2, grand_total).number_format = money_format
    for column in range(1, 3):
        cell = transfer_sheet.cell(transfer_total_row, column)
        cell.fill = total_fill
        cell.border = section_border
        cell.font = Font(bold=True, color="C00000" if column == 2 else "0F172A")
    transfer_sheet.auto_filter.ref = f"A1:B{max(transfer_total_row - 1, 1)}"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    date_suffix = report_date.strftime("%Y%m%d") if report_date else datetime.now().strftime("%Y%m%d-%H%M")
    filename = f"lenh-nop-tien-lai-xe-{date_suffix}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/reports/driver-revenue.xlsx")
def export_driver_revenue_report(tuNgay: str = "", denNgay: str = "") -> Response:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover
        raise HTTPException(status_code=500, detail="Thiếu thư viện xuất Excel.") from exc

    start_date = parse_existing_date(tuNgay) if tuNgay else None
    end_date = parse_existing_date(denNgay) if denNgay else start_date
    if start_date and end_date and start_date.date() > end_date.date():
        raise HTTPException(status_code=422, detail="Từ ngày không được lớn hơn đến ngày.")

    def in_range(value: Any) -> bool:
        parsed = parse_existing_datetime(value)
        if not parsed:
            return not start_date and not end_date
        if start_date and parsed.date() < start_date.date():
            return False
        if end_date and parsed.date() > end_date.date():
            return False
        return True

    orders = all_order_records()
    shared_rows = all_shared_ride_records()
    orders_by_id = {str(row.get("id") or ""): row for row in orders}
    shared_gross_by_order: dict[str, float] = {}
    for passenger in shared_rows:
        order_id = str(passenger.get("donHangId") or "")
        shared_gross_by_order[order_id] = shared_gross_by_order.get(order_id, 0) + money_value(passenger.get("soTien"))
    rows: list[list[Any]] = []

    def append_report_row(
        order: dict[str, Any],
        customer: Any,
        contract_type: str,
        gross: Any,
        manual: Any,
        benefits: Any,
        vat: Any,
        surcharge: Any,
        deposit: Any,
        date_source: Any,
        debt_source: dict[str, Any] | None = None,
    ) -> None:
        if "huy" in normalize_text(order.get("trangThai")) or not in_range(date_source):
            return
        driver_revenue = driver_revenue_amount(gross, manual, benefits, vat, surcharge)
        debt_record = debt_source or order
        is_debt = normalize_text(debt_record.get("congNo")) in {"co", "yes", "true", "1"}
        actual = 0 if is_debt else max(driver_revenue + money_value(vat) - money_value(deposit), 0)
        commission = 0.0
        if "thuong quyen" in normalize_text(order.get("loaiXeDieuDong")):
            commission = money_value(order.get("soTienNopLai"))
            if "ghep" in normalize_text(contract_type):
                order_gross = shared_gross_by_order.get(str(order.get("id") or ""), 0)
                commission = round(commission * money_value(gross) / order_gross) if order_gross > 0 else 0
        parsed = parse_existing_datetime(date_source)
        rows.append(
            [
                parsed or datetime.min,
                parsed.strftime("%d/%m/%Y %H:%M") if parsed else str(date_source or ""),
                order.get("id") or "",
                order.get("hoTenLaiXe") or "Chưa gán lái xe",
                order.get("bienKiemSoat") or "",
                customer or "",
                order.get("tuyen") or route_text(order.get("diemDon"), order.get("diemTra")),
                contract_type,
                money_value(gross),
                money_value(manual),
                money_value(benefits),
                money_value(vat),
                driver_revenue,
                money_value(deposit),
                actual,
                "Có" if is_debt else "Không",
                debt_record.get("congNoChoAi") or "",
                order.get("loaiXeDieuDong") or "Chưa xác định",
                money_value(surcharge),
                commission,
                debt_record.get("ghiChu") or order.get("ghiChu") or "",
            ]
        )

    for order in orders:
        if "ghep" in normalize_text(order.get("loaiHopDong")):
            continue
        append_report_row(
            order,
            order.get("tenKhach"),
            "Xe nguyên chuyến",
            order.get("giaTien"),
            order.get("giamGia"),
            order.get("tongUuDai"),
            order.get("thueVAT"),
            order.get("phuThu"),
            order.get("daCoc"),
            order.get("ngayGioDi") or order.get("createdAt"),
        )
    for passenger in shared_rows:
        order = orders_by_id.get(str(passenger.get("donHangId") or ""))
        if not order:
            continue
        append_report_row(
            order,
            passenger.get("hoTen") or passenger.get("soDienThoai"),
            "Xe ghép",
            passenger.get("soTien"),
            passenger.get("giamGia"),
            passenger.get("tongUuDai"),
            passenger.get("thueVAT"),
            passenger.get("phuThu"),
            passenger.get("daCoc"),
            passenger.get("ngayGioDi") or order.get("ngayGioDi") or passenger.get("createdAt"),
            passenger,
        )
    rows.sort(key=lambda row: (str(row[3]), row[0]))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Chi tiết đơn hàng"
    headers = [
        "STT", "Ngày giờ", "Mã đơn", "Lái xe", "Biển số", "Khách hàng", "Tuyến",
        "Loại đơn", "Giá tiền", "Phụ thu", "Giảm giá thủ công", "Voucher/khuyến mãi",
        "Doanh thu lái xe", "VAT", "Đã cọc", "Công nợ", "Thực thu", "Trạng thái công nợ", "Đối tượng công nợ",
        "Đơn vị vận hành xe", "Hoa hồng xe thương quyền", "Ghi chú",
    ]
    widths = [7, 18, 24, 24, 14, 24, 28, 16, 15, 15, 18, 20, 18, 14, 14, 16, 16, 18, 28, 24, 24, 32]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title = sheet.cell(1, 1, "BÁO CÁO DOANH THU LÁI XE")
    title.font = Font(bold=True, size=16)
    title.alignment = Alignment(horizontal="center")
    period = "Tất cả thời gian"
    if start_date:
        period = f"Từ {start_date.strftime('%d/%m/%Y')} đến {(end_date or start_date).strftime('%d/%m/%Y')}"
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    sheet.cell(2, 1, period).alignment = Alignment(horizontal="center")
    fill = PatternFill("solid", fgColor="DDEBFF")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(4, column, header)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for index, row in enumerate(rows, start=1):
        debt_amount = (
            max(money_value(row[12]) + money_value(row[11]) - money_value(row[13]), 0)
            if normalize_text(row[15]) == "co"
            else 0
        )
        values = (
            [index] + row[1:9]
            + [row[18], row[9], row[10], row[12], row[11], row[13], debt_amount, row[14]]
            + row[15:18] + [row[19], row[20]]
        )
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(index + 4, column, value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if 9 <= column <= 17 or column == 21:
                cell.number_format = '#,##0'
    total_row = len(rows) + 5
    sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=8)
    total_label = sheet.cell(total_row, 1, "TỔNG CỘNG")
    total_label.alignment = Alignment(horizontal="right", vertical="center")
    for column in [9, 10, 11, 12, 13, 14, 15, 16, 17, 21]:
        letter = get_column_letter(column)
        cell = sheet.cell(
            total_row,
            column,
            f"=SUM({letter}5:{letter}{total_row - 1})" if rows else 0,
        )
        cell.number_format = '#,##0'
    for column in range(1, len(headers) + 1):
        cell = sheet.cell(total_row, column)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.border = border
    sheet.freeze_panes = "A5"
    sheet.auto_filter.ref = f"A4:V{max(total_row - 1, 4)}"

    summary_sheet = workbook.create_sheet("Tổng hợp lái xe")
    summary_headers = [
        "STT", "Lái xe", "Biển số", "Đơn vị vận hành xe", "Số đơn hàng", "Tổng giá tiền", "Tổng phụ thu",
        "Tổng giảm giá thủ công", "Tổng voucher/khuyến mãi", "Tổng doanh thu lái xe", "Tổng VAT",
        "Tổng hoa hồng xe thương quyền phải nộp",
        "Tổng đã cọc", "Tổng thực thu", "Số đơn công nợ", "Tổng số tiền công nợ",
    ]
    summary_widths = [7, 28, 18, 24, 14, 18, 16, 22, 24, 16, 22, 28, 16, 18, 16, 22]
    for index, width in enumerate(summary_widths, start=1):
        summary_sheet.column_dimensions[get_column_letter(index)].width = width
    summary_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(summary_headers))
    summary_title = summary_sheet.cell(1, 1, "TỔNG HỢP DOANH THU LÁI XE")
    summary_title.font = Font(bold=True, size=16)
    summary_title.alignment = Alignment(horizontal="center")
    summary_sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(summary_headers))
    summary_sheet.cell(2, 1, period).alignment = Alignment(horizontal="center")
    for column, header in enumerate(summary_headers, start=1):
        cell = summary_sheet.cell(4, column, header)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    driver_totals: dict[str, dict[str, Any]] = {}
    for row in rows:
        driver = str(row[3] or "Chưa gán lái xe")
        item = driver_totals.setdefault(
            driver,
            {
                "plates": set(),
                "vehicle_types": set(),
                "orders": set(),
                "gross": 0.0,
                "surcharge": 0.0,
                "manual": 0.0,
                "benefits": 0.0,
                "vat": 0.0,
                "revenue": 0.0,
                "deposit": 0.0,
                "actual": 0.0,
                "commission": 0.0,
                "commission_orders": set(),
                "debt_amount": 0.0,
                "debt_orders": set(),
            },
        )
        if row[4]:
            item["plates"].add(str(row[4]))
        if row[17]:
            item["vehicle_types"].add(str(row[17]))
        if row[2]:
            item["orders"].add(str(row[2]))
        item["gross"] += money_value(row[8])
        item["surcharge"] += money_value(row[18])
        item["manual"] += money_value(row[9])
        item["benefits"] += money_value(row[10])
        item["vat"] += money_value(row[11])
        item["revenue"] += money_value(row[12])
        item["deposit"] += money_value(row[13])
        item["actual"] += money_value(row[14])
        order_id = str(row[2] or "")
        if order_id and order_id not in item["commission_orders"]:
            source_order = orders_by_id.get(order_id) or {}
            item["commission"] += money_value(source_order.get("soTienNopLai"))
            item["commission_orders"].add(order_id)
        if normalize_text(row[15]) == "co" and row[2]:
            item["debt_orders"].add(str(row[2]))
            item["debt_amount"] += max(
                money_value(row[12]) + money_value(row[11]) - money_value(row[13]),
                0,
            )

    for index, (driver, item) in enumerate(sorted(driver_totals.items()), start=1):
        values = [
            index,
            driver,
            ", ".join(sorted(item["plates"])),
            ", ".join(sorted(item["vehicle_types"])),
            len(item["orders"]),
            item["gross"],
            item["surcharge"],
            item["manual"],
            item["benefits"],
            item["revenue"],
            item["vat"],
            item["commission"],
            item["deposit"],
            item["actual"],
            len(item["debt_orders"]),
            item["debt_amount"],
        ]
        for column, value in enumerate(values, start=1):
            cell = summary_sheet.cell(index + 4, column, excel_safe_value(value))
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if 6 <= column <= 14 or column == 16:
                cell.number_format = '#,##0'
    summary_total_row = len(driver_totals) + 5
    summary_sheet.cell(summary_total_row, 5, "Tổng cộng").font = Font(bold=True)
    for column in range(6, 17):
        letter = get_column_letter(column)
        summary_sheet.cell(
            summary_total_row,
            column,
            f"=SUM({letter}5:{letter}{summary_total_row - 1})" if driver_totals else 0,
        ).number_format = '#,##0'
    summary_sheet.freeze_panes = "A5"
    summary_sheet.auto_filter.ref = f"A4:P{max(summary_total_row - 1, 4)}"
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"bao-cao-doanh-thu-lai-xe-{(start_date or datetime.now()).strftime('%Y%m%d')}-{(end_date or start_date or datetime.now()).strftime('%Y%m%d')}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/reports/summary.xlsx")
def export_summary_report(ngay: str = "", thang: str = "") -> Response:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover
        raise HTTPException(status_code=500, detail="Thiếu thư viện xuất Excel.") from exc

    if thang:
        try:
            report_date = datetime.strptime(thang, "%Y-%m")
        except ValueError:
            raise HTTPException(status_code=422, detail="Tháng báo cáo không hợp lệ.")
    else:
        report_date = parse_existing_date(ngay) or datetime.now()
    month_start = report_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    next_month = (month_start.replace(day=28) + timedelta(days=4)).replace(day=1)
    all_order_rows = all_order_records()
    shared_report_rows = all_shared_ride_records()
    orders = [
        row
        for row in all_order_rows
        if month_start <= (parse_existing_datetime(row.get("ngayGioDi")) or datetime.min) < next_month
        and "huy" not in normalize_text(row.get("trangThai"))
    ]
    customers = customer_records()
    franchise_rows = franchise_vehicle_records()
    roster = roster_rows()
    roster_by_plate: dict[str, dict[str, Any]] = {}
    for row in roster:
        plate = normalize_text(row.get("bienKiemSoat"))
        if plate:
            roster_by_plate[plate] = better_roster_row(roster_by_plate.get(plate), row)
    franchise_by_plate = {normalize_text(row.get("bienKiemSoat")): row for row in franchise_rows}

    def is_shared_order(row: dict[str, Any]) -> bool:
        normalized = normalize_text(row.get("loaiHopDong")).replace("_", " ")
        return normalized == "xe ghep"

    def is_franchise_order(row: dict[str, Any]) -> bool:
        return "thuong quyen" in normalize_text(row.get("loaiXeDieuDong"))

    def is_company_order(row: dict[str, Any]) -> bool:
        return not is_franchise_order(row)

    def order_start(row: dict[str, Any]) -> datetime:
        return parse_existing_datetime(row.get("ngayGioDi")) or datetime.min

    def order_gross(row: dict[str, Any]) -> float:
        return money_value(row.get("giaTien")) + money_value(row.get("phuThu"))

    def order_discount(row: dict[str, Any]) -> float:
        return money_value(row.get("giamGia")) + money_value(row.get("tongUuDai"))

    def order_deposit(row: dict[str, Any]) -> float:
        return money_value(row.get("daCoc"))

    def order_due(row: dict[str, Any]) -> float:
        return max(order_gross(row) - order_discount(row) - order_deposit(row), 0)

    def display_date(value: Any) -> str:
        parsed = parse_existing_datetime(value)
        if parsed and parsed != datetime.min:
            return parsed.strftime("%d/%m/%Y")
        date_value = parse_existing_date(value)
        return date_value.strftime("%d/%m/%Y") if date_value else str(value or "")

    def customer_address_map_from_sheet() -> dict[str, str]:
        values = customers_worksheet().get_all_values()
        if len(values) < 2:
            return {}
        headers = values[0]

        def key(value: Any) -> str:
            return re.sub(r"[^a-z0-9]", "", normalize_text(value))

        phone_index = next((index for index, header in enumerate(headers) if key(header) in {"sodienthoai", "sdt"}), None)
        address_index = next((index for index, header in enumerate(headers) if key(header) in {"diachi", "diachikhach", "diachikhachhang"}), None)
        if phone_index is None or address_index is None:
            return {}
        result: dict[str, str] = {}
        for raw in values[1:]:
            phone = normalize_phone(raw[phone_index] if phone_index < len(raw) else "")
            address = str(raw[address_index] if address_index < len(raw) else "").strip()
            if phone and address and not result.get(phone):
                result[phone] = address
        return result

    active_all_orders = [row for row in all_order_rows if "huy" not in normalize_text(row.get("trangThai"))]
    raw_customer_address_by_phone = customer_address_map_from_sheet()
    customer_address_by_phone: dict[str, str] = {}
    for customer in customers:
        phone_key = normalize_phone(customer.get("soDienThoai"))
        address = str(customer.get("diaChi") or "").strip()
        if phone_key and address and not customer_address_by_phone.get(phone_key):
            customer_address_by_phone[phone_key] = address

    def customer_activity(customer: dict[str, Any]) -> dict[str, Any]:
        phone = normalize_phone(customer.get("soDienThoai"))
        related_orders = [row for row in active_all_orders if phone and normalize_phone(row.get("soDienThoai")) == phone]
        related_shared = [row for row in shared_report_rows if phone and normalize_phone(row.get("soDienThoai")) == phone]
        dates = [order_start(row) for row in related_orders]
        dates.extend(parse_existing_datetime(row.get("ngayGioDi")) or datetime.min for row in related_shared)
        last_date = max([date for date in dates if date != datetime.min], default=None)
        revenue = sum(order_gross(row) for row in related_orders) + sum(money_value(row.get("soTien")) for row in related_shared)
        address = str(customer.get("diaChi") or "").strip() or customer_address_by_phone.get(phone, "")
        if not address:
            related_shared_by_date = sorted(
                related_shared,
                key=lambda item: parse_existing_datetime(item.get("ngayGioDi")) or datetime.min,
                reverse=True,
            )
            address = next((str(row.get("diaChi") or "").strip() for row in related_shared_by_date if str(row.get("diaChi") or "").strip()), "")
        if not address:
            related_orders_by_date = sorted(related_orders, key=order_start, reverse=True)
            address = next(
                (
                    str(row.get("diaChiHoaDon") or "").strip()
                    for row in related_orders_by_date
                    if str(row.get("diaChiHoaDon") or "").strip()
                ),
                "",
            )
        return {
            "last_date": last_date.strftime("%d/%m/%Y") if last_date else "",
            "count": len(related_orders) + len(related_shared),
            "revenue": revenue,
            "address": address,
        }

    def franchise_activity(vehicle: dict[str, Any]) -> dict[str, Any]:
        plate = normalize_text(vehicle.get("bienKiemSoat"))
        related_orders = [row for row in active_all_orders if plate and normalize_text(row.get("bienKiemSoat")) == plate]
        dates = [order_start(row) for row in related_orders]
        last_date = max([date for date in dates if date != datetime.min], default=None)
        return {
            "last_date": last_date.strftime("%d/%m/%Y") if last_date else "",
            "count": len(related_orders),
            "revenue": sum(order_gross(row) for row in related_orders),
        }

    def route_area(row: dict[str, Any]) -> str:
        selected_area = str(row.get("khuVucDatXe") or "").strip()
        if selected_area:
            return selected_area.upper()
        plate_key = normalize_text(row.get("bienKiemSoat"))
        vehicle = roster_by_plate.get(plate_key) or franchise_by_plate.get(plate_key) or {}
        area = str(vehicle.get("khuVucHoatDong") or "").strip()
        if area:
            return area.upper()
        start, _ = split_route(row.get("tuyen"))
        return (start or "TỈNH KHÁC").strip().upper()

    def source_label(row: dict[str, Any]) -> str:
        customer = next((item for item in customers if normalize_phone(item.get("soDienThoai")) == normalize_phone(row.get("soDienThoai"))), {})
        return str(customer.get("nguonKhach") or "Không rõ").strip() or "Không rõ"

    def vehicle_info(row: dict[str, Any]) -> dict[str, Any]:
        plate_key = normalize_text(row.get("bienKiemSoat"))
        vehicle = franchise_by_plate.get(plate_key) or roster_by_plate.get(plate_key) or {}
        driver_text = roster_driver_text(vehicle)
        driver_name = str(row.get("hoTenLaiXe") or "").strip() or str(vehicle.get("hoTenLaiXe") or "").strip() or roster_driver_name(driver_text)
        driver_code = str(row.get("maNVLaiXe") or "").strip() or roster_driver_code(driver_text)
        return {
            "driver": driver_name,
            "driver_code": driver_code,
            "plate": row.get("bienKiemSoat") or vehicle.get("bienKiemSoat") or "",
            "code": row.get("soHieuXe") or vehicle.get("dongXe") or vehicle.get("soHieuXe") or "",
            "seats": vehicle.get("soCho") or vehicle.get("so_cho") or "",
            "type": vehicle.get("hieuXe") or vehicle.get("loaiXe") or vehicle.get("loai_xe") or "",
            "group": "Xe Thương Quyền" if is_franchise_order(row) else "Xe Công ty",
        }

    workbook = Workbook()
    workbook.remove(workbook.active)
    thin = Side(style="thin", color="222222")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="A9D18E")
    title_fill = PatternFill("solid", fgColor="FFFFFF")
    money_format = '#,##0'
    percent_format = '0.00%'

    def setup_sheet(title: str, headers: list[str], widths: list[int] | None = None):
        sheet = workbook.create_sheet(title[:31])
        sheet.sheet_view.showGridLines = False
        for column_index, header in enumerate(headers, start=1):
            cell = sheet.cell(2, column_index, header)
            cell.fill = header_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
            sheet.column_dimensions[get_column_letter(column_index)].width = (widths or [16] * len(headers))[column_index - 1]
        sheet.freeze_panes = "A3"
        return sheet

    def write_title(sheet, title: str, end_column: int) -> None:
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
        cell = sheet.cell(1, 1, title)
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = title_fill

    def append_row(sheet, row_index: int, values: list[Any], money_columns: set[int] | None = None, percent_columns: set[int] | None = None) -> None:
        for column_index, value in enumerate(values, start=1):
            cell = sheet.cell(row_index, column_index, value)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if money_columns and column_index in money_columns:
                cell.number_format = money_format
                cell.alignment = Alignment(horizontal="right", vertical="center")
            if percent_columns and column_index in percent_columns:
                cell.number_format = percent_format
                cell.alignment = Alignment(horizontal="right", vertical="center")

    def append_summary_total(
        sheet,
        row_index: int,
        label_end_column: int,
        sum_columns: list[int],
        percent_columns: list[int] | None = None,
    ) -> None:
        sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=label_end_column)
        label = sheet.cell(row_index, 1, "TỔNG CỘNG")
        label.alignment = Alignment(horizontal="right")
        data_start = 3
        data_end = row_index - 1
        for column in sum_columns:
            letter = get_column_letter(column)
            sheet.cell(row_index, column, f"=SUM({letter}{data_start}:{letter}{data_end})" if data_end >= data_start else 0)
        for column in percent_columns or []:
            sheet.cell(row_index, column, 1 if data_end >= data_start else 0)
            sheet.cell(row_index, column).number_format = percent_format
        for column in range(1, sheet.max_column + 1):
            cell = sheet.cell(row_index, column)
            cell.border = border
            cell.fill = header_fill
            cell.font = Font(bold=True)
            if column in sum_columns:
                cell.number_format = money_format if "doanh thu" in normalize_text(sheet.title) or column >= 10 else '#,##0'

    period_label = f"T{report_date.month}.{report_date.year}"

    overview_headers = ["STT", "Sản phẩm DV", "Xe Công ty", "Xe Thương quyền", "Tổng cộng"]
    overview = setup_sheet("Tong quan", overview_headers, [8, 28, 16, 18, 16])
    write_title(overview, f"BÁO CÁO LƯỢT KHÁCH {period_label}", len(overview_headers))
    categories = [("Xe Ghép", True), ("Xe Nguyên chiếc", False)]
    for index, (label, shared) in enumerate(categories, start=1):
        category_orders = [row for row in orders if is_shared_order(row) == shared]
        company_count = sum(1 for row in category_orders if is_company_order(row))
        franchise_count = sum(1 for row in category_orders if is_franchise_order(row))
        append_row(overview, index + 2, [index, label, company_count, franchise_count, company_count + franchise_count])
    append_row(overview, 5, ["", "Tỷ lệ lượt khách TQ", "", "=D6/E6", ""], percent_columns={4})
    append_row(overview, 6, ["", "Tổng cộng", "=SUM(C3:C4)", "=SUM(D3:D4)", "=SUM(E3:E4)"])

    revenue_sheet = setup_sheet("Tong doanh thu", overview_headers, [8, 28, 16, 18, 16])
    write_title(revenue_sheet, f"BÁO CÁO TỔNG DOANH THU {period_label}", len(overview_headers))
    for index, (label, shared) in enumerate(categories, start=1):
        category_orders = [row for row in orders if is_shared_order(row) == shared]
        company_revenue = sum(order_gross(row) for row in category_orders if is_company_order(row))
        franchise_revenue = sum(order_gross(row) for row in category_orders if is_franchise_order(row))
        append_row(revenue_sheet, index + 2, [index, label, company_revenue, franchise_revenue, company_revenue + franchise_revenue], money_columns={3, 4, 5})
    append_row(revenue_sheet, 5, ["", "Tỷ lệ doanh thu TQ", "", "=D6/E6", ""], percent_columns={4})
    append_row(revenue_sheet, 6, ["", "Tổng cộng", "=SUM(C3:C4)", "=SUM(D3:D4)", "=SUM(E3:E4)"], money_columns={3, 4, 5})

    area_headers = ["STT", "Khu vực đặt xe", "Tỷ lệ theo khu vực", "Xe Công ty", "Xe TQ", "Tổng số chuyến"]
    area_sheet = setup_sheet("Luot khach khu vuc", area_headers, [8, 24, 18, 16, 16, 18])
    write_title(area_sheet, f"BẢNG KÊ LƯỢT KHÁCH THEO KHU VỰC {period_label}", len(area_headers))
    areas = sorted({route_area(row) for row in orders})
    total_area_count = len(orders)
    for index, area in enumerate(areas, start=1):
        rows = [row for row in orders if route_area(row) == area]
        company_count = sum(1 for row in rows if is_company_order(row))
        franchise_count = sum(1 for row in rows if is_franchise_order(row))
        append_row(area_sheet, index + 2, [index, area, (len(rows) / total_area_count) if total_area_count else 0, company_count, franchise_count, len(rows)], percent_columns={3})
    append_summary_total(area_sheet, len(areas) + 3, 2, [4, 5, 6], [3])

    area_revenue_sheet = setup_sheet("Doanh thu khu vuc", ["STT", "Khu vực đặt xe", "Tỷ lệ theo khu vực", "DT Xe Công ty", "DT xe TQ", "Tổng Doanh thu"], [8, 24, 18, 18, 18, 20])
    write_title(area_revenue_sheet, f"BẢNG KÊ DOANH THU THEO KHU VỰC {period_label}", 6)
    total_revenue = sum(order_gross(row) for row in orders)
    for index, area in enumerate(areas, start=1):
        rows = [row for row in orders if route_area(row) == area]
        company_revenue = sum(order_gross(row) for row in rows if is_company_order(row))
        franchise_revenue = sum(order_gross(row) for row in rows if is_franchise_order(row))
        row_total = company_revenue + franchise_revenue
        append_row(area_revenue_sheet, index + 2, [index, area, (row_total / total_revenue) if total_revenue else 0, company_revenue, franchise_revenue, row_total], money_columns={4, 5, 6}, percent_columns={3})
    append_summary_total(area_revenue_sheet, len(areas) + 3, 2, [4, 5, 6], [3])

    source_sheet = setup_sheet("Nguon khach", ["STT", "Nguồn khách", "Tỷ lệ", "Xe ghép", "Xe Nguyên chiếc", "Tổng cộng"], [8, 24, 14, 14, 18, 16])
    write_title(source_sheet, f"BÁO CÁO LƯỢT KHÁCH THEO NGUỒN {period_label}", 6)
    report_order_ids = {str(row.get("id") or "") for row in orders}
    source_entries: list[tuple[str, bool]] = [
        (source_label(row), False) for row in orders if not is_shared_order(row)
    ]
    for shared_row in shared_report_rows:
        if str(shared_row.get("donHangId") or "") not in report_order_ids:
            continue
        source = str(shared_row.get("nguonKhach") or "").strip()
        if not source:
            customer = next(
                (
                    item
                    for item in customers
                    if normalize_phone(item.get("soDienThoai")) == normalize_phone(shared_row.get("soDienThoai"))
                ),
                {},
            )
            source = str(customer.get("nguonKhach") or "").strip()
        source_entries.append((source or "Không rõ", True))
    sources = sorted({source for source, _ in source_entries})
    for index, source in enumerate(sources, start=1):
        entries = [entry for entry in source_entries if entry[0] == source]
        shared_count = sum(1 for _, is_shared in entries if is_shared)
        private_count = len(entries) - shared_count
        append_row(
            source_sheet,
            index + 2,
            [index, source, (len(entries) / len(source_entries)) if source_entries else 0, shared_count, private_count, len(entries)],
            percent_columns={3},
        )
    append_summary_total(source_sheet, len(sources) + 3, 2, [4, 5, 6], [3])

    detail_headers = ["STT", "Lái xe", "Nhóm lái xe", "Biển kiểm soát", "Số chỗ", "Loại xe", "NGÀY", "Mã đơn hàng", "Khách hàng", "Tuyến", "DT Xe Nguyên chuyến", "DT Xe Ghép", "Cọc", "Ưu đãi", "Phải Thu", "Hoa hồng", "Ghi chú"]
    detail_sheet = setup_sheet("Doanh thu lai xe", detail_headers, [8, 24, 18, 18, 12, 16, 14, 24, 24, 28, 18, 16, 14, 14, 14, 16, 28])
    write_title(detail_sheet, f"BÁO CÁO TỔNG DOANH THU LÁI XE {period_label}", len(detail_headers))
    for index, row in enumerate(sorted(orders, key=lambda item: (str(item.get("hoTenLaiXe") or ""), order_start(item))), start=1):
        info = vehicle_info(row)
        is_shared = is_shared_order(row)
        commission = money_value(row.get("soTienNopLai")) if is_franchise_order(row) else ""
        append_row(
            detail_sheet,
            index + 2,
            [
                index,
                info["driver"],
                info["group"],
                info["plate"],
                info["seats"],
                info["type"],
                order_start(row).strftime("%d/%m/%Y") if order_start(row) != datetime.min else "",
                row.get("id") or "",
                "Khách xe ghép" if is_shared else row.get("tenKhach") or "",
                row.get("tuyen") or "",
                "" if is_shared else order_gross(row),
                order_gross(row) if is_shared else "",
                order_deposit(row) or "",
                order_discount(row) or "",
                order_due(row),
                commission,
                row.get("ghiChu") or "",
            ],
            money_columns={11, 12, 13, 14, 15, 16},
        )
    append_summary_total(detail_sheet, len(orders) + 3, 10, [11, 12, 13, 14, 15, 16])

    for sheet_name, title, predicate in [
        ("DT xe cong ty", "BÁO CÁO DOANH THU LÁI XE CÔNG TY", is_company_order),
        ("DT xe thuong quyen", "BÁO CÁO DOANH THU LÁI XE THƯỞNG QUYỀN", is_franchise_order),
        ("Doanh thu B2B", "BÁO CÁO DOANH THU B2B", lambda row: normalize_text(row.get("loaiKhach")) == "b2b"),
        ("Doanh thu B2C", "BÁO CÁO DOANH THU B2C", lambda row: normalize_text(row.get("loaiKhach")) == "b2c"),
    ]:
        sheet = setup_sheet(sheet_name, detail_headers, [8, 24, 18, 18, 12, 16, 14, 24, 24, 28, 18, 16, 14, 14, 14, 16, 28])
        write_title(sheet, f"{title} {period_label}", len(detail_headers))
        filtered = [row for row in orders if predicate(row)]
        for index, row in enumerate(filtered, start=1):
            info = vehicle_info(row)
            is_shared = is_shared_order(row)
            commission = money_value(row.get("soTienNopLai")) if is_franchise_order(row) else ""
            append_row(
                sheet,
                index + 2,
                [
                    index,
                    info["driver"],
                    info["group"],
                    info["plate"],
                    info["seats"],
                    info["type"],
                    order_start(row).strftime("%d/%m/%Y") if order_start(row) != datetime.min else "",
                    row.get("id") or "",
                    "Khách xe ghép" if is_shared else row.get("tenKhach") or "",
                    row.get("tuyen") or "",
                    "" if is_shared else order_gross(row),
                    order_gross(row) if is_shared else "",
                    order_deposit(row) or "",
                    order_discount(row) or "",
                    order_due(row),
                    commission,
                    row.get("ghiChu") or "",
                ],
                money_columns={11, 12, 13, 14, 15, 16},
            )
        append_summary_total(sheet, len(filtered) + 3, 10, [11, 12, 13, 14, 15, 16])

    customer_headers = ["STT", "Ngày tạo khách hàng", "Nguồn Khách", "Tên Khách Hàng", "Giới tính", "Ngày sinh", "SĐT", "Địa Chỉ", "Ngày giao dịch cuối", "Số lượt đơn hàng", "Tổng doanh thu", "Chú thích"]
    customer_sheet = setup_sheet("Danh sach khach hang", customer_headers, [8, 18, 18, 28, 14, 14, 18, 24, 18, 18, 18, 28])
    write_title(customer_sheet, f"DANH SÁCH KHÁCH HÀNG {period_label}", len(customer_headers))
    fresh_customer_address_by_phone = {
        normalize_phone(customer.get("soDienThoai")): str(customer.get("diaChi") or "").strip()
        for customer in customer_records()
        if normalize_phone(customer.get("soDienThoai")) and str(customer.get("diaChi") or "").strip()
    }
    for index, row in enumerate(customers, start=1):
        activity = customer_activity(row)
        customer_address = (
            str(row.get("diaChi") or "").strip()
            or str(activity.get("address") or "").strip()
            or customer_address_by_phone.get(normalize_phone(row.get("soDienThoai")), "")
            or fresh_customer_address_by_phone.get(normalize_phone(row.get("soDienThoai")), "")
            or raw_customer_address_by_phone.get(normalize_phone(row.get("soDienThoai")), "")
        )
        note_parts = [part for part in [row.get("loaiKhachHang"), row.get("soCCCD"), row.get("nhanVienNhap")] if part]
        append_row(
            customer_sheet,
            index + 2,
            [
                index,
                display_date(row.get("createdAt")),
                row.get("nguonKhach") or "",
                row.get("tenKhach") or "",
                row.get("gioiTinh") or "",
                row.get("namSinh") or "",
                row.get("soDienThoai") or "",
                customer_address,
                activity["last_date"],
                activity["count"],
                activity["revenue"],
                " | ".join(note_parts),
            ],
            money_columns={11},
        )
    append_summary_total(customer_sheet, len(customers) + 3, 9, [10, 11])
    for row_index in range(3, customer_sheet.max_row + 1):
        if str(customer_sheet.cell(row_index, 8).value or "").strip():
            continue
        phone = normalize_phone(customer_sheet.cell(row_index, 7).value)
        address = raw_customer_address_by_phone.get(phone) or fresh_customer_address_by_phone.get(phone) or customer_address_by_phone.get(phone, "")
        if address:
            customer_sheet.cell(row_index, 8, address)

    franchise_headers = ["STT", "Lái xe", "BKS", "Số chỗ", "Hiệu xe", "Chủ xe / đơn vị", "SĐT", "Địa chỉ lái xe", "Ngày giao dịch cuối", "Số lượt đơn hàng", "Tổng doanh thu", "Chú thích"]
    franchise_sheet = setup_sheet("Danh sach xe TQ", franchise_headers, [8, 24, 16, 12, 16, 28, 18, 24, 18, 18, 18, 28])
    write_title(franchise_sheet, "DANH SÁCH XE THƯỞNG QUYỀN", len(franchise_headers))
    for index, row in enumerate(franchise_rows, start=1):
        activity = franchise_activity(row)
        append_row(
            franchise_sheet,
            index + 2,
            [
                index,
                row.get("hoTenLaiXe") or "",
                row.get("bienKiemSoat") or "",
                row.get("soCho") or "",
                row.get("hieuXe") or "",
                row.get("tenChuXe") or "",
                row.get("soDienThoaiLaiXe") or row.get("soDienThoaiChuXe") or "",
                row.get("diaChiLaiXe") or "",
                activity["last_date"],
                activity["count"],
                activity["revenue"],
                row.get("ghiChu") or row.get("trangThai") or "",
            ],
            money_columns={11},
        )
    append_summary_total(franchise_sheet, len(franchise_rows) + 3, 9, [10, 11])

    combined_sheet_names = [
        "Tong quan",
        "Tong doanh thu",
        "Luot khach khu vuc",
        "Doanh thu khu vuc",
        "Nguon khach",
        "Danh sach khach hang",
        "Danh sach xe TQ",
    ]

    def copy_sheet_section(source, target, start_row: int) -> int:
        from copy import copy
        from openpyxl.formula.translate import Translator

        for column_index in range(1, source.max_column + 1):
            letter = get_column_letter(column_index)
            source_width = source.column_dimensions[letter].width or 12
            target_width = target.column_dimensions[letter].width or 0
            target.column_dimensions[letter].width = max(target_width, source_width)

        for merged_range in source.merged_cells.ranges:
            target.merge_cells(
                start_row=start_row + merged_range.min_row - 1,
                start_column=merged_range.min_col,
                end_row=start_row + merged_range.max_row - 1,
                end_column=merged_range.max_col,
            )

        for row_index in range(1, source.max_row + 1):
            target_row = start_row + row_index - 1
            target.row_dimensions[target_row].height = source.row_dimensions[row_index].height
            for column_index in range(1, source.max_column + 1):
                source_cell = source.cell(row_index, column_index)
                target_coordinate = f"{get_column_letter(column_index)}{target_row}"
                value = source_cell.value
                if isinstance(value, str) and value.startswith("="):
                    value = Translator(value, origin=source_cell.coordinate).translate_formula(target_coordinate)
                target_cell = target.cell(target_row, column_index, value)
                if source_cell.has_style:
                    target_cell.font = copy(source_cell.font)
                    target_cell.fill = copy(source_cell.fill)
                    target_cell.border = copy(source_cell.border)
                    target_cell.alignment = copy(source_cell.alignment)
                    target_cell.number_format = source_cell.number_format
                    target_cell.protection = copy(source_cell.protection)
        return start_row + source.max_row + 2

    combined_sheet = workbook.create_sheet("Tong hop", 0)
    combined_sheet.sheet_view.showGridLines = False
    next_section_row = 1
    for sheet_name in combined_sheet_names:
        if sheet_name in workbook.sheetnames:
            next_section_row = copy_sheet_section(workbook[sheet_name], combined_sheet, next_section_row)
    combined_sheet.page_setup.orientation = "landscape"
    combined_sheet.page_setup.fitToWidth = 1
    combined_sheet.page_setup.fitToHeight = 0
    combined_sheet.sheet_properties.pageSetUpPr.fitToPage = True
    for sheet_name in combined_sheet_names:
        if sheet_name in workbook.sheetnames:
            workbook.remove(workbook[sheet_name])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"bao-cao-tong-hop-{report_date.strftime('%Y%m')}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/reports/customers.xlsx")
def export_customers_report() -> Response:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover
        raise HTTPException(status_code=500, detail="Thiếu thư viện xuất Excel.") from exc

    customers = customer_records()
    orders = all_order_records()
    shared_passengers = all_shared_ride_records()

    def customer_key(row: dict[str, Any]) -> str:
        customer_id = str(row.get("id") or row.get("khachHangId") or "").strip()
        if customer_id:
            return f"id:{customer_id}"
        phone = normalize_phone(row.get("soDienThoai"))
        return f"phone:{phone}" if phone else ""

    def order_customer_key(row: dict[str, Any]) -> str:
        customer_id = str(row.get("khachHangId") or "").strip()
        if customer_id:
            return f"id:{customer_id}"
        phone = normalize_phone(row.get("soDienThoai"))
        return f"phone:{phone}" if phone else ""

    def order_start(row: dict[str, Any]) -> datetime | None:
        return parse_existing_datetime(row.get("ngayGioDi")) or parse_existing_datetime(row.get("createdAt"))

    def display_datetime(value: Any) -> str:
        parsed = parse_existing_datetime(value)
        return parsed.strftime("%d/%m/%Y %H:%M") if parsed else str(value or "")

    def is_later_datetime(candidate: datetime, current: datetime | None) -> bool:
        if current is None:
            return True
        candidate_compare = candidate.replace(tzinfo=timezone.utc) if candidate.tzinfo is None else candidate.astimezone(timezone.utc)
        current_compare = current.replace(tzinfo=timezone.utc) if current.tzinfo is None else current.astimezone(timezone.utc)
        return candidate_compare > current_compare

    activity: dict[str, dict[str, Any]] = {}
    for order in orders:
        key = order_customer_key(order)
        if not key:
            continue
        current = activity.setdefault(key, {"count": 0, "revenue": 0.0, "last": None, "address": ""})
        current["count"] += 1
        current["revenue"] += money_value(order.get("giaTien")) + money_value(order.get("phuThu")) - money_value(order.get("giamGia")) - money_value(order.get("tongUuDai"))
        started_at = order_start(order)
        if started_at and is_later_datetime(started_at, current["last"]):
            current["last"] = started_at
        if not current["address"] and str(order.get("diaChiHoaDon") or "").strip():
            current["address"] = str(order.get("diaChiHoaDon") or "").strip()

    customers_by_phone = {
        normalize_phone(customer.get("soDienThoai")): customer
        for customer in customers
        if normalize_phone(customer.get("soDienThoai"))
    }
    for passenger in shared_passengers:
        phone = normalize_phone(passenger.get("soDienThoai"))
        customer = customers_by_phone.get(phone)
        key = customer_key(customer) if customer else (f"phone:{phone}" if phone else "")
        if not key:
            continue
        current = activity.setdefault(key, {"count": 0, "revenue": 0.0, "last": None, "address": ""})
        current["count"] += 1
        current["revenue"] += money_value(passenger.get("thucThu"))
        started_at = order_start(passenger)
        if started_at and is_later_datetime(started_at, current["last"]):
            current["last"] = started_at
        if not current["address"] and str(passenger.get("diaChi") or "").strip():
            current["address"] = str(passenger.get("diaChi") or "").strip()

    customer_address_by_phone: dict[str, str] = {}
    for customer in customers:
        phone_key = normalize_phone(customer.get("soDienThoai"))
        address = str(customer.get("diaChi") or "").strip()
        if phone_key and address and not customer_address_by_phone.get(phone_key):
            customer_address_by_phone[phone_key] = address

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Danh sach khach hang"
    headers = [
        "STT",
        "Mã khách",
        "Tên khách hàng",
        "Số điện thoại",
        "CCCD",
        "Địa chỉ",
        "Loại khách",
        "Năm sinh",
        "Giới tính",
        "Nguồn khách",
        "Nhân viên nhập",
        "Ngày tạo",
        "Ngày giao dịch cuối",
        "Số đơn hàng",
        "Tổng doanh thu",
    ]
    widths = [8, 18, 28, 18, 18, 32, 18, 12, 12, 18, 20, 18, 20, 14, 18]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    title = "DANH SÁCH KHÁCH HÀNG"
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    sheet.cell(row=1, column=1, value=title)
    sheet.cell(row=1, column=1).font = Font(bold=True, size=15, color="0F172A")
    sheet.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    header_fill = PatternFill("solid", fgColor="DDEBFF")
    border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=3, column=column, value=header)
        cell.font = Font(bold=True, color="0F172A")
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_index, customer in enumerate(customers, start=4):
        item = activity.get(customer_key(customer), {})
        phone_key = normalize_phone(customer.get("soDienThoai"))
        last_transaction = item.get("last")
        values = [
            row_index - 3,
            customer.get("id", ""),
            customer.get("tenKhach", ""),
            customer.get("soDienThoai", ""),
            customer.get("soCCCD", ""),
            customer.get("diaChi", "") or customer_address_by_phone.get(phone_key, "") or item.get("address", ""),
            customer.get("loaiKhachHang", ""),
            customer.get("namSinh", ""),
            customer.get("gioiTinh", ""),
            customer.get("nguonKhach", ""),
            customer.get("nhanVienNhap", ""),
            display_datetime(customer.get("createdAt")),
            last_transaction.strftime("%d/%m/%Y %H:%M") if last_transaction else "",
            item.get("count", 0),
            item.get("revenue", 0),
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=column, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if column in {1, 8, 9, 14}:
                cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
            if column == 15:
                cell.number_format = '#,##0'

    total_row = len(customers) + 4
    sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=13)
    sheet.cell(total_row, 1, "TỔNG CỘNG").alignment = Alignment(horizontal="right")
    sheet.cell(total_row, 14, f"=SUM(N4:N{total_row - 1})" if customers else 0)
    sheet.cell(total_row, 15, f"=SUM(O4:O{total_row - 1})" if customers else 0)
    sheet.cell(total_row, 15).number_format = '#,##0'
    for column in range(1, len(headers) + 1):
        cell = sheet.cell(total_row, column)
        cell.font = Font(bold=True, color="0F172A")
        cell.fill = header_fill
        cell.border = border

    sheet.freeze_panes = "A4"
    sheet.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{max(len(customers) + 3, 3)}"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"danh-sach-khach-hang-{datetime.now().strftime('%Y%m%d')}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/reports/debts.xlsx")
def export_debts_report(tuNgay: str = "", denNgay: str = "", ngay: str = "") -> Response:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover
        raise HTTPException(status_code=500, detail="Thiếu thư viện xuất Excel.") from exc

    start_date, end_date = selected_report_range(tuNgay, denNgay, ngay)
    debt_orders = rows_for_departure_range(debt_order_rows(), start_date, end_date)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Cong no"
    headers = [
        "STT", "Mã đơn", "Ngày đi", "Khách hàng", "Số điện thoại", "Tuyến",
        "Đối tượng ghi nhận công nợ", "Giá tiền", "Giảm giá thủ công", "Voucher/khuyến mãi",
        "VAT 8%", "Đã cọc", "Còn công nợ", "Trạng thái thu hồi", "Ngày thu hồi", "Người xác nhận", "Ghi chú",
    ]
    widths = [7, 24, 18, 24, 17, 28, 28, 15, 18, 20, 14, 14, 16, 20, 20, 22, 32]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title = sheet.cell(1, 1, "BÁO CÁO CÔNG NỢ ĐƠN HÀNG")
    title.font = Font(bold=True, size=15)
    title.alignment = Alignment(horizontal="center")
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    subtitle = sheet.cell(2, 1, f"Từ {start_date.strftime('%d/%m/%Y')} đến {end_date.strftime('%d/%m/%Y')}")
    subtitle.alignment = Alignment(horizontal="center")
    fill = PatternFill("solid", fgColor="DDEBFF")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(3, column, header)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for index, row in enumerate(debt_orders, start=1):
        gross = money_value(row.get("giaTien"))
        manual = money_value(row.get("giamGia"))
        benefits = money_value(row.get("tongUuDai"))
        vat = money_value(row.get("thueVAT"))
        deposit = money_value(row.get("daCoc"))
        outstanding = max(gross + money_value(row.get("phuThu")) - manual - benefits + vat - deposit, 0)
        started = parse_existing_datetime(row.get("ngayGioDi"))
        recovered = parse_existing_datetime(row.get("ngayThuHoiCongNo"))
        values = [
            index,
            row.get("orderCode") or row.get("id") or "",
            started.strftime("%d/%m/%Y %H:%M") if started else row.get("ngayGioDi") or "",
            row.get("tenKhach") or "",
            row.get("soDienThoai") or "",
            row.get("tuyen") or route_text(row.get("diemDon"), row.get("diemTra")),
            row.get("congNoChoAi") or "",
            gross,
            manual,
            benefits,
            vat,
            deposit,
            outstanding,
            debt_status(row),
            recovered.strftime("%d/%m/%Y %H:%M") if recovered else "",
            row.get("nguoiThuHoiCongNo") or "",
            row.get("ghiChu") or "",
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(index + 3, column, excel_safe_value(value))
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if 8 <= column <= 13:
                cell.number_format = '#,##0'
    total_row = len(debt_orders) + 4
    sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=7)
    total_label = sheet.cell(total_row, 1, "TỔNG CỘNG")
    total_label.font = Font(bold=True)
    total_label.alignment = Alignment(horizontal="right")
    debt_totals = {
        8: sum(money_value(row.get("giaTien")) for row in debt_orders),
        9: sum(money_value(row.get("giamGia")) for row in debt_orders),
        10: sum(money_value(row.get("tongUuDai")) for row in debt_orders),
        11: sum(money_value(row.get("thueVAT")) for row in debt_orders),
        12: sum(money_value(row.get("daCoc")) for row in debt_orders),
        13: sum(max(money_value(row.get("giaTien")) + money_value(row.get("phuThu")) - money_value(row.get("giamGia")) - money_value(row.get("tongUuDai")) + money_value(row.get("thueVAT")) - money_value(row.get("daCoc")), 0) for row in debt_orders),
    }
    for column, total in debt_totals.items():
        total_cell = sheet.cell(total_row, column, total)
        total_cell.font = Font(bold=True)
        total_cell.number_format = '#,##0'
    for column in range(1, len(headers) + 1):
        sheet.cell(total_row, column).border = border
        sheet.cell(total_row, column).fill = fill
    sheet.freeze_panes = "A4"
    sheet.auto_filter.ref = f"A3:Q{max(len(debt_orders) + 3, 3)}"
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"bao-cao-cong-no-{start_date.strftime('%Y%m%d')}-{end_date.strftime('%Y%m%d')}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/reports/commissions.xlsx")
def export_commissions_report(tuNgay: str = "", denNgay: str = "", ngay: str = "") -> Response:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover
        raise HTTPException(status_code=500, detail="Thiếu thư viện xuất Excel.") from exc

    start_date, end_date = selected_report_range(tuNgay, denNgay, ngay)
    rows = rows_for_departure_range(commission_order_rows(), start_date, end_date)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Hoa hong xe TQ"
    headers = [
        "STT", "Mã đơn", "Ngày giờ đi", "Khách hàng", "Số điện thoại", "Tuyến",
        "Biển kiểm soát", "Lái xe", "Doanh thu tính hoa hồng", "Tỷ lệ hoa hồng (%)",
        "Số tiền hoa hồng", "Trạng thái thu", "Ngày thu", "Người xác nhận", "Ghi chú",
    ]
    widths = [7, 24, 20, 24, 17, 28, 16, 24, 22, 20, 20, 18, 20, 22, 32]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title = sheet.cell(1, 1, "BÁO CÁO HOA HỒNG XE THƯƠNG QUYỀN")
    title.font = Font(bold=True, size=15)
    title.alignment = Alignment(horizontal="center")
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    subtitle = sheet.cell(2, 1, f"Từ {start_date.strftime('%d/%m/%Y')} đến {end_date.strftime('%d/%m/%Y')}")
    subtitle.alignment = Alignment(horizontal="center")
    header_fill = PatternFill("solid", fgColor="E0E7FF")
    paid_fill = PatternFill("solid", fgColor="DCFCE7")
    pending_fill = PatternFill("solid", fgColor="FEF3C7")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(3, column, header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for index, row in enumerate(rows, start=1):
        started = parse_existing_datetime(row.get("ngayGioDi"))
        collected = parse_existing_datetime(row.get("ngayThuHoaHong"))
        rate = money_value(row.get("tyLeNopLai"))
        commission = money_value(row.get("soTienNopLai"))
        revenue = commission * 100 / rate if rate > 0 else 0
        status = commission_status(row)
        values = [
            index,
            row.get("id") or "",
            started.strftime("%d/%m/%Y %H:%M") if started else row.get("ngayGioDi") or "",
            row.get("tenKhach") or "",
            row.get("soDienThoai") or "",
            row.get("tuyen") or route_text(row.get("diemDon"), row.get("diemTra")),
            row.get("bienKiemSoat") or "",
            row.get("hoTenLaiXe") or "",
            revenue,
            rate,
            commission,
            status,
            collected.strftime("%d/%m/%Y %H:%M") if collected else "",
            row.get("nguoiThuHoaHong") or "",
            row.get("ghiChu") or "",
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(index + 3, column, excel_safe_value(value))
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if column in {9, 10, 11}:
                cell.number_format = '#,##0'
            if column == 12:
                paid = normalize_text(status) == "da thu"
                cell.fill = paid_fill if paid else pending_fill
                cell.font = Font(bold=True, color="065F46" if paid else "92400E")
    total_row = len(rows) + 4
    sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=8)
    total_label = sheet.cell(total_row, 1, "TỔNG CỘNG")
    total_label.font = Font(bold=True)
    total_label.alignment = Alignment(horizontal="right")
    commission_totals = {
        9: sum((money_value(row.get("soTienNopLai")) * 100 / money_value(row.get("tyLeNopLai"))) if money_value(row.get("tyLeNopLai")) > 0 else 0 for row in rows),
        10: sum(money_value(row.get("tyLeNopLai")) for row in rows),
        11: sum(money_value(row.get("soTienNopLai")) for row in rows),
    }
    for column, total in commission_totals.items():
        total_cell = sheet.cell(total_row, column, total)
        total_cell.font = Font(bold=True)
        total_cell.number_format = '#,##0'
    for column in range(1, len(headers) + 1):
        sheet.cell(total_row, column).border = border
        sheet.cell(total_row, column).fill = header_fill
    sheet.freeze_panes = "A4"
    sheet.auto_filter.ref = f"A3:O{max(len(rows) + 3, 3)}"
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"bao-cao-hoa-hong-xe-thuong-quyen-{start_date.strftime('%Y%m%d')}-{end_date.strftime('%Y%m%d')}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/reports/invoices.xlsx")
def export_invoices_report(tuNgay: str = "", denNgay: str = "", ngay: str = "") -> Response:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover
        raise HTTPException(status_code=500, detail="Thiếu thư viện xuất Excel.") from exc

    start_date, end_date = selected_report_range(tuNgay, denNgay, ngay)
    rows = rows_for_departure_range(invoice_rows(), start_date, end_date)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Don can xuat hoa don"
    headers = [
        "STT",
        "Mã đơn",
        "Khách hàng",
        "SĐT",
        "Tuyến",
        "Ngày giờ đi",
        "Yêu cầu HĐ",
        "Tên công ty",
        "Mã số thuế",
        "Địa chỉ hóa đơn",
        "Email hóa đơn",
        "Thành tiền trước VAT",
        "Thuế VAT",
        "Tổng thanh toán",
        "Trạng thái hóa đơn",
        "Ngày xuất",
        "Người xuất",
        "Mã nhóm hóa đơn",
        "Ghi chú",
    ]
    widths = [8, 22, 28, 18, 28, 20, 16, 28, 18, 34, 28, 18, 16, 18, 20, 20, 22, 24, 34]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    sheet.cell(row=1, column=1, value="DANH SÁCH ĐƠN CẦN XUẤT HÓA ĐƠN")
    sheet.cell(row=1, column=1).font = Font(bold=True, size=15, color="0F172A")
    sheet.cell(row=1, column=1).alignment = Alignment(horizontal="center")
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    sheet.cell(row=2, column=1, value=f"Từ {start_date.strftime('%d/%m/%Y')} đến {end_date.strftime('%d/%m/%Y')}")
    sheet.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    header_fill = PatternFill("solid", fgColor="E0F2FE")
    issued_fill = PatternFill("solid", fgColor="DCFCE7")
    pending_fill = PatternFill("solid", fgColor="FEF3C7")
    border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=3, column=column, value=header)
        cell.font = Font(bold=True, color="0F172A")
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def display_datetime(value: Any) -> str:
        parsed = parse_existing_datetime(value)
        return parsed.strftime("%d/%m/%Y %H:%M") if parsed else str(value or "")

    def invoice_amounts(row: dict[str, Any]) -> tuple[float, float, float]:
        if row.get("invoiceEntityType") == "invoiceGroup":
            vat = money_value(row.get("tongVAT"))
            total = money_value(row.get("tongThanhToan")) or money_value(row.get("giaTien"))
            before_vat = money_value(row.get("tongTruocVAT")) or max(total - vat, 0)
            return before_vat, vat, total
        before_vat = max(
            money_value(row.get("giaTien")) + money_value(row.get("phuThu"))
            - money_value(row.get("giamGia")) - money_value(row.get("tongUuDai")),
            0,
        )
        vat = money_value(row.get("thueVAT"))
        total = money_value(row.get("tongThanhToan")) or before_vat + vat
        return before_vat, vat, total

    for index, row in enumerate(rows, start=1):
        status = invoice_status(row)
        is_issued = normalize_text(status) == "da xuat"
        before_vat, vat, total_payment = invoice_amounts(row)
        values = [
            index,
            row.get("orderCode") or row.get("id") or "",
            row.get("tenKhach") or "",
            row.get("soDienThoai") or "",
            row.get("tuyen") or "",
            display_datetime(row.get("ngayGioDi")),
            row.get("yeuCauHoaDon") or "",
            row.get("tenCongTy") or "",
            row.get("maSoThue") or "",
            row.get("diaChiHoaDon") or "",
            row.get("emailHoaDon") or "",
            before_vat,
            vat,
            total_payment,
            status,
            display_datetime(row.get("ngayXuatHoaDon")),
            row.get("nguoiXuatHoaDon") or "",
            row.get("nhomHoaDonId") or "",
            row.get("ghiChu") or "",
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=index + 3, column=column, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if column in {12, 13, 14}:
                cell.number_format = '#,##0'
            if column == 15:
                cell.fill = issued_fill if is_issued else pending_fill
                cell.font = Font(bold=True, color="065F46" if is_issued else "92400E")

    total_row = len(rows) + 4
    sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=11)
    total_label = sheet.cell(total_row, 1, "TỔNG CỘNG")
    total_label.font = Font(bold=True)
    total_label.alignment = Alignment(horizontal="right")
    for column, value in enumerate(
        [
            sum(invoice_amounts(row)[0] for row in rows),
            sum(invoice_amounts(row)[1] for row in rows),
            sum(invoice_amounts(row)[2] for row in rows),
        ],
        start=12,
    ):
        total_cell = sheet.cell(total_row, column, value)
        total_cell.font = Font(bold=True)
        total_cell.number_format = '#,##0'
    for column in range(1, len(headers) + 1):
        sheet.cell(total_row, column).border = border
        sheet.cell(total_row, column).fill = header_fill

    sheet.freeze_panes = "A4"
    sheet.auto_filter.ref = f"A3:S{max(len(rows) + 3, 3)}"
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"danh-sach-hoa-don-{start_date.strftime('%Y%m%d')}-{end_date.strftime('%Y%m%d')}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/reports/vouchers.xlsx")
def export_voucher_report(thang: str = "") -> Response:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover
        raise HTTPException(status_code=500, detail="Thiếu thư viện xuất Excel.") from exc

    if thang:
        try:
            report_date = datetime.strptime(thang, "%Y-%m")
        except ValueError:
            raise HTTPException(status_code=422, detail="Tháng báo cáo không hợp lệ.")
    else:
        report_date = datetime.now()

    month_start = report_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    next_month = (month_start.replace(day=28) + timedelta(days=4)).replace(day=1)
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

    vouchers = worksheet_records(vouchers_worksheet(), VOUCHER_HEADERS)
    for voucher in vouchers:
        voucher["tenVoucher"] = voucher_campaign_name(voucher.get("tenVoucher"))
    usages = [
        row
        for row in order_benefit_records()
        if normalize_text(row.get("loaiUuDai")) == "voucher"
    ]
    orders = all_order_records()
    shared_rows = all_shared_ride_records()
    orders_by_id = {str(row.get("id") or "").strip(): row for row in orders}
    shared_by_order: dict[str, list[dict[str, Any]]] = {}
    for row in shared_rows:
        order_id = str(row.get("donHangId") or "").strip()
        if order_id:
            shared_by_order.setdefault(order_id, []).append(row)

    def display_date(value: Any) -> str:
        parsed = parse_existing_date(value)
        return parsed.strftime("%d/%m/%Y") if parsed else str(value or "")

    def display_datetime(value: Any) -> str:
        parsed = parse_existing_datetime(value)
        return parsed.strftime("%d/%m/%Y %H:%M") if parsed else str(value or "")

    def voucher_usages(voucher: dict[str, Any]) -> list[dict[str, Any]]:
        voucher_code = normalize_text(voucher.get("maVoucher"))
        voucher_id = normalize_text(voucher.get("id"))
        matched: list[dict[str, Any]] = []
        for usage in usages:
            usage_code = normalize_text(usage.get("maUuDai"))
            usage_id = normalize_text(usage.get("uuDaiId"))
            if voucher_code and usage_code:
                if voucher_code == usage_code:
                    matched.append(usage)
                continue
            if voucher_id and usage_id and voucher_id == usage_id:
                matched.append(usage)
        return matched

    def voucher_status(voucher: dict[str, Any], used: bool) -> str:
        if used:
            return "Đã sử dụng"
        status = normalize_text(voucher.get("trangThai"))
        if "ngung" in status or "tam ngung" in status:
            return "Tạm ngưng"
        start = parse_existing_date(voucher.get("ngayBatDau"))
        end = parse_existing_date(voucher.get("ngayHetHan"))
        if start and today < start:
            return "Chưa đến hạn"
        if end and today > end:
            return "Hết hạn"
        return "Còn hạn - chưa sử dụng"

    def format_value(row: dict[str, Any]) -> str:
        value = money_value(row.get("giaTri"))
        if normalize_text(row.get("loaiGiaTri")) in {"percent", "phan tram", "%"}:
            return f"{value:g}%"
        return value

    def usage_order(usage: dict[str, Any]) -> dict[str, Any]:
        return orders_by_id.get(str(usage.get("donHangId") or "").strip(), {})

    def usage_date(usage: dict[str, Any]) -> datetime | None:
        order = usage_order(usage)
        return parse_existing_datetime(order.get("ngayGioDi")) or parse_existing_datetime(usage.get("createdAt"))

    def in_report_month(usage: dict[str, Any]) -> bool:
        date_value = usage_date(usage)
        return bool(date_value and month_start <= date_value < next_month)

    def shared_customer_line(order_id: str, customer_name: str) -> dict[str, Any]:
        normalized_name = normalize_text(customer_name)
        for row in shared_by_order.get(order_id, []):
            if normalized_name and normalize_text(row.get("hoTen")) == normalized_name:
                return row
        return {}

    report_rows: list[dict[str, Any]] = []
    detail_rows: list[dict[str, Any]] = []
    for voucher in vouchers:
        matched_usages = voucher_usages(voucher)
        status = voucher_status(voucher, bool(matched_usages))
        report_rows.append({"voucher": voucher, "uses": matched_usages, "status": status})
        for usage in matched_usages:
            order = usage_order(usage)
            order_id = str(usage.get("donHangId") or "").strip()
            shared_line = shared_customer_line(order_id, str(usage.get("tenKhach") or ""))
            detail_rows.append(
                {
                    "voucher": voucher,
                    "usage": usage,
                    "order": order,
                    "shared": shared_line,
                    "status": status,
                    "usageDate": usage_date(usage),
                }
            )

    workbook = Workbook()
    workbook.remove(workbook.active)
    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    title_fill = PatternFill("solid", fgColor="0F172A")
    total_fill = PatternFill("solid", fgColor="ECFDF5")
    warning_fill = PatternFill("solid", fgColor="FEF3C7")
    danger_fill = PatternFill("solid", fgColor="FEE2E2")
    used_fill = PatternFill("solid", fgColor="DBEAFE")
    money_format = '#,##0'

    def setup_sheet(title: str, headers: list[str], widths: list[int] | None = None):
        sheet = workbook.create_sheet(title[:31])
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A3"
        for column_index, header in enumerate(headers, start=1):
            cell = sheet.cell(2, column_index, header)
            cell.fill = header_fill
            cell.font = Font(bold=True, color="0F172A")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
            sheet.column_dimensions[get_column_letter(column_index)].width = (widths or [16] * len(headers))[column_index - 1]
        return sheet

    def write_title(sheet, title: str, end_column: int) -> None:
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
        cell = sheet.cell(1, 1, title)
        cell.fill = title_fill
        cell.font = Font(bold=True, size=14, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[1].height = 24

    def append_row(sheet, row_index: int, values: list[Any], money_columns: set[int] | None = None, fill=None) -> None:
        for column_index, value in enumerate(values, start=1):
            cell = sheet.cell(row_index, column_index, value)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if fill:
                cell.fill = fill
            if money_columns and column_index in money_columns:
                cell.number_format = money_format
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)

    def status_fill(status: str):
        normalized = normalize_text(status)
        if "da su dung" in normalized:
            return used_fill
        if "het han" in normalized or "tam ngung" in normalized:
            return danger_fill
        if "chua den han" in normalized:
            return warning_fill
        return total_fill

    def write_voucher_sheet(title: str, rows: list[dict[str, Any]]) -> None:
        display_titles = {
            "Tat ca voucher": "TẤT CẢ VOUCHER",
            "Con han chua dung": "CÒN HẠN CHƯA DÙNG",
            "Het han": "HẾT HẠN",
            "Da su dung": "ĐÃ SỬ DỤNG",
        }
        headers = [
            "STT",
            "Mã voucher",
            "Tên chiến dịch",
            "Loại giá trị",
            "Giá trị",
            "Ngày bắt đầu",
            "Ngày hết hạn",
            "Trạng thái",
            "Đã dùng cho đơn",
            "Khách sử dụng",
            "Số tiền giảm",
            "Ghi chú",
        ]
        sheet = setup_sheet(title, headers, [8, 18, 30, 14, 14, 16, 16, 22, 24, 24, 16, 32])
        write_title(sheet, f"BÁO CÁO VOUCHER - {display_titles.get(title, title.upper())}", len(headers))
        for index, item in enumerate(rows, start=1):
            voucher = item["voucher"]
            uses = item["uses"]
            used_orders = ", ".join(str(row.get("donHangId") or "") for row in uses if row.get("donHangId"))
            used_customers = ", ".join(str(row.get("tenKhach") or "") for row in uses if row.get("tenKhach"))
            discount_total = sum(money_value(row.get("soTienGiam")) for row in uses)
            append_row(
                sheet,
                index + 2,
                [
                    index,
                    voucher.get("maVoucher") or voucher.get("id") or "",
                    voucher.get("tenVoucher") or "",
                    voucher.get("loaiGiaTri") or "",
                    format_value(voucher),
                    display_date(voucher.get("ngayBatDau")),
                    display_date(voucher.get("ngayHetHan")),
                    item["status"],
                    used_orders,
                    used_customers,
                    discount_total,
                    voucher.get("ghiChu") or "",
                ],
                money_columns={11},
                fill=status_fill(item["status"]),
            )
        total_row = len(rows) + 3
        sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=10)
        sheet.cell(total_row, 1, f"TỔNG CỘNG ({len(rows)} voucher)").alignment = Alignment(horizontal="right")
        sheet.cell(total_row, 11, sum(sum(money_value(usage.get("soTienGiam")) for usage in item["uses"]) for item in rows))
        for column in range(1, 13):
            cell = sheet.cell(total_row, column)
            cell.border = border
            cell.fill = total_fill
            cell.font = Font(bold=True)
            if column == 11:
                cell.number_format = money_format

    period_label = f"{report_date.month:02d}/{report_date.year}"
    total_vouchers = len(report_rows)
    used_vouchers = [row for row in report_rows if row["uses"]]
    expired_vouchers = [row for row in report_rows if normalize_text(row["status"]) == "het han"]
    available_vouchers = [row for row in report_rows if normalize_text(row["status"]) == "con han - chua su dung"]
    paused_vouchers = [row for row in report_rows if normalize_text(row["status"]) in {"tam ngung", "chua den han"}]
    period_usages = [row for row in detail_rows if row["usageDate"] and month_start <= row["usageDate"] < next_month]

    overview = setup_sheet("Tong quan", ["Chỉ tiêu", "Số lượng", "Ghi chú"], [32, 16, 48])
    write_title(overview, f"TỔNG QUAN VOUCHER - Kỳ {period_label}", 3)
    overview_items = [
        ("Tổng voucher", total_vouchers, "Tất cả mã voucher trên hệ thống"),
        ("Còn hạn - chưa sử dụng", len(available_vouchers), "Có thể áp vào đơn hàng"),
        ("Đã sử dụng", len(used_vouchers), "Đã gắn vào ít nhất một đơn hàng"),
        ("Hết hạn", len(expired_vouchers), "Quá ngày hết hạn"),
        ("Tạm ngưng / chưa đến hạn", len(paused_vouchers), "Chưa nên áp vào đơn"),
        ("Lượt sử dụng trong kỳ", len(period_usages), f"Kỳ {period_label}"),
        ("Tổng tiền giảm trong kỳ", sum(money_value(row["usage"].get("soTienGiam")) for row in period_usages), f"Kỳ {period_label}"),
    ]
    for index, (label, value, note) in enumerate(overview_items, start=3):
        append_row(overview, index, [label, value, note], money_columns={2} if "tien" in normalize_text(label) else None)
    overview_total_row = len(overview_items) + 3
    append_row(
        overview,
        overview_total_row,
        [
            "TỔNG CỘNG",
            total_vouchers,
            f"Tổng tiền giảm trong kỳ: {sum(money_value(row['usage'].get('soTienGiam')) for row in period_usages):,.0f}",
        ],
        fill=total_fill,
    )
    for cell in overview[overview_total_row]:
        cell.font = Font(bold=True)

    write_voucher_sheet("Tat ca voucher", report_rows)
    write_voucher_sheet("Con han chua dung", available_vouchers)
    write_voucher_sheet("Het han", expired_vouchers)
    write_voucher_sheet("Da su dung", used_vouchers)

    detail_headers = [
        "STT",
        "Mã voucher",
        "Tên chiến dịch",
        "Ngày sử dụng",
        "Mã đơn hàng",
        "Khách sử dụng",
        "SĐT khách",
        "Loại đơn",
        "Tuyến",
        "Điểm đón",
        "Điểm trả",
        "Giờ đi",
        "Xe",
        "Lái xe",
        "Giá tiền",
        "Giảm giá thủ công",
        "Ưu đãi đơn",
        "Voucher giảm",
        "Thực thu",
        "Trạng thái đơn",
    ]
    widths = [8, 18, 28, 18, 24, 24, 16, 18, 24, 22, 22, 18, 16, 24, 16, 18, 16, 16, 16, 18]

    def write_detail_sheet(title: str, rows: list[dict[str, Any]]) -> None:
        display_titles = {
            "Su dung trong ky": "SỬ DỤNG TRONG KỲ",
            "Chi tiet don hang": "CHI TIẾT ĐƠN HÀNG",
        }
        sheet = setup_sheet(title, detail_headers, widths)
        write_title(sheet, f"CHI TIẾT SỬ DỤNG VOUCHER - {display_titles.get(title, title.upper())}", len(detail_headers))
        for index, item in enumerate(rows, start=1):
            voucher = item["voucher"]
            usage = item["usage"]
            order = item["order"]
            shared = item["shared"]
            order_id = str(usage.get("donHangId") or "").strip()
            customer_name = usage.get("tenKhach") or order.get("tenKhach") or shared.get("hoTen") or ""
            customer_phone = shared.get("soDienThoai") or order.get("soDienThoai") or ""
            append_row(
                sheet,
                index + 2,
                [
                    index,
                    voucher.get("maVoucher") or voucher.get("id") or usage.get("maUuDai") or "",
                    voucher.get("tenVoucher") or usage.get("tenUuDai") or "",
                    item["usageDate"].strftime("%d/%m/%Y") if item["usageDate"] else "",
                    order_id,
                    customer_name,
                    customer_phone,
                    order.get("loaiHopDong") or "",
                    order.get("tuyen") or shared.get("tuyen") or "",
                    shared.get("diemDon") or order.get("diemDon") or "",
                    shared.get("diemTra") or order.get("diemTra") or "",
                    display_datetime(order.get("ngayGioDi") or shared.get("ngayGioDi")),
                    order.get("bienKiemSoat") or shared.get("bienKiemSoat") or "",
                    order.get("hoTenLaiXe") or "",
                    money_value(shared.get("soTien") if shared else order.get("giaTien")),
                    money_value(shared.get("giamGia") if shared else order.get("giamGia")),
                    money_value(shared.get("tongUuDai") if shared else order.get("tongUuDai")),
                    money_value(usage.get("soTienGiam")),
                    money_value(shared.get("thucThu") if shared else order.get("thucThu")),
                    order.get("trangThai") or "",
                ],
                money_columns={15, 16, 17, 18, 19},
            )
        total_row = len(rows) + 3
        sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=14)
        sheet.cell(total_row, 1, f"TỔNG CỘNG ({len(rows)} lượt)").alignment = Alignment(horizontal="right")
        for column in range(15, 20):
            sheet.cell(total_row, column, f"=SUM({get_column_letter(column)}3:{get_column_letter(column)}{total_row - 1})" if rows else 0)
            sheet.cell(total_row, column).number_format = money_format
        for column in range(1, len(detail_headers) + 1):
            cell = sheet.cell(total_row, column)
            cell.border = border
            cell.fill = total_fill
            cell.font = Font(bold=True)

    write_detail_sheet("Su dung trong ky", period_usages)
    write_detail_sheet("Chi tiet don hang", detail_rows)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"bao-cao-voucher-{report_date.strftime('%Y%m')}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/reports/work-performance.xlsx")
def export_work_performance_report(tuNgay: str = "", denNgay: str = "") -> Response:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover
        raise HTTPException(status_code=500, detail="Thiếu thư viện xuất Excel.") from exc

    from_date = parse_existing_date(tuNgay) or datetime.now()
    to_date = parse_existing_date(denNgay) or from_date
    start_date = from_date.replace(hour=0, minute=0, second=0, microsecond=0)
    end_date = to_date.replace(hour=23, minute=59, second=59, microsecond=999999)
    if end_date < start_date:
        raise HTTPException(status_code=422, detail="Khoảng ngày báo cáo không hợp lệ.")

    customers = customer_records()
    customer_source_by_id = {
        str(row.get("id") or "").strip(): str(row.get("nguonKhach") or "").strip()
        for row in customers if str(row.get("id") or "").strip()
    }
    customer_source_by_phone = {
        normalize_phone(row.get("soDienThoai")): str(row.get("nguonKhach") or "").strip()
        for row in customers if normalize_phone(row.get("soDienThoai"))
    }

    def customer_source(customer: dict[str, Any]) -> str:
        source = str(customer.get("nguonKhach") or "").strip()
        source = source or customer_source_by_id.get(str(customer.get("khachHangId") or "").strip(), "")
        return source or customer_source_by_phone.get(normalize_phone(customer.get("soDienThoai")), "")

    orders = []
    for order in all_order_records():
        created_at = parse_existing_datetime(order.get("createdAt"))
        if created_at and start_date <= created_at <= end_date:
            orders.append((created_at, order))
    orders.sort(key=lambda item: item[0])

    shared_by_order: dict[str, list[dict[str, Any]]] = {}
    for passenger in all_shared_ride_records():
        order_id = str(passenger.get("donHangId") or "").strip()
        if order_id:
            shared_by_order.setdefault(order_id, []).append(passenger)

    report_rows: list[tuple[datetime, dict[str, Any], dict[str, Any] | None]] = []
    for created_at, order in orders:
        is_shared = "ghep" in normalize_text(order.get("loaiHopDong"))
        passengers = shared_by_order.get(str(order.get("id") or "").strip(), []) if is_shared else []
        if passengers:
            report_rows.extend(
                (created_at, order, passenger)
                for passenger in passengers
                if str(passenger.get("loaiKhach") or order.get("loaiKhach") or "").strip().upper() == "B2C"
            )
        elif str(order.get("loaiKhach") or "").strip().upper() == "B2C":
            report_rows.append((created_at, order, None))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Hieu suat lam viec"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    headers = ["STT", "Ngày tạo đơn", "Nguồn khách", "Tên khách hàng", "Loại đơn", "Ngày khách đi", "Loại khách", "Trạng thái", "Ghi chú"]
    widths = [7, 21, 22, 28, 21, 21, 14, 20, 42]
    header_fill = PatternFill("solid", fgColor="FFF200")
    thin = Side(style="thin", color="1F2937")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for column, (header, width) in enumerate(zip(headers, widths), start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
        cell = sheet.cell(1, column, header)
        cell.font = Font(bold=True, color="000000")
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 28

    type_labels = {"xe_nguyen_chuyen": "Xe nguyên chuyến", "xe_ghep": "Xe ghép"}
    for index, (created_at, order, passenger) in enumerate(report_rows, start=1):
        customer = passenger or order
        departure_value = customer.get("ngayGioDi") or order.get("ngayGioDi")
        departure = parse_existing_datetime(departure_value)
        order_type = type_labels.get(str(order.get("loaiHopDong") or ""), str(order.get("loaiHopDong") or ""))
        customer_type = str(customer.get("loaiKhach") or order.get("loaiKhach") or "").strip().upper()
        if customer_type not in {"B2B", "B2C"}:
            customer_type = ""
        values = [
            index,
            created_at,
            customer_source(customer),
            customer.get("hoTen") or customer.get("tenKhach") or order.get("tenKhach") or "",
            order_type,
            departure or str(departure_value or ""),
            customer_type,
            order.get("trangThai") or "Chưa hoàn thành",
            order.get("ghiChu") or "",
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(index + 1, column, value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=column in {4, 9})
        sheet.cell(index + 1, 2).number_format = "dd/mm/yyyy hh:mm"
        if departure:
            sheet.cell(index + 1, 6).number_format = "dd/mm/yyyy hh:mm"
    total_row = len(report_rows) + 2
    sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=8)
    total_cell = sheet.cell(total_row, 1, "TỔNG CỘNG")
    total_cell.font = Font(bold=True)
    total_cell.fill = header_fill
    total_cell.alignment = Alignment(horizontal="right")
    sheet.cell(total_row, 9, f"{len(report_rows)} bản ghi")
    for column in range(1, len(headers) + 1):
        cell = sheet.cell(total_row, column)
        cell.border = border
        cell.fill = header_fill
        cell.font = Font(bold=True)
    sheet.auto_filter.ref = f"A1:I{max(1, len(report_rows) + 1)}"
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"bao-cao-hieu-suat-{start_date.strftime('%Y%m%d')}-{end_date.strftime('%Y%m%d')}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/reports/orders.xlsx")
def export_orders_detail_report(tuNgay: str = "", denNgay: str = "") -> Response:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover
        raise HTTPException(status_code=500, detail="Thiếu thư viện xuất Excel.") from exc

    from_date = parse_existing_date(tuNgay) or datetime.now()
    to_date = parse_existing_date(denNgay) or from_date
    start_date = from_date.replace(hour=0, minute=0, second=0, microsecond=0)
    end_date = to_date.replace(hour=23, minute=59, second=59, microsecond=999999)
    if end_date < start_date:
        raise HTTPException(status_code=422, detail="Khoảng ngày báo cáo không hợp lệ.")

    all_orders = all_order_records()
    shared_rows = all_shared_ride_records()
    benefits = order_benefit_records()
    customers = customer_records()

    customer_source_by_id = {
        str(customer.get("id") or "").strip(): str(customer.get("nguonKhach") or "").strip()
        for customer in customers
        if str(customer.get("id") or "").strip()
    }
    customer_source_by_phone = {
        normalize_phone(customer.get("soDienThoai")): str(customer.get("nguonKhach") or "").strip()
        for customer in customers
        if normalize_phone(customer.get("soDienThoai"))
    }

    def order_customer_source(row: dict[str, Any]) -> str:
        customer_id = str(row.get("khachHangId") or "").strip()
        source = customer_source_by_id.get(customer_id, "")
        if source:
            return source
        return customer_source_by_phone.get(normalize_phone(row.get("soDienThoai")), "")

    def order_report_datetime(row: dict[str, Any]) -> datetime | None:
        return parse_existing_datetime(row.get("ngayGioDi")) or parse_existing_datetime(row.get("createdAt"))

    orders = [
        row
        for row in all_orders
        if (order_report_datetime(row) and start_date <= order_report_datetime(row) <= end_date)
    ]
    order_ids = {str(row.get("id") or "").strip() for row in orders}
    shared_for_orders = [row for row in shared_rows if str(row.get("donHangId") or "").strip() in order_ids]
    benefits_for_orders = [row for row in benefits if str(row.get("donHangId") or "").strip() in order_ids]

    def display_datetime(value: Any) -> str:
        parsed = parse_existing_datetime(value)
        return parsed.strftime("%d/%m/%Y %H:%M") if parsed else str(value or "")

    def display_date(value: Any) -> str:
        parsed = parse_existing_date(value)
        return parsed.strftime("%d/%m/%Y") if parsed else str(value or "")

    def format_benefit_list(order_id: str, kind: str) -> str:
        rows = [
            row for row in benefits_for_orders
            if str(row.get("donHangId") or "").strip() == order_id and normalize_text(row.get("loaiUuDai")) == kind
        ]
        parts = []
        for row in rows:
            name = str(row.get("maUuDai") or row.get("tenUuDai") or "").strip()
            amount = money_value(row.get("soTienGiam"))
            if name:
                parts.append(f"{name} ({amount:,.0f})")
        return "; ".join(parts)

    def invoice_label(row: dict[str, Any]) -> str:
        if normalize_text(row.get("yeuCauHoaDon")) in {"co", "yes", "true", "1"}:
            details = [row.get("tenCongTy"), row.get("maSoThue"), row.get("diaChiHoaDon"), row.get("emailHoaDon")]
            return "Có - " + " | ".join(str(item) for item in details if item)
        return "Không"

    workbook = Workbook()
    workbook.remove(workbook.active)
    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    title_fill = PatternFill("solid", fgColor="0F172A")
    section_fill = PatternFill("solid", fgColor="F8FAFC")
    money_format = '#,##0'

    def setup_sheet(title: str, headers: list[str], widths: list[int] | None = None):
        sheet = workbook.create_sheet(title[:31])
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A3"
        for column_index, header in enumerate(headers, start=1):
            cell = sheet.cell(2, column_index, header)
            cell.fill = header_fill
            cell.font = Font(bold=True, color="0F172A")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
            sheet.column_dimensions[get_column_letter(column_index)].width = (widths or [16] * len(headers))[column_index - 1]
        return sheet

    def write_title(sheet, title: str, end_column: int) -> None:
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
        cell = sheet.cell(1, 1, title)
        cell.fill = title_fill
        cell.font = Font(bold=True, size=14, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[1].height = 24

    def append_row(sheet, row_index: int, values: list[Any], money_columns: set[int] | None = None) -> None:
        for column_index, value in enumerate(values, start=1):
            cell = sheet.cell(row_index, column_index, value)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if row_index % 2 == 1:
                cell.fill = section_fill
            if money_columns and column_index in money_columns:
                cell.number_format = money_format
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)

    period_label = f"{start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
    order_headers = [
        "STT",
        "Mã đơn",
        "Trạng thái",
        "Ngày giờ đi",
        "Dự kiến kết thúc",
        "Ngày giờ hoàn thành",
        "Khách hàng",
        "Nguồn khách",
        "SĐT",
        "Tuyến",
        "Khu vực đặt xe",
        "Điểm đón",
        "Điểm trả",
        "Loại đơn",
        "Số vé",
        "Loại khách",
        "Xe",
        "Số hiệu xe",
        "Loại xe điều động",
        "Lái xe",
        "Mã NV lái xe",
        "Giá tiền",
        "Phụ thu",
        "Lý do phụ thu",
        "Giảm giá",
        "Voucher",
        "Khuyến mãi",
        "Tổng ưu đãi",
        "Đã cọc",
        "Thực thu",
        "Còn phải thu",
        "Tỷ lệ nộp lại",
        "Số tiền nộp lại",
        "Hóa đơn",
        "Ghi chú",
    ]
    order_widths = [8, 24, 18, 18, 18, 18, 24, 18, 16, 26, 16, 22, 22, 18, 10, 12, 16, 14, 18, 24, 14, 16, 16, 28, 16, 16, 28, 28, 16, 14, 16, 16, 14, 16, 34]
    order_sheet = setup_sheet("Don hang", order_headers, order_widths)
    write_title(order_sheet, f"BÁO CÁO CHI TIẾT ĐƠN HÀNG ({period_label})", len(order_headers))

    for index, row in enumerate(sorted(orders, key=lambda item: order_report_datetime(item) or datetime.min), start=1):
        order_id = str(row.get("id") or "").strip()
        gross = money_value(row.get("giaTien"))
        manual_discount = money_value(row.get("giamGia"))
        benefit_total = money_value(row.get("tongUuDai"))
        deposit = money_value(row.get("daCoc"))
        net = money_value(row.get("thucThu"))
        due = max(net - deposit, 0)
        append_row(
            order_sheet,
            index + 2,
            [
                index,
                order_id,
                row.get("trangThai") or "",
                display_datetime(row.get("ngayGioDi")),
                display_datetime(row.get("ngayGioDuKienKetThuc")),
                display_datetime(row.get("ngayGioHoanThanh")),
                row.get("tenKhach") or "",
                order_customer_source(row),
                row.get("soDienThoai") or "",
                row.get("tuyen") or "",
                row.get("khuVucDatXe") or "",
                row.get("diemDon") or "",
                row.get("diemTra") or "",
                row.get("loaiHopDong") or "",
                row.get("soVe") or "",
                row.get("loaiKhach") or "",
                row.get("bienKiemSoat") or "",
                row.get("soHieuXe") or "",
                row.get("loaiXeDieuDong") or "",
                row.get("hoTenLaiXe") or "",
                row.get("maNVLaiXe") or "",
                gross,
                money_value(row.get("phuThu")),
                row.get("lyDoPhuThu") or "",
                manual_discount,
                format_benefit_list(order_id, "voucher"),
                format_benefit_list(order_id, "promotion"),
                benefit_total,
                deposit,
                net,
                due,
                row.get("tyLeNopLai") or "",
                money_value(row.get("soTienNopLai")),
                invoice_label(row),
                row.get("ghiChu") or "",
            ],
            money_columns={22, 23, 25, 28, 29, 30, 31, 33},
        )

    shared_headers = [
        "STT",
        "Mã đơn",
        "Họ tên",
        "SĐT",
        "CCCD",
        "Giới tính",
        "Năm sinh",
        "Nguồn khách",
        "Nhân viên nhập",
        "Tuyến",
        "Điểm đón",
        "Điểm trả",
        "Số tiền",
        "Phụ thu",
        "Lý do phụ thu",
        "Giảm giá",
        "Voucher",
        "Khuyến mãi",
        "Tổng ưu đãi",
        "Đã cọc",
        "Thực thu",
        "Hóa đơn VAT",
    ]
    shared_sheet = setup_sheet("Khach xe ghep", shared_headers, [8, 24, 24, 16, 16, 12, 12, 18, 18, 26, 22, 22, 16, 16, 28, 16, 28, 28, 16, 14, 16, 34])
    write_title(shared_sheet, f"KHÁCH XE GHÉP ({period_label})", len(shared_headers))
    for index, row in enumerate(shared_for_orders, start=1):
        order_id = str(row.get("donHangId") or "").strip()
        passenger_key = normalize_text(row.get("hoTen"))
        row_benefits = [
            item for item in benefits_for_orders
            if str(item.get("donHangId") or "").strip() == order_id
            and (not passenger_key or normalize_text(item.get("tenKhach")) == passenger_key)
        ]
        voucher_text = "; ".join(str(item.get("maUuDai") or item.get("tenUuDai") or "") for item in row_benefits if normalize_text(item.get("loaiUuDai")) == "voucher")
        promo_text = "; ".join(str(item.get("tenUuDai") or item.get("maUuDai") or "") for item in row_benefits if normalize_text(item.get("loaiUuDai")) == "promotion")
        append_row(
            shared_sheet,
            index + 2,
            [
                index,
                order_id,
                row.get("hoTen") or "",
                row.get("soDienThoai") or "",
                row.get("soCCCD") or "",
                row.get("gioiTinh") or "",
                row.get("namSinh") or "",
                row.get("nguonKhach") or "",
                row.get("nhanVienNhap") or "",
                row.get("tuyen") or "",
                row.get("diemDon") or "",
                row.get("diemTra") or "",
                money_value(row.get("soTien")),
                money_value(row.get("phuThu")),
                row.get("lyDoPhuThu") or "",
                money_value(row.get("giamGia")),
                voucher_text,
                promo_text,
                money_value(row.get("tongUuDai")),
                money_value(row.get("daCoc")),
                money_value(row.get("thucThu")),
                invoice_label(row),
            ],
            money_columns={13, 14, 16, 19, 20, 21},
        )

    benefit_headers = [
        "STT",
        "Mã đơn",
        "Loại ưu đãi",
        "Mã/Tên ưu đãi",
        "Khách áp dụng",
        "Loại giá trị",
        "Giá trị",
        "Số tiền giảm",
        "Ngày ghi nhận",
    ]
    benefit_sheet = setup_sheet("Uu dai da ap", benefit_headers, [8, 24, 14, 30, 24, 14, 14, 16, 18])
    write_title(benefit_sheet, f"VOUCHER VÀ KHUYẾN MÃI ĐÃ ÁP ({period_label})", len(benefit_headers))
    for index, row in enumerate(benefits_for_orders, start=1):
        append_row(
            benefit_sheet,
            index + 2,
            [
                index,
                row.get("donHangId") or "",
                row.get("loaiUuDai") or "",
                row.get("maUuDai") or row.get("tenUuDai") or "",
                row.get("tenKhach") or "",
                row.get("loaiGiaTri") or "",
                row.get("giaTri") or "",
                money_value(row.get("soTienGiam")),
                display_date(row.get("createdAt")),
            ],
            money_columns={8},
        )

    def financial_values(row: dict[str, Any], shared: bool = False) -> dict[str, float]:
        gross = money_value(row.get("soTien") if shared else row.get("giaTien"))
        discount = money_value(row.get("giamGia"))
        benefit = money_value(row.get("tongUuDai"))
        surcharge = money_value(row.get("phuThu"))
        revenue = max(gross - discount - benefit, 0) + surcharge
        vat = money_value(row.get("thueVAT"))
        total_payment = money_value(row.get("tongThanhToan")) or revenue + vat
        deposit = money_value(row.get("daCoc"))
        due = money_value(row.get("thucThu"))
        return {
            "gross": gross,
            "discount": discount,
            "benefit": benefit,
            "surcharge": surcharge,
            "revenue": revenue,
            "vat": vat,
            "total_payment": total_payment,
            "deposit": deposit,
            "due": due,
        }

    def add_financials(target: dict[str, Any], values: dict[str, float]) -> None:
        for key, value in values.items():
            target[key] = money_value(target.get(key)) + value

    customer_summary: dict[str, dict[str, Any]] = {}
    shared_by_order: dict[str, list[dict[str, Any]]] = {}
    for shared in shared_for_orders:
        shared_by_order.setdefault(str(shared.get("donHangId") or "").strip(), []).append(shared)

    def customer_key(name: Any, phone: Any) -> str:
        normalized_phone = normalize_phone(phone)
        return f"phone:{normalized_phone}" if normalized_phone else f"name:{normalize_text(name)}"

    for order in orders:
        order_id = str(order.get("id") or "").strip()
        passengers = shared_by_order.get(order_id, [])
        customer_rows = passengers or [order]
        for customer_row in customer_rows:
            is_shared = bool(passengers)
            name = customer_row.get("hoTen") if is_shared else customer_row.get("tenKhach")
            phone = customer_row.get("soDienThoai")
            base_key = customer_key(name, phone)
            if not base_key.removeprefix("name:"):
                continue
            customer_type = normalize_customer_segment(
                customer_row.get("loaiKhach") if is_shared else order.get("loaiKhach")
            ) or "Chưa xác định"
            # Một khách có cả đơn B2C và B2B phải xuất thành hai dòng độc lập.
            key = f"{base_key}|type:{customer_type}"
            current = customer_summary.setdefault(
                key,
                {
                    "name": str(name or "Chưa xác định"),
                    "phone": str(phone or ""),
                    "customer_type": customer_type,
                    "trip_ids": set(),
                },
            )
            current["trip_ids"].add(order_id)
            add_financials(current, financial_values(customer_row, shared=is_shared))

    driver_summary: dict[str, dict[str, Any]] = {}
    for order in orders:
        order_id = str(order.get("id") or "").strip()
        driver_name = str(order.get("hoTenLaiXe") or "Chưa phân công").strip()
        driver_code = str(order.get("maNVLaiXe") or "").strip()
        plate = str(order.get("bienKiemSoat") or "").strip()
        key = f"{normalize_text(driver_code)}|{normalize_text(driver_name)}|{normalize_text(plate)}"
        current = driver_summary.setdefault(
            key,
            {
                "name": driver_name,
                "code": driver_code,
                "plate": plate,
                "vehicle_type": str(order.get("loaiXeDieuDong") or ""),
                "trip_ids": set(),
                "remittance": 0.0,
            },
        )
        current["trip_ids"].add(order_id)
        current["remittance"] += money_value(order.get("soTienNopLai"))
        add_financials(current, financial_values(order))

    customer_headers = [
        "STT",
        "Khách hàng",
        "SĐT",
        "Loại khách",
        "Tổng cuốc",
        "Tổng giá tiền",
        "Tổng giảm giá",
        "Tổng ưu đãi",
        "Doanh thu sau ưu đãi",
        "Tổng VAT",
        "Tổng thanh toán",
        "Đã cọc",
        "Còn phải thu",
        "Doanh thu bình quân/cuốc",
    ]
    customer_sheet = setup_sheet(
        "Tong hop khach hang",
        customer_headers,
        [8, 26, 16, 16, 12, 17, 17, 17, 21, 15, 18, 15, 18, 22],
    )
    write_title(customer_sheet, f"TỔNG HỢP THEO KHÁCH HÀNG ({period_label})", len(customer_headers))
    sorted_customers = sorted(customer_summary.values(), key=lambda item: (-item.get("revenue", 0), item.get("name", "")))
    for index, item in enumerate(sorted_customers, start=1):
        trip_count = len(item["trip_ids"])
        append_row(
            customer_sheet,
            index + 2,
            [
                index,
                item["name"],
                item["phone"],
                item["customer_type"],
                trip_count,
                item.get("gross", 0),
                item.get("discount", 0),
                item.get("benefit", 0),
                item.get("revenue", 0),
                item.get("vat", 0),
                item.get("total_payment", 0),
                item.get("deposit", 0),
                item.get("due", 0),
                item.get("revenue", 0) / trip_count if trip_count else 0,
            ],
            money_columns=set(range(6, 15)),
        )

    driver_headers = [
        "STT",
        "Lái xe",
        "Mã NV",
        "BKS",
        "Loại xe điều động",
        "Tổng cuốc",
        "Tổng giá tiền",
        "Tổng giảm giá",
        "Tổng ưu đãi",
        "Doanh thu sau ưu đãi",
        "Tổng VAT",
        "Tổng thanh toán",
        "Đã cọc",
        "Còn phải thu",
        "Tổng tiền nộp lại",
        "Doanh thu bình quân/cuốc",
    ]
    driver_sheet = setup_sheet(
        "Tong hop lai xe",
        driver_headers,
        [8, 26, 14, 16, 20, 12, 17, 17, 17, 21, 15, 18, 15, 18, 19, 22],
    )
    write_title(driver_sheet, f"TỔNG HỢP THEO LÁI XE ({period_label})", len(driver_headers))
    sorted_drivers = sorted(driver_summary.values(), key=lambda item: (-item.get("revenue", 0), item.get("name", "")))
    for index, item in enumerate(sorted_drivers, start=1):
        trip_count = len(item["trip_ids"])
        append_row(
            driver_sheet,
            index + 2,
            [
                index,
                item["name"],
                item["code"],
                item["plate"],
                item["vehicle_type"],
                trip_count,
                item.get("gross", 0),
                item.get("discount", 0),
                item.get("benefit", 0),
                item.get("revenue", 0),
                item.get("vat", 0),
                item.get("total_payment", 0),
                item.get("deposit", 0),
                item.get("due", 0),
                item.get("remittance", 0),
                item.get("revenue", 0) / trip_count if trip_count else 0,
            ],
            money_columns=set(range(7, 17)),
        )

    def append_total_row(sheet, row_index: int, label_column: int, count_column: int, money_columns: list[int]) -> None:
        sheet.cell(row_index, label_column, "TỔNG CỘNG")
        sheet.cell(row_index, count_column, sum(sheet.cell(row, count_column).value or 0 for row in range(3, row_index)))
        for column in money_columns:
            sheet.cell(row_index, column, sum(money_value(sheet.cell(row, column).value) for row in range(3, row_index)))
        for cell in sheet[row_index]:
            cell.border = border
            cell.fill = title_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="right" if cell.column >= count_column else "left", vertical="center")
            if cell.column in money_columns:
                cell.number_format = money_format

    def append_financial_total_row(sheet, row_index: int, label_end_column: int, money_columns: list[int]) -> None:
        sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=label_end_column)
        sheet.cell(row_index, 1, "TỔNG CỘNG").alignment = Alignment(horizontal="right")
        for column in money_columns:
            sheet.cell(row_index, column, sum(money_value(sheet.cell(row, column).value) for row in range(3, row_index)))
        for cell in sheet[row_index]:
            cell.border = border
            cell.fill = title_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="right" if cell.column >= label_end_column else "left", vertical="center")
            if cell.column in money_columns:
                cell.number_format = money_format

    append_financial_total_row(order_sheet, len(orders) + 3, 21, [22, 23, 25, 28, 29, 30, 31, 33])
    append_financial_total_row(shared_sheet, len(shared_for_orders) + 3, 12, [13, 14, 16, 19, 20, 21])
    append_financial_total_row(benefit_sheet, len(benefits_for_orders) + 3, 7, [8])
    append_total_row(customer_sheet, len(sorted_customers) + 3, 2, 5, list(range(6, 14)))
    append_total_row(driver_sheet, len(sorted_drivers) + 3, 2, 6, list(range(7, 16)))

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"bao-cao-don-hang-{start_date.strftime('%Y%m%d')}-{end_date.strftime('%Y%m%d')}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/api/orders")
def create_order(request: Request, payload: OrderInput) -> dict[str, Any]:
    customers = customer_records()
    tours = tour_records()
    vouchers = voucher_records()
    promotions = promotion_records()
    benefit_usage = order_benefit_records()
    tour = row_by_id(tours, payload.hopDongTourId) if payload.hopDongTourId else None
    if not str(payload.ngayGioDi or "").strip():
        raise HTTPException(status_code=422, detail="Vui lòng nhập ngày giờ đi.")
    if not str(payload.soCho or "").strip():
        raise HTTPException(status_code=422, detail="Vui lòng chọn số chỗ.")
    try:
        parse_datetime(payload.ngayGioDi)
    except (TypeError, ValueError) as exc:
        raise HTTPException(status_code=422, detail="Ngày giờ đi không hợp lệ.") from exc
    if tour is None:
        raise HTTPException(status_code=404, detail="Không tìm thấy hợp đồng/tuyến.")

    customer: dict[str, Any]
    if payload.loaiHopDong == "xe_nguyen_chuyen":
        payload.soDienThoai = validate_customer_phone(payload.soDienThoai, "Số điện thoại khách hàng")
        if not payload.diemDon.strip() or not payload.diemTra.strip():
            raise HTTPException(status_code=422, detail="Vui lÃ²ng nháº­p Ä‘iá»ƒm Ä‘Ã³n vÃ  Ä‘iá»ƒm tráº£.")
        if not payload.khuVucDatXe.strip():
            raise HTTPException(status_code=422, detail="Vui lòng chọn khu vực đặt xe.")
        if payload.giaTien <= 0:
            raise HTTPException(status_code=422, detail="Vui lÃ²ng nháº­p giÃ¡ tiá»n.")
        if payload.loaiKhach not in ("B2B", "B2C"):
            raise HTTPException(status_code=422, detail="Vui lÃ²ng chá»n loáº¡i khÃ¡ch B2B/B2C.")
        customer = row_by_id(customers, payload.khachHangId) if payload.khachHangId else find_customer_by_phone(customers, payload.soDienThoai)
        if customer is None:
            required_customer_fields = {
                "số điện thoại": payload.soDienThoai,
                "họ tên": payload.tenKhach,
                "giới tính": payload.gioiTinh,
                "nguồn khách": payload.nguonKhach,
            }
            missing_fields = [label for label, value in required_customer_fields.items() if not str(value or "").strip()]
            if missing_fields:
                raise HTTPException(
                    status_code=422,
                    detail=f"Vui lòng nhập đầy đủ thông tin khách hàng bắt buộc: {', '.join(missing_fields)}.",
                )
            payload.nhanVienNhap = current_user_display_name(request) or payload.nhanVienNhap
            customer = create_customer_from_order(payload)
    else:
        if payload.soVe <= 0:
            raise HTTPException(status_code=422, detail="Xe ghép phải nhập số vé lớn hơn 0.")
        if len(payload.khachXeGhep) != payload.soVe:
            raise HTTPException(status_code=422, detail="Số khách lẻ phải khớp với số vé.")
        customer = {
            "id": "",
            "tenKhach": f"Khách xe ghép ({payload.soVe} vé)",
            "soDienThoai": "",
        }

    is_franchise_vehicle = False
    commission_rate = 0
    vehicle_code = ""
    driver_name = ""
    driver_code = ""
    if payload.loaiHopDong == "xe_nguyen_chuyen" and payload.yeuCauHoaDon and not (payload.tenCongTy and payload.maSoThue and payload.diaChiHoaDon):
        raise HTTPException(status_code=422, detail="Vui lòng nhập đủ tên công ty, mã số thuế và địa chỉ hóa đơn.")
    if payload.giamGia > 0 and not payload.ghiChuGiamGia.strip():
        raise HTTPException(status_code=422, detail="Vui lòng nhập ghi chú giảm giá thủ công.")
    if payload.loaiHopDong == "xe_nguyen_chuyen" and payload.phuThu > 0 and not payload.lyDoPhuThu.strip():
        raise HTTPException(status_code=422, detail="Vui lòng nhập lý do phụ thu.")
    if payload.loaiHopDong == "xe_nguyen_chuyen" and payload.congNo and not payload.congNoChoAi.strip():
        raise HTTPException(status_code=422, detail="Vui lòng nhập đối tượng ghi nhận công nợ.")

    order_id = make_id("DH")
    used_voucher_ids: set[str] = set()
    benefit_rows: list[list[Any]] = []
    selected_vouchers: list[dict[str, Any]] = []
    selected_promotions: list[dict[str, Any]] = []
    shared_rows: list[list[Any]] = []
    new_shared_customer_rows: list[list[Any]] = []
    vat_amount = 0.0

    if payload.loaiHopDong == "xe_nguyen_chuyen":
        base_amount = payload.giaTien
        manual_discount = min(payload.giamGia, max(base_amount, 0))
        rows, total_benefit_discount, selected_vouchers, selected_promotions = build_benefit_rows(
            order_id,
            customer.get("id", ""),
            customer.get("tenKhach", ""),
            payload.voucherIds,
            payload.promotionIds,
            max(base_amount - manual_discount, 0),
            vouchers,
            promotions,
            benefit_usage,
            used_voucher_ids,
            percent_base_amount=base_amount,
        )
        benefit_rows.extend(rows)
        revenue_amount = max(base_amount - manual_discount - total_benefit_discount, 0) + payload.phuThu
        vat_amount = round(revenue_amount * 0.08) if payload.yeuCauHoaDon else 0
        total_payment = revenue_amount + vat_amount
        deposit_amount = min(payload.daCoc, total_payment)
        net_amount = max(total_payment - deposit_amount, 0)
        order_pickup = payload.diemDon
        order_dropoff = payload.diemTra
        invoice_label = "Có" if payload.yeuCauHoaDon else "Không"
        invoice_fields = [payload.tenCongTy, payload.maSoThue, payload.diaChiHoaDon, payload.emailHoaDon]
    else:
        base_amount = 0.0
        manual_discount = 0.0
        total_benefit_discount = 0.0
        deposit_amount = 0.0
        revenue_amount = 0.0
        for index, passenger in enumerate(payload.khachXeGhep, start=1):
            if passenger.loaiKhach not in ("B2C", "B2B"):
                raise HTTPException(status_code=422, detail=f"Khách lẻ {index} phải chọn loại khách B2C/B2B.")
            if passenger.soTien <= 0:
                raise HTTPException(status_code=422, detail=f"Khách lẻ {index} phải nhập số tiền.")
            if passenger.yeuCauHoaDon and not (passenger.tenCongTy and passenger.maSoThue and passenger.diaChiHoaDon):
                raise HTTPException(status_code=422, detail=f"Khách lẻ {index} cần nhập đủ thông tin hóa đơn.")
            if passenger.giamGia > 0 and not passenger.ghiChuGiamGia.strip():
                raise HTTPException(status_code=422, detail=f"Khách lẻ {index} cần nhập ghi chú giảm giá thủ công.")
            if passenger.phuThu > 0 and not passenger.lyDoPhuThu.strip():
                raise HTTPException(status_code=422, detail=f"Khách lẻ {index} cần nhập lý do phụ thu.")
            if passenger.congNo and not passenger.congNoChoAi.strip():
                raise HTTPException(status_code=422, detail=f"Khách lẻ {index} cần nhập đối tượng ghi nhận công nợ.")
            passenger_phone = validate_customer_phone(
                passenger.soDienThoai,
                f"Số điện thoại khách lẻ {index}",
            )
            passenger.soDienThoai = passenger_phone
            passenger_staff = current_user_display_name(request) or passenger.nhanVienNhap
            passenger_customer = find_customer_by_phone(customers, passenger.soDienThoai)
            if passenger_customer is None:
                passenger_customer = {
                    "id": next_customer_id(customers),
                    "tenKhach": passenger.hoTen,
                    "soDienThoai": passenger.soDienThoai,
                    "soCCCD": passenger.soCCCD,
                    "diaChi": passenger.diaChi,
                    "loaiKhachHang": "Khách cá nhân",
                    "namSinh": passenger.namSinh,
                    "gioiTinh": passenger.gioiTinh,
                    "nguonKhach": passenger.nguonKhach,
                    "nhanVienNhap": passenger_staff,
                    "createdAt": now_iso(),
                }
                customers.append(passenger_customer)
                new_shared_customer_rows.append(
                    [passenger_customer.get(header, "") for header in CUSTOMER_HEADERS]
                )
            passenger_manual_discount = min(passenger.giamGia, passenger.soTien)
            rows, passenger_benefit_discount, passenger_vouchers, passenger_promotions = build_benefit_rows(
                order_id,
                passenger_customer.get("id", ""),
                passenger.hoTen,
                passenger.voucherIds,
                passenger.promotionIds,
                max(passenger.soTien - passenger_manual_discount, 0),
                vouchers,
                promotions,
                benefit_usage,
                used_voucher_ids,
                percent_base_amount=passenger.soTien,
            )
            benefit_rows.extend(rows)
            selected_vouchers.extend(passenger_vouchers)
            selected_promotions.extend(passenger_promotions)
            passenger_revenue = max(passenger.soTien - passenger_manual_discount - passenger_benefit_discount, 0) + passenger.phuThu
            passenger_vat = round(passenger_revenue * 0.08) if passenger.yeuCauHoaDon else 0
            passenger_total_payment = passenger_revenue + passenger_vat
            passenger_deposit = min(passenger.daCoc, passenger_total_payment)
            passenger_net = max(passenger_total_payment - passenger_deposit, 0)
            base_amount += passenger.soTien
            manual_discount += passenger_manual_discount
            total_benefit_discount += passenger_benefit_discount
            deposit_amount += passenger_deposit
            revenue_amount += passenger_revenue
            vat_amount += passenger_vat
            shared_rows.append(
                [
                    make_id("GX"),
                    order_id,
                    tour.get("id", "") if tour else "",
                    tour.get("tuyen", "") if tour else "",
                    payload.bienKiemSoat,
                    payload.ngayGioDi,
                    payload.ngayGioDuKienKetThuc,
                    passenger.hoTen,
                    passenger.soDienThoai,
                    passenger.soCCCD,
                    passenger.diaChi,
                    passenger.gioiTinh,
                    passenger.namSinh,
                    passenger.nguonKhach,
                    passenger_staff,
                    passenger.diemDon,
                    passenger.diemTra,
                    passenger.soTien,
                    passenger_manual_discount,
                    passenger_deposit,
                    passenger_net,
                    ", ".join(voucher.get("id", "") for voucher in passenger_vouchers),
                    ", ".join(voucher.get("maVoucher", "") for voucher in passenger_vouchers),
                    ", ".join(promotion.get("id", "") for promotion in passenger_promotions),
                    ", ".join(promotion.get("tenChuongTrinh", "") for promotion in passenger_promotions),
                    passenger_benefit_discount,
                    "Có" if passenger.yeuCauHoaDon else "Không",
                    passenger.tenCongTy,
                    passenger.maSoThue,
                    passenger.diaChiHoaDon,
                    passenger.emailHoaDon,
                    now_iso(),
                    "",
                    "",
                    "",
                    passenger.ghiChuGiamGia,
                    passenger_vat,
                    passenger_total_payment,
                    "Có" if passenger.congNo else "Không",
                    passenger.congNoChoAi if passenger.congNo else "",
                    "Chưa thu hồi" if passenger.congNo else "",
                    "",
                    "",
                    passenger.phuThu,
                    passenger.lyDoPhuThu if passenger.phuThu > 0 else "",
                    passenger.loaiKhach,
                ]
            )
        net_amount = max(revenue_amount + vat_amount - deposit_amount, 0)
        order_pickup = ", ".join(dict.fromkeys(passenger.diemDon for passenger in payload.khachXeGhep if passenger.diemDon))
        order_dropoff = ", ".join(dict.fromkeys(passenger.diemTra for passenger in payload.khachXeGhep if passenger.diemTra))
        invoice_label = "Theo từng khách"
        invoice_fields = ["", "", "", ""]
    total_payment = revenue_amount + vat_amount
    commission_amount = round_up_ten_thousand(revenue_amount * commission_rate / 100)
    row = [
        order_id,
        customer.get("id", ""),
        customer.get("tenKhach", ""),
        customer.get("soDienThoai", ""),
        tour.get("id", "") if tour else "",
        tour.get("tuyen", "") if tour else "",
        order_pickup,
        order_dropoff,
        base_amount,
        payload.bienKiemSoat,
        vehicle_code,
        driver_name,
        driver_code,
        payload.ngayGioDi,
        payload.ngayGioDuKienKetThuc,
        "Xe nguyên chuyến" if payload.loaiHopDong == "xe_nguyen_chuyen" else "Xe ghép",
        payload.soVe if payload.loaiHopDong == "xe_ghep" else "",
        payload.loaiKhach if payload.loaiHopDong == "xe_nguyen_chuyen" else "",
        invoice_label,
        invoice_fields[0],
        invoice_fields[1],
        invoice_fields[2],
        invoice_fields[3],
        payload.ghiChu,
        "Chưa hoàn thành",
        "",
        now_iso(),
        manual_discount,
        deposit_amount,
        net_amount,
        "",
        commission_rate,
        commission_amount,
        ", ".join(voucher.get("maVoucher", "") for voucher in selected_vouchers),
        ", ".join(promotion.get("tenChuongTrinh", "") for promotion in selected_promotions),
        total_benefit_discount,
        payload.khuVucDatXe,
        "",
        "",
        "",
        "",
        "",
        payload.ghiChuGiamGia,
        "Có" if payload.loaiHopDong == "xe_nguyen_chuyen" and payload.congNo else "Không",
        payload.congNoChoAi if payload.loaiHopDong == "xe_nguyen_chuyen" and payload.congNo else "",
        vat_amount,
        total_payment,
        "",
        "",
        "",
        payload.phuThu if payload.loaiHopDong == "xe_nguyen_chuyen" else sum(passenger.phuThu for passenger in payload.khachXeGhep),
        payload.lyDoPhuThu if payload.loaiHopDong == "xe_nguyen_chuyen" and payload.phuThu > 0 else "",
        "",
        "",
        "",
        "",
        payload.soCho,
        "Chưa gửi tài xế",
        "Công nợ" if payload.loaiHopDong == "xe_nguyen_chuyen" and payload.congNo else "",
        "",
        "",
        current_user_display_name(request),
    ]
    if len(row) != len(ORDER_HEADERS):
        raise HTTPException(status_code=500, detail="Cấu trúc dữ liệu đơn hàng không hợp lệ.")
    append_worksheet_row(orders_worksheet(), row)
    if new_shared_customer_rows:
        append_worksheet_rows(customers_worksheet(), new_shared_customer_rows)
    if benefit_rows:
        append_worksheet_rows(order_benefits_worksheet(), benefit_rows)
    if shared_rows:
        append_worksheet_rows(shared_ride_worksheet(), shared_rows)
    return {"ok": True, "id": order_id}


def update_shared_order(
    order_id: str,
    payload: OrderInput,
    request: Request,
    worksheet: Any,
    row_number: int,
    order: dict[str, Any],
    duplicate_row_numbers: list[int] | None = None,
) -> dict[str, Any]:
    tour = row_by_id(tour_records(), payload.hopDongTourId) if payload.hopDongTourId else None
    if tour is None:
        raise HTTPException(status_code=404, detail="Không tìm thấy hợp đồng/tuyến.")
    if not payload.khachXeGhep:
        raise HTTPException(status_code=422, detail="Vui lòng khai báo ít nhất một khách xe ghép.")
    if int(payload.soVe or 0) != len(payload.khachXeGhep):
        raise HTTPException(status_code=422, detail="Số vé không khớp với số khách xe ghép.")
    if not str(payload.ngayGioDi or "").strip() or not str(payload.soCho or "").strip():
        raise HTTPException(status_code=422, detail="Vui lòng nhập ngày giờ đi và số chỗ.")
    try:
        updated_start_at = parse_datetime(payload.ngayGioDi)
    except (TypeError, ValueError) as exc:
        raise HTTPException(status_code=422, detail="Ngày giờ đi không hợp lệ.") from exc
    existing_end_at = parse_existing_datetime(order.get("ngayGioDuKienKetThuc"))
    if existing_end_at and existing_end_at <= updated_start_at:
        raise HTTPException(status_code=422, detail="Ngày giờ đi phải trước ngày giờ đến dự kiến.")

    customers = customer_records()
    new_customer_rows: list[list[Any]] = []
    vouchers = voucher_records()
    promotions = promotion_records()
    benefit_usage = [
        item for item in order_benefit_records()
        if str(item.get("donHangId") or "") != str(order_id)
    ]
    used_voucher_ids: set[str] = set()
    benefit_rows: list[list[Any]] = []
    shared_rows: list[list[Any]] = []
    selected_vouchers: list[dict[str, Any]] = []
    selected_promotions: list[dict[str, Any]] = []
    base_amount = manual_discount = benefit_discount = deposit_amount = revenue_amount = vat_amount = 0.0

    for index, passenger in enumerate(payload.khachXeGhep, start=1):
        if passenger.loaiKhach not in ("B2C", "B2B"):
            raise HTTPException(status_code=422, detail=f"Khách lẻ {index} phải chọn loại khách B2C/B2B.")
        if passenger.soTien <= 0 or not passenger.diemDon.strip() or not passenger.diemTra.strip():
            raise HTTPException(status_code=422, detail=f"Khách lẻ {index} chưa đủ hành trình hoặc số tiền.")
        if passenger.yeuCauHoaDon and not (passenger.tenCongTy and passenger.maSoThue and passenger.diaChiHoaDon):
            raise HTTPException(status_code=422, detail=f"Khách lẻ {index} cần nhập đủ thông tin hóa đơn.")
        if passenger.giamGia > 0 and not passenger.ghiChuGiamGia.strip():
            raise HTTPException(status_code=422, detail=f"Khách lẻ {index} cần nhập ghi chú giảm giá.")
        if passenger.phuThu > 0 and not passenger.lyDoPhuThu.strip():
            raise HTTPException(status_code=422, detail=f"Khách lẻ {index} cần nhập lý do phụ thu.")
        if passenger.congNo and not passenger.congNoChoAi.strip():
            raise HTTPException(status_code=422, detail=f"Khách lẻ {index} cần nhập đối tượng công nợ.")
        passenger.soDienThoai = validate_customer_phone(passenger.soDienThoai, f"Số điện thoại khách lẻ {index}")
        passenger_staff = current_user_display_name(request) or passenger.nhanVienNhap
        customer = find_customer_by_phone(customers, passenger.soDienThoai)
        if customer is None:
            customer = {
                "id": next_customer_id(customers), "tenKhach": passenger.hoTen,
                "soDienThoai": passenger.soDienThoai, "soCCCD": passenger.soCCCD,
                "diaChi": passenger.diaChi, "loaiKhachHang": "Khách cá nhân",
                "namSinh": passenger.namSinh, "gioiTinh": passenger.gioiTinh,
                "nguonKhach": passenger.nguonKhach, "nhanVienNhap": passenger_staff,
                "createdAt": now_iso(),
            }
            customers.append(customer)
            new_customer_rows.append([customer.get(header, "") for header in CUSTOMER_HEADERS])
        passenger_manual = min(passenger.giamGia, passenger.soTien)
        rows, passenger_benefit, passenger_vouchers, passenger_promotions = build_benefit_rows(
            order_id, customer.get("id", ""), passenger.hoTen,
            passenger.voucherIds, passenger.promotionIds,
            max(passenger.soTien - passenger_manual, 0),
            vouchers, promotions, benefit_usage, used_voucher_ids,
            percent_base_amount=passenger.soTien,
        )
        benefit_rows.extend(rows)
        selected_vouchers.extend(passenger_vouchers)
        selected_promotions.extend(passenger_promotions)
        passenger_revenue = max(passenger.soTien - passenger_manual - passenger_benefit, 0) + passenger.phuThu
        passenger_vat = round(passenger_revenue * 0.08) if passenger.yeuCauHoaDon else 0
        passenger_total = passenger_revenue + passenger_vat
        passenger_deposit = min(passenger.daCoc, passenger_total)
        passenger_net = max(passenger_total - passenger_deposit, 0)
        base_amount += passenger.soTien
        manual_discount += passenger_manual
        benefit_discount += passenger_benefit
        deposit_amount += passenger_deposit
        revenue_amount += passenger_revenue
        vat_amount += passenger_vat
        shared_rows.append([
            make_id("GX"), order_id, tour.get("id", ""), tour.get("tuyen", ""),
            order.get("bienKiemSoat", ""), payload.ngayGioDi, order.get("ngayGioDuKienKetThuc", ""),
            passenger.hoTen, passenger.soDienThoai, passenger.soCCCD, passenger.diaChi,
            passenger.gioiTinh, passenger.namSinh, passenger.nguonKhach, passenger_staff,
            passenger.diemDon, passenger.diemTra, passenger.soTien, passenger_manual,
            passenger_deposit, passenger_net,
            ", ".join(item.get("id", "") for item in passenger_vouchers),
            ", ".join(item.get("maVoucher", "") for item in passenger_vouchers),
            ", ".join(item.get("id", "") for item in passenger_promotions),
            ", ".join(item.get("tenChuongTrinh", "") for item in passenger_promotions),
            passenger_benefit, "Có" if passenger.yeuCauHoaDon else "Không",
            passenger.tenCongTy, passenger.maSoThue, passenger.diaChiHoaDon,
            passenger.emailHoaDon, now_iso(), "", "", "", passenger.ghiChuGiamGia,
            passenger_vat, passenger_total, "Có" if passenger.congNo else "Không",
            passenger.congNoChoAi if passenger.congNo else "",
            "Chưa thu hồi" if passenger.congNo else "", "", "",
            passenger.phuThu, passenger.lyDoPhuThu if passenger.phuThu > 0 else "",
            passenger.loaiKhach,
        ])

    before = dict(order)
    order.update({
        "khachHangId": "", "tenKhach": f"Khách xe ghép ({len(payload.khachXeGhep)} vé)",
        "soDienThoai": "", "hopDongTourId": tour.get("id", ""), "tuyen": tour.get("tuyen", ""),
        "diemDon": ", ".join(dict.fromkeys(item.diemDon for item in payload.khachXeGhep if item.diemDon)),
        "diemTra": ", ".join(dict.fromkeys(item.diemTra for item in payload.khachXeGhep if item.diemTra)),
        "giaTien": base_amount, "ngayGioDi": payload.ngayGioDi, "loaiHopDong": "Xe ghép",
        "soVe": len(payload.khachXeGhep), "loaiKhach": "", "yeuCauHoaDon": "Theo từng khách",
        "ghiChu": payload.ghiChu, "giamGia": manual_discount, "daCoc": deposit_amount,
        "thucThu": max(revenue_amount + vat_amount - deposit_amount, 0),
        "voucherCodes": ", ".join(item.get("maVoucher", "") for item in selected_vouchers),
        "khuyenMai": ", ".join(item.get("tenChuongTrinh", "") for item in selected_promotions),
        "tongUuDai": benefit_discount, "khuVucDatXe": payload.khuVucDatXe,
        "ghiChuGiamGia": "", "congNo": "Không", "congNoChoAi": "",
        "thueVAT": vat_amount, "tongThanhToan": revenue_amount + vat_amount,
        "phuThu": sum(item.phuThu for item in payload.khachXeGhep),
        "lyDoPhuThu": "", "soCho": payload.soCho,
    })
    update_row_by_headers(worksheet, row_number, ORDER_HEADERS, order)
    for duplicate_row_number in duplicate_row_numbers or []:
        soft_delete_row(worksheet, duplicate_row_number, ORDER_HEADERS, order, request)
    shared_worksheet = shared_ride_worksheet()
    values = worksheet_values(shared_worksheet)
    order_column = SHARED_RIDE_HEADERS.index("donHangId")
    end_column = re.sub(r"\d+$", "", gspread.utils.rowcol_to_a1(1, len(SHARED_RIDE_HEADERS)))
    ranges = [
        f"A{number}:{end_column}{number}" for number, row in enumerate(values[1:], start=2)
        if len(row) > order_column and str(row[order_column] or "") == str(order_id)
    ]
    for start in range(0, len(ranges), 200):
        shared_worksheet.batch_clear(ranges[start:start + 200])
    if ranges:
        invalidate_worksheet_cache(shared_worksheet)
    append_worksheet_rows(shared_worksheet, shared_rows)
    if new_customer_rows:
        append_worksheet_rows(customers_worksheet(), new_customer_rows)
    replace_order_benefits(order_id, benefit_rows)
    log_action(request, "update_order", "order", order_id, before=before, after=order)
    return {"ok": True, "id": order_id}


@app.put("/api/orders/{order_id}")
def update_order(order_id: str, payload: OrderInput, request: Request) -> dict[str, Any]:
    worksheet = orders_worksheet()
    row_numbers = find_rows_by_id(worksheet, order_id, force_refresh=True)
    row_number = row_numbers[0] if row_numbers else None
    if row_number is None:
        raise HTTPException(status_code=404, detail="Không tìm thấy đơn hàng.")

    orders = worksheet_records(worksheet, ORDER_HEADERS)
    order = row_by_id(orders, order_id)
    if order is None:
        raise HTTPException(status_code=404, detail="Không tìm thấy đơn hàng.")
    if order_is_done(order):
        raise HTTPException(status_code=409, detail="Đơn hàng đã hoàn thành, không thể chỉnh sửa.")
    existing_shared = normalize_text(order.get("loaiHopDong")).find("ghep") >= 0
    requested_shared = payload.loaiHopDong == "xe_ghep"
    if existing_shared != requested_shared:
        raise HTTPException(status_code=422, detail="Không thể đổi loại đơn hàng khi chỉnh sửa.")
    if requested_shared:
        return update_shared_order(
            order_id,
            payload,
            request,
            worksheet,
            row_number,
            order,
            duplicate_row_numbers=row_numbers[1:],
        )

    tours = tour_records()
    tour = row_by_id(tours, payload.hopDongTourId) if payload.hopDongTourId else None
    if tour is None:
        raise HTTPException(status_code=404, detail="Không tìm thấy hợp đồng/tuyến.")
    if not payload.diemDon.strip() or not payload.diemTra.strip():
        raise HTTPException(status_code=422, detail="Vui lòng nhập điểm đón và điểm trả.")
    if payload.giaTien <= 0:
        raise HTTPException(status_code=422, detail="Vui lòng nhập giá tiền.")
    if payload.loaiKhach not in ("B2B", "B2C"):
        raise HTTPException(status_code=422, detail="Vui lòng chọn loại khách B2B/B2C.")
    if payload.yeuCauHoaDon and not (payload.tenCongTy and payload.maSoThue and payload.diaChiHoaDon):
        raise HTTPException(status_code=422, detail="Vui lòng nhập đủ tên công ty, mã số thuế và địa chỉ hóa đơn.")
    if payload.giamGia > 0 and not payload.ghiChuGiamGia.strip():
        raise HTTPException(status_code=422, detail="Vui lòng nhập ghi chú giảm giá thủ công.")
    if payload.phuThu > 0 and not payload.lyDoPhuThu.strip():
        raise HTTPException(status_code=422, detail="Vui lòng nhập lý do phụ thu.")
    if payload.congNo and not payload.congNoChoAi.strip():
        raise HTTPException(status_code=422, detail="Vui lòng nhập đối tượng ghi nhận công nợ.")
    if not str(payload.ngayGioDi or "").strip():
        raise HTTPException(status_code=422, detail="Vui lòng nhập ngày giờ đi.")
    if not str(payload.soCho or "").strip():
        raise HTTPException(status_code=422, detail="Vui lòng chọn số chỗ.")
    try:
        updated_start_at = parse_datetime(payload.ngayGioDi)
    except (TypeError, ValueError) as exc:
        raise HTTPException(status_code=422, detail="Ngày giờ đi không hợp lệ.") from exc
    existing_end_at = parse_existing_datetime(order.get("ngayGioDuKienKetThuc"))
    if existing_end_at and existing_end_at <= updated_start_at:
        raise HTTPException(status_code=422, detail="Ngày giờ đi phải trước ngày giờ đến dự kiến đã điều xe.")

    customers = customer_records()
    payload.soDienThoai = validate_customer_phone(payload.soDienThoai, "Số điện thoại khách hàng")
    customer = row_by_id(customers, payload.khachHangId) if payload.khachHangId else find_customer_by_phone(customers, payload.soDienThoai)
    if customer is None:
        payload.nhanVienNhap = current_user_display_name(request) or payload.nhanVienNhap
        customer = create_customer_from_order(payload)

    manual_discount = min(payload.giamGia, max(payload.giaTien, 0))
    all_benefit_usage = order_benefit_records()
    other_benefit_usage = [
        item for item in all_benefit_usage if str(item.get("donHangId") or "") != str(order_id)
    ]
    benefit_rows, benefit_discount, selected_vouchers, selected_promotions = build_benefit_rows(
        order_id,
        str(customer.get("id") or ""),
        str(customer.get("tenKhach") or payload.tenKhach),
        payload.voucherIds,
        payload.promotionIds,
        max(payload.giaTien - manual_discount, 0),
        voucher_records(),
        promotion_records(),
        other_benefit_usage,
        set(),
        percent_base_amount=payload.giaTien,
    )
    revenue_amount = max(payload.giaTien - manual_discount - benefit_discount, 0) + payload.phuThu
    vat_amount = round(revenue_amount * 0.08) if payload.yeuCauHoaDon else 0
    total_payment = revenue_amount + vat_amount
    deposit_amount = min(payload.daCoc, total_payment)
    net_amount = max(total_payment - deposit_amount, 0)
    commission_rate = money_value(order.get("tyLeNopLai"))
    updated_commission_amount = round_up_ten_thousand(revenue_amount * commission_rate / 100)
    was_debt = normalize_text(order.get("congNo")) in {"co", "yes", "true", "1"}
    debt_status = str(order.get("trangThaiCongNo") or "") if payload.congNo and was_debt else ("Chưa thu hồi" if payload.congNo else "")

    before = dict(order)
    order.update(
        {
            "khachHangId": customer.get("id", ""),
            "tenKhach": customer.get("tenKhach", payload.tenKhach),
            "soDienThoai": customer.get("soDienThoai", payload.soDienThoai),
            "hopDongTourId": tour.get("id", ""),
            "tuyen": tour.get("tuyen", ""),
            "diemDon": payload.diemDon,
            "diemTra": payload.diemTra,
            "ngayGioDi": payload.ngayGioDi,
            "giaTien": payload.giaTien,
            "giamGia": manual_discount,
            "daCoc": deposit_amount,
            "thucThu": net_amount,
            "loaiHopDong": "Xe nguyên chuyến",
            "soVe": "",
            "loaiKhach": payload.loaiKhach,
            "yeuCauHoaDon": "Có" if payload.yeuCauHoaDon else "Không",
            "tenCongTy": payload.tenCongTy if payload.yeuCauHoaDon else "",
            "maSoThue": payload.maSoThue if payload.yeuCauHoaDon else "",
            "diaChiHoaDon": payload.diaChiHoaDon if payload.yeuCauHoaDon else "",
            "emailHoaDon": payload.emailHoaDon if payload.yeuCauHoaDon else "",
            "trangThaiHoaDon": order.get("trangThaiHoaDon", "") if payload.yeuCauHoaDon else "",
            "ngayXuatHoaDon": order.get("ngayXuatHoaDon", "") if payload.yeuCauHoaDon else "",
            "nguoiXuatHoaDon": order.get("nguoiXuatHoaDon", "") if payload.yeuCauHoaDon else "",
            "ghiChu": payload.ghiChu,
            "voucherCodes": ", ".join(voucher.get("maVoucher", "") for voucher in selected_vouchers),
            "khuyenMai": ", ".join(promotion.get("tenChuongTrinh", "") for promotion in selected_promotions),
            "tongUuDai": benefit_discount,
            "ghiChuGiamGia": payload.ghiChuGiamGia if manual_discount > 0 else "",
            "phuThu": payload.phuThu,
            "lyDoPhuThu": payload.lyDoPhuThu if payload.phuThu > 0 else "",
            "congNo": "Có" if payload.congNo else "Không",
            "congNoChoAi": payload.congNoChoAi if payload.congNo else "",
            "trangThaiNopTien": "Công nợ" if payload.congNo else str(order.get("trangThaiNopTien") or ""),
            "ngayXacNhanNopTien": "" if payload.congNo else str(order.get("ngayXacNhanNopTien") or ""),
            "nguoiXacNhanNopTien": "" if payload.congNo else str(order.get("nguoiXacNhanNopTien") or ""),
            "trangThaiCongNo": debt_status,
            "ngayThuHoiCongNo": order.get("ngayThuHoiCongNo", "") if payload.congNo and was_debt else "",
            "nguoiThuHoiCongNo": order.get("nguoiThuHoiCongNo", "") if payload.congNo and was_debt else "",
            "thueVAT": vat_amount,
            "tongThanhToan": total_payment,
            "soTienNopLai": updated_commission_amount,
            "khuVucDatXe": payload.khuVucDatXe,
            "soCho": payload.soCho,
        }
    )
    if updated_commission_amount != money_value(before.get("soTienNopLai")):
        order["trangThaiHoaHong"] = ""
        order["ngayThuHoaHong"] = ""
        order["nguoiThuHoaHong"] = ""
    update_row_by_headers(worksheet, row_number, ORDER_HEADERS, order)
    for duplicate_row_number in row_numbers[1:]:
        soft_delete_row(worksheet, duplicate_row_number, ORDER_HEADERS, order, request)
    replace_order_benefits(order_id, benefit_rows)
    log_action(request, "update_order", "order", order_id, before=before, after=order)
    return {"ok": True, "id": order_id}


@app.delete("/api/orders/{order_id}")
def delete_order(order_id: str, request: Request) -> dict[str, Any]:
    worksheet = orders_worksheet()
    row_numbers = find_rows_by_id(worksheet, order_id, force_refresh=True)
    if not row_numbers:
        raise HTTPException(status_code=404, detail="Không tìm thấy đơn hàng.")
    order = row_by_id(worksheet_records(worksheet, ORDER_HEADERS), order_id)
    if order is None:
        raise HTTPException(status_code=404, detail="Không tìm thấy đơn hàng.")
    if order_is_done(order):
        raise HTTPException(status_code=409, detail="Đơn hàng đã hoàn thành, không thể xóa.")

    deleted = dict(order)
    for row_number in row_numbers:
        deleted = soft_delete_row(worksheet, row_number, ORDER_HEADERS, order, request)
    replace_order_benefits(order_id, [])
    shared_worksheet = shared_ride_worksheet()
    shared_values = worksheet_values(shared_worksheet)
    order_id_column = SHARED_RIDE_HEADERS.index("donHangId")
    end_column = re.sub(r"\d+$", "", gspread.utils.rowcol_to_a1(1, len(SHARED_RIDE_HEADERS)))
    shared_ranges = [
        f"A{shared_row_number}:{end_column}{shared_row_number}"
        for shared_row_number, shared_row in enumerate(shared_values[1:], start=2)
        if len(shared_row) > order_id_column and str(shared_row[order_id_column] or "") == str(order_id)
    ]
    for start in range(0, len(shared_ranges), 200):
        shared_worksheet.batch_clear(shared_ranges[start : start + 200])
    if shared_ranges:
        invalidate_worksheet_cache(shared_worksheet)
    log_action(request, "delete_order", "order", order_id, before=order, after=deleted)
    return {"ok": True, "id": order_id}


@app.post("/api/orders/{order_id}/assign-vehicle")
def assign_order_vehicle(order_id: str, payload: AssignVehicleInput, request: Request) -> dict[str, Any]:
    worksheet = orders_worksheet()
    row_number = find_row_by_id(worksheet, order_id)
    if row_number is None:
        raise HTTPException(status_code=404, detail="Không tìm thấy đơn hàng.")
    orders = worksheet_records(worksheet, ORDER_HEADERS)
    order = row_by_id(orders, order_id)
    if order is None:
        raise HTTPException(status_code=404, detail="Không tìm thấy đơn hàng.")
    if order_is_done(order):
        raise HTTPException(status_code=409, detail="Đơn hàng đã hoàn thành, không thể điều xe lại.")
    before_order = dict(order)

    stored_start = str(order.get("ngayGioDi") or "").strip()
    if not stored_start:
        raise HTTPException(status_code=422, detail="Đơn hàng chưa có ngày giờ đi.")
    start_at = parse_datetime(stored_start)
    end_at = parse_datetime(payload.ngayGioDuKienKetThuc)
    if end_at <= start_at:
        raise HTTPException(status_code=422, detail="Giờ kết thúc dự kiến phải sau giờ đi.")

    if not str(payload.bienKiemSoat or "").strip():
        order["bienKiemSoat"] = ""
        order["soHieuXe"] = ""
        order["hoTenLaiXe"] = ""
        order["maNVLaiXe"] = ""
        order["ngayGioDi"] = stored_start
        order["ngayGioDuKienKetThuc"] = payload.ngayGioDuKienKetThuc
        order["loaiXeDieuDong"] = ""
        order["tyLeNopLai"] = 0
        order["soTienNopLai"] = 0
        order["trangThaiHoaHong"] = ""
        order["ngayThuHoaHong"] = ""
        order["nguoiThuHoaHong"] = ""
        update_row_by_headers(worksheet, row_number, ORDER_HEADERS, order)
        log_action(request, "unassign_order_vehicle", "order", order_id, before=before_order, after=order)

        shared_worksheet = shared_ride_worksheet()
        shared_values = worksheet_values(shared_worksheet)
        for shared_row_number, values in enumerate(shared_values[1:], start=2):
            padded = values + [""] * max(len(SHARED_RIDE_HEADERS) - len(values), 0)
            shared_row = {header: padded[index] for index, header in enumerate(SHARED_RIDE_HEADERS)}
            if str(shared_row.get("donHangId") or "") != order_id:
                continue
            shared_row["bienKiemSoat"] = ""
            shared_row["ngayGioDi"] = stored_start
            shared_row["ngayGioDuKienKetThuc"] = payload.ngayGioDuKienKetThuc
            update_row_by_headers(shared_worksheet, shared_row_number, SHARED_RIDE_HEADERS, shared_row)
        return {"ok": True, "id": order_id}

    roster_vehicle = roster_vehicle_by_plate(payload.bienKiemSoat, date_key(start_at))
    franchise_vehicle = None if roster_vehicle else franchise_vehicle_by_plate(payload.bienKiemSoat)
    if roster_vehicle is None and franchise_vehicle is None:
        raise HTTPException(status_code=404, detail="Xe này không có ca hợp lệ hoặc chưa được khai báo trong xe thương quyền.")

    conflict = conflicting_order(orders, payload.bienKiemSoat, start_at, end_at, exclude_order_id=order_id)
    if conflict is not None:
        raise HTTPException(
            status_code=409,
            detail=f"Xe {payload.bienKiemSoat} đang bận đơn {conflict.get('id')} trong khung giờ này.",
        )

    if franchise_vehicle is not None:
        vehicle_code = str(franchise_vehicle.get("dongXe") or "").strip()
        driver_name = str(franchise_vehicle.get("hoTenLaiXe") or "").strip()
        driver_code = ""
        vehicle_type = "Xe thương quyền hợp tác"
        commission_rate = payload.tyLeNopLai
    else:
        driver_text = roster_driver_text(roster_vehicle)
        if not driver_text.strip():
            raise HTTPException(status_code=422, detail="Xe này có lên ca nhưng chưa có lái xe.")
        vehicle_code = str(roster_vehicle.get("soHieuXe") or "").strip()
        driver_name = roster_driver_name(driver_text)
        driver_code = roster_driver_code(driver_text)
        vehicle_type = "Xe Công ty"
        commission_rate = 0

    order["bienKiemSoat"] = payload.bienKiemSoat
    order["soHieuXe"] = vehicle_code
    order["hoTenLaiXe"] = driver_name
    order["maNVLaiXe"] = driver_code
    order["ngayGioDi"] = stored_start
    order["ngayGioDuKienKetThuc"] = payload.ngayGioDuKienKetThuc
    order["loaiXeDieuDong"] = vehicle_type
    order["tyLeNopLai"] = commission_rate
    driver_revenue = max(
        money_value(order.get("giaTien"))
        + money_value(order.get("phuThu"))
        - money_value(order.get("giamGia"))
        - money_value(order.get("tongUuDai")),
        0,
    )
    new_commission_amount = round_up_ten_thousand(driver_revenue * commission_rate / 100)
    old_commission_amount = money_value(before_order.get("soTienNopLai"))
    order["soTienNopLai"] = new_commission_amount
    if new_commission_amount <= 0 or new_commission_amount != old_commission_amount:
        order["trangThaiHoaHong"] = ""
        order["ngayThuHoaHong"] = ""
        order["nguoiThuHoaHong"] = ""
    update_row_by_headers(worksheet, row_number, ORDER_HEADERS, order)
    log_action(request, "assign_order_vehicle", "order", order_id, before=before_order, after=order)

    shared_worksheet = shared_ride_worksheet()
    shared_values = worksheet_values(shared_worksheet)
    for shared_row_number, values in enumerate(shared_values[1:], start=2):
        padded = values + [""] * max(len(SHARED_RIDE_HEADERS) - len(values), 0)
        shared_row = {header: padded[index] for index, header in enumerate(SHARED_RIDE_HEADERS)}
        if str(shared_row.get("donHangId") or "") != order_id:
            continue
        shared_row["bienKiemSoat"] = payload.bienKiemSoat
        shared_row["ngayGioDi"] = stored_start
        shared_row["ngayGioDuKienKetThuc"] = payload.ngayGioDuKienKetThuc
        update_row_by_headers(shared_worksheet, shared_row_number, SHARED_RIDE_HEADERS, shared_row)

    return {"ok": True, "id": order_id}


@app.post("/api/orders/{order_id}/complete")
def complete_order(order_id: str, payload: CompleteOrderInput, request: Request) -> dict[str, Any]:
    worksheet = orders_worksheet()
    row_number = find_row_by_id(worksheet, order_id)
    if row_number is None:
        raise HTTPException(status_code=404, detail="Không tìm thấy đơn hàng.")
    order = row_by_id(worksheet_records(worksheet, ORDER_HEADERS), order_id)
    if order and order_is_done(order):
        raise HTTPException(status_code=409, detail="ÄÆ¡n hÃ ng Ä‘Ã£ hoÃ n thÃ nh, khÃ´ng thá»ƒ thao tÃ¡c láº¡i.")
    if order and not (str(order.get("bienKiemSoat") or "").strip() and str(order.get("ngayGioDi") or "").strip()):
        raise HTTPException(status_code=422, detail="Đơn hàng chưa điều xe, chưa thể hoàn thành.")
    if order and normalize_text(order.get("trangThaiGuiTaiXe")) != "da gui tai xe":
        raise HTTPException(
            status_code=422,
            detail="Đơn hàng chưa được đánh dấu Đã gửi tài xế, chưa thể hoàn thành.",
        )
    try:
        completed_at = parse_datetime(payload.ngayGioHoanThanh)
        started_at = parse_datetime(str(order.get("ngayGioDi") or ""))
    except (TypeError, ValueError) as exc:
        raise HTTPException(status_code=422, detail="Ngày giờ hoàn thành không hợp lệ.") from exc
    if completed_at < started_at:
        raise HTTPException(status_code=422, detail="Giờ hoàn thành không được trước giờ đi.")
    status_col = ORDER_HEADERS.index("trangThai") + 1
    completed_col = ORDER_HEADERS.index("ngayGioHoanThanh") + 1
    worksheet.update(gspread.utils.rowcol_to_a1(row_number, status_col), [["Đã hoàn thành"]], value_input_option="RAW")
    worksheet.update(
        gspread.utils.rowcol_to_a1(row_number, completed_col),
        [[payload.ngayGioHoanThanh]],
        value_input_option="RAW",
    )
    invalidate_worksheet_cache(worksheet)
    updated_order = dict(order or {})
    updated_order["trangThai"] = "Đã hoàn thành"
    updated_order["ngayGioHoanThanh"] = payload.ngayGioHoanThanh
    log_action(request, "complete_order", "order", order_id, before=order, after=updated_order)
    return {"ok": True, "id": order_id}


@app.post("/api/orders/{order_id}/driver-notification-status")
def update_driver_notification_status(
    order_id: str, payload: DriverNotificationStatusInput, request: Request
) -> dict[str, Any]:
    worksheet = orders_worksheet()
    row_number = find_row_by_id(worksheet, order_id)
    if row_number is None:
        raise HTTPException(status_code=404, detail="Không tìm thấy đơn hàng.")
    order = row_by_id(worksheet_records(worksheet, ORDER_HEADERS), order_id) or {}
    if order_is_done(order):
        raise HTTPException(status_code=409, detail="Chỉ cập nhật trạng thái gửi tài xế cho đơn chưa hoàn thành.")
    before_order = dict(order)
    order["trangThaiGuiTaiXe"] = payload.trangThaiGuiTaiXe
    update_row_by_headers(worksheet, row_number, ORDER_HEADERS, order)
    log_action(
        request,
        "update_driver_notification_status",
        "order",
        order_id,
        before=before_order,
        after=order,
    )
    return {"ok": True, "id": order_id, "trangThaiGuiTaiXe": payload.trangThaiGuiTaiXe}


@app.post("/api/orders/{order_id}/remittance-status")
def update_order_remittance_status(
    order_id: str, payload: RemittanceStatusInput, request: Request
) -> dict[str, Any]:
    user = current_user(request)
    if not has_action(user, "manage_remittance_status"):
        raise HTTPException(status_code=403, detail="Chỉ Kế toán hoặc Admin được xác nhận nộp tiền.")
    worksheet = orders_worksheet()
    row_number = find_row_by_id(worksheet, order_id)
    if row_number is None:
        raise HTTPException(status_code=404, detail="Không tìm thấy đơn hàng.")
    order = row_by_id(worksheet_records(worksheet, ORDER_HEADERS), order_id) or {}
    if not order_is_done(order):
        raise HTTPException(status_code=409, detail="Chỉ đơn hàng đã hoàn thành mới được xác nhận nộp tiền.")
    if (
        normalize_text(order.get("congNo")) in {"co", "yes", "true", "1"}
        and payload.trangThaiNopTien != "Đã nộp tiền"
    ):
        raise HTTPException(status_code=409, detail="Đơn công nợ có trạng thái cố định là Công nợ.")
    before = dict(order)
    order["trangThaiNopTien"] = payload.trangThaiNopTien
    if payload.trangThaiNopTien == "Đã nộp tiền":
        order["ngayXacNhanNopTien"] = now_iso()
        order["nguoiXacNhanNopTien"] = current_user_display_name(request)
    else:
        order["ngayXacNhanNopTien"] = ""
        order["nguoiXacNhanNopTien"] = ""
    update_row_by_headers(worksheet, row_number, ORDER_HEADERS, order)
    log_action(request, "update_remittance_status", "order", order_id, before=before, after=order)
    return {
        "ok": True,
        "id": order_id,
        "trangThaiNopTien": order["trangThaiNopTien"],
        "ngayXacNhanNopTien": order["ngayXacNhanNopTien"],
        "nguoiXacNhanNopTien": order["nguoiXacNhanNopTien"],
    }
